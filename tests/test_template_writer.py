"""Item E3b — the XLSM template writer decides no value of its own.

Two gates:

1. `test_no_numeric_literals_in_write_paths` — an AST walk over
   `output/template_writer.py`. Structural, runs anywhere, needs no
   workbook. This is the acceptance criterion the scoped backlog asked
   be enforced "by a test (grep or AST), not by inspection".
2. Behavior tests against a SYNTHETIC workbook. The real
   `template_uw.xlsm` is a 3 MB proprietary macro file and is
   gitignored, so CI has never had it and never will. The stub below
   carries only the cells and formulas under test, which is enough to
   assert what the writer writes — and the formulas it must overwrite
   are reproduced verbatim from the real workbook so a stub that
   drifted would fail loudly rather than pass vacuously.
"""

import ast
import os
from pathlib import Path

import openpyxl
import pytest

import config as cfg
from model.debt import resolve_debt_terms
from model.waterfall import resolve_waterfall_terms
from output import template_writer
from registry import ScenarioType

WRITER_SOURCE = Path(template_writer.__file__)


# ── Gate 1: no numeric literal decides a cell value ──────────────────

def _is_cell_write(node):
    """True for `ws["A1"] = ...` and `ws.cell(...).value = ...`.

    Row and column arguments sit on the TARGET side and are deliberately
    not checked: a column index is the template's schema, not an
    underwriting assumption. What the rule forbids is the writer
    choosing the number that lands in the cell.
    """
    for target in node.targets:
        if isinstance(target, ast.Subscript):
            return True
        if (isinstance(target, ast.Attribute)
                and target.attr == "value"
                and isinstance(target.value, ast.Call)
                and isinstance(target.value.func, ast.Attribute)
                and target.value.func.attr == "cell"):
            return True
    return False


def _local_assignments(func):
    """name -> [expressions assigned to it] inside one function."""
    out = {}
    for node in ast.walk(func):
        targets = []
        if isinstance(node, ast.Assign):
            targets = [t for t in node.targets if isinstance(t, ast.Name)]
        elif (isinstance(node, (ast.AnnAssign, ast.AugAssign))
                and isinstance(node.target, ast.Name) and node.value):
            targets = [node.target]
        for target in targets:
            out.setdefault(target.id, []).append(node.value)
    return out


def _reachable_literals(expr, locals_, seen=None):
    """Numeric literals reachable from `expr`, following local names.

    A bare `ws["K180"] = value` where `value = 0.065` two lines earlier
    is the same defect as writing the literal inline, so the walk does
    not stop at the name — it follows every expression assigned to that
    name within the function. `seen` guards the self-referential case
    (`capex = ... if capex is None else capex`).
    """
    seen = set() if seen is None else seen
    found = []
    for node in ast.walk(expr):
        if (isinstance(node, ast.Constant)
                and isinstance(node.value, (int, float))
                and not isinstance(node.value, bool)):
            found.append(node.value)
        elif isinstance(node, ast.Name) and node.id not in seen:
            seen.add(node.id)
            for assigned in locals_.get(node.id, []):
                found += _reachable_literals(assigned, locals_, seen)
    return found


def _cell_write_values(tree):
    """Yield (function, assign-node, value-expr, local-assignment map)."""
    for func in ast.walk(tree):
        if not isinstance(func, (ast.FunctionDef, ast.AsyncFunctionDef)):
            continue
        locals_ = _local_assignments(func)
        for node in ast.walk(func):
            if isinstance(node, ast.Assign) and _is_cell_write(node):
                yield func, node, node.value, locals_


def test_no_numeric_literals_in_write_paths():
    """No numeric literal may appear on the value side of a cell write.

    This is the whole item in one assertion. Every number written into
    the workbook must trace to config, to a resolved terms object, to
    the run's results, or to a named structural constant at the top of
    the module — because a literal here is a second underwriting opinion
    that nothing reconciles against the memo and the .xlsx.

    If this fails on a legitimately structural number (a column index, a
    zero that means "at closing"), the fix is to NAME it at the top of
    template_writer.py with a comment saying why it is schema and not an
    assumption. The fix is never to add an exception here.
    """
    tree = ast.parse(WRITER_SOURCE.read_text())
    offenders = []

    for func, assign, value, locals_ in _cell_write_values(tree):
        for literal in _reachable_literals(value, locals_):
            offenders.append(
                f"{func.name} line {assign.lineno}: {literal!r}")

    assert not offenders, (
        "numeric literals decide a cell value in template_writer.py:\n  "
        + "\n  ".join(offenders)
        + "\nName it as a structural constant, or read it from config / "
          "the resolved terms / the run's results."
    )


def test_the_literal_gate_catches_a_literal_behind_a_variable():
    """The gate is only worth having if it fails when it should.

    A literal parked in a local and written a line later is the same
    defect as writing it inline, and the obvious implementation — check
    only the RHS of the cell-write statement — misses it silently. This
    pins that the walk follows local names, so a future simplification
    of `_reachable_literals` cannot quietly reopen the hole.
    """
    hidden = ast.parse(
        "def f(ws):\n"
        "    value = 0.065\n"
        "    ws['K180'] = value\n"
    )
    caught = [lit
              for _f, _a, value, locals_ in _cell_write_values(hidden)
              for lit in _reachable_literals(value, locals_)]
    assert caught == [0.065]

    # ...and does not fire on a value that traces to config.
    clean = ast.parse(
        "def f(ws):\n"
        "    value = cfg.DEBT_TERMS['rate']\n"
        "    ws['K180'] = value\n"
    )
    assert not [lit
                for _f, _a, value, locals_ in _cell_write_values(clean)
                for lit in _reachable_literals(value, locals_)]


def test_no_environment_reads_outside_the_template_path():
    """The GP_EQUITY_SHARE / GP_AM_FEE_RATE / GP_PROMOTE_PCT block is
    deleted, not re-defaulted (scoped-backlog rule 2).

    `UW_TEMPLATE_PATH` survives: it locates the file, it is not an
    underwriting assumption. Everything else that used to come from the
    environment now lives in config, where it can be found.
    """
    tree = ast.parse(WRITER_SOURCE.read_text())
    reads = [
        node.args[0].value
        for node in ast.walk(tree)
        if isinstance(node, ast.Call)
        and isinstance(node.func, ast.Attribute)
        and node.func.attr == "get"
        and isinstance(node.func.value, ast.Attribute)
        and node.func.value.attr == "environ"
        and node.args and isinstance(node.args[0], ast.Constant)
    ]
    assert reads == ["UW_TEMPLATE_PATH"], reads


# ── Synthetic workbook ───────────────────────────────────────────────

# Formulas copied verbatim from Self-Storage-Acquisition-Model v1.3.
# The writer must overwrite the first with a value; the second is the
# ONLY chained hurdle in v1.3 (v1.2 chained three), which is why the
# writer now writes the pref into H249/H250/H251 individually.
REAL_K174_FORMULA = "=K173+0.5%"
REAL_H252_FORMULA = "=H251"
#: v1.3's shipped hurdle literal. The writer overwrites it, and a test
#: proves it does — an 8% pref would otherwise pass by coincidence.
REAL_SHIPPED_HURDLE = 0.08
#: E128 — "Concessions Avg. Length", which v1.3's other-income formulas
#: reuse as their annualization divisor.
REAL_E128_PERIOD = 13


def build_stub_template(path):
    """Write a minimal .xlsm standing in for the proprietary template.

    Shared with `tests/test_web_runs.py`, which drives the real
    `engine.run_analysis` against it.

    Carries the v1.3 shape markers as well as the cells under test:
    `_assert_template_shape` refuses to write into a workbook whose
    labels are not v1.3's, so a stub without them would fail every
    behavior test — which is exactly the protection the guard is for.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Underwriting"
    wb.create_sheet("Summary")
    for address, label in template_writer._SHAPE_MARKERS.items():
        ws[address] = label
    ws["K174"] = REAL_K174_FORMULA
    ws["H250"] = REAL_SHIPPED_HURDLE
    ws["H251"] = REAL_SHIPPED_HURDLE
    ws["H252"] = REAL_H252_FORMULA
    ws["E128"] = REAL_E128_PERIOD
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def stub_template(tmp_path, monkeypatch):
    path = build_stub_template(tmp_path / "stub_template.xlsm")
    monkeypatch.setattr(template_writer, "TEMPLATE_PATH", str(path))
    return path


class _Unit:
    def __init__(self, size_label, count, sf, rate, climate_controlled=False):
        self.size_label = size_label
        self.count = count
        self.sf = sf
        self.rate = rate
        self.climate_controlled = climate_controlled


class _CIM:
    """Only the attributes the writer touches."""

    def __init__(self, **kwargs):
        defaults = dict(
            property_name="Test Storage", address="1 Main St", city="Abilene",
            state="TX", acreage=4.0, year_built=2005, asking_price=10_000_000,
            physical_occupancy=0.92, nrsf=50_000, other_income=25_000,
            total_units=None, mgmt_fee_pct=None, capex_estimate=0,
            unit_mix=[_Unit("10x10", 100, 100, 95.0)],
        )
        defaults.update(kwargs)
        for key, value in defaults.items():
            setattr(self, key, value)


FINANCIALS = {"adjusted_ttm_noi": {"analyst_adjusted_noi": 650_000},
              "income_analysis": {"egr": 1_200_000},
              "expense_analysis": {"lines": []}}


def _scenarios(**param_overrides):
    params = dict(cfg.SCENARIO_DEFAULTS[ScenarioType.BASE])
    params.update(param_overrides)
    return {ScenarioType.BASE: {
        "params": params,
        "exit_cap": 0.0685,
        "exit_cap_detail": {"market_cap": 0.0625},
    }}


def _generate(tmp_path, cim=None, **kwargs):
    kwargs.setdefault("financial_analysis", FINANCIALS)
    kwargs.setdefault("scenario_results", _scenarios())
    out = template_writer.generate_template(
        cim_data=cim or _CIM(),
        output_dir=str(tmp_path),
        **kwargs,
    )
    wb = openpyxl.load_workbook(out, keep_vba=True)
    try:
        yield_ws = wb["Underwriting"]
        return {c.coordinate: c.value
                for row in yield_ws.iter_rows() for c in row}
    finally:
        wb.close()


# ── Gate 2: the workbook carries the run's resolved terms ────────────

def test_debt_block_equals_resolved_debt_terms(tmp_path, stub_template):
    """The XLSM's loan terms are the loan the app actually priced.

    Pre-E3b these cells asserted 60-month term / 12-month IO / 360-month
    amortization / 6.5% against a model running 120 / 0 / 300 / 6.25%.
    """
    terms = resolve_debt_terms()
    cells = _generate(tmp_path, debt_terms=terms,
                      sources_uses={"ltv": 0.62})

    assert cells["F74"] == terms.term_years * 12 == 120
    assert cells["G74"] == terms.io_months == 0
    assert cells["H74"] == terms.amort_years * 12 == 300
    assert cells["I74"] == terms.all_in_rate() == 0.0625
    # H64 is the run's own leverage, not a hardcoded all-equity 0.
    assert cells["H65"] == 0.62
    assert cells["H66"] == 0        # we model no junior/mezz tranche


def test_debt_block_survives_an_overridden_rate(tmp_path, stub_template):
    """A per-deal override reaches the workbook. This is the test that
    would fail if the writer re-read config instead of the resolved
    object handed to it."""
    terms = resolve_debt_terms({"rate": 0.075, "amort_years": 30,
                                "term_years": 7, "io_months": 24})
    cells = _generate(tmp_path, debt_terms=terms)

    assert cells["I74"] == 0.075
    assert cells["H74"] == 360
    assert cells["F74"] == 84
    assert cells["G74"] == 24


def test_no_sources_uses_leaves_the_block_all_equity(tmp_path, stub_template):
    """The CLI path, and any deal that never sized. Terms still land, so
    flipping LTC in Excel gives the terms the app would have used."""
    cells = _generate(tmp_path, sources_uses=None)
    assert cells["H65"] == 0
    assert cells["I74"] == resolve_debt_terms().all_in_rate()


def test_waterfall_block_equals_resolved_waterfall_terms(tmp_path,
                                                         stub_template):
    """Pref, promote and GP co-invest come from WATERFALL_TERMS and
    GP_COINVEST_PCT — not from the deleted env vars, which defaulted GP
    equity to 6% against a config value of 10%."""
    terms = resolve_waterfall_terms()
    cells = _generate(tmp_path, waterfall_terms=terms, am_fee_pct=cfg.AM_FEE_PCT)

    assert cells["H60"] == terms.gp_coinvest_pct == 0.10
    assert cells["G245"] == cfg.AM_FEE_PCT == 0.01
    for row in ("I250", "I251", "I252"):
        assert cells[row] == terms.promote_split == 0.20
    # v1.3 does NOT chain H250/H251 to H249 — they ship as their own
    # literals — so the pref must reach all three. Only H252 chains,
    # and it is left alone.
    for row in ("H249", "H250", "H251"):
        assert cells[row] == terms.pref_rate == 0.08
    assert cells["H252"] == REAL_H252_FORMULA
    assert cells["C244"] == cfg.GP_ENTITY_NAME


def test_the_pref_reaches_every_unchained_hurdle_row(tmp_path,
                                                     stub_template):
    """The v1.2→v1.3 trap, pinned as a red test.

    v1.2 chained H259-H261 to H258, so one write set every hurdle. v1.3
    ships H250/H251 as independent 0.08 literals. A writer that still
    wrote only the first row would pass every test at the DEFAULT pref —
    the shipped literal and the fund's rate both being 8% — and silently
    produce a multi-hurdle workbook for any deal that moved it. So this
    asserts at a pref that is not 8%.
    """
    terms = resolve_waterfall_terms({"pref_rate": 0.09})
    cells = _generate(tmp_path, waterfall_terms=terms)
    for row in ("H249", "H250", "H251"):
        assert cells[row] == 0.09, (
            f"{row} kept the template's shipped hurdle — v1.3 does not "
            f"chain the tier rows, so each must be written")
    assert REAL_SHIPPED_HURDLE == 0.08          # the coincidence guarded


def test_waterfall_carries_a_per_deal_coinvest(tmp_path, stub_template):
    """The trap `resolve_waterfall_terms` documents: a deal edited to 25%
    co-invest must not print a 10/90 waterfall next to a 25/75 stack."""
    terms = resolve_waterfall_terms(capital_structure={"gp_coinvest_pct": 0.25})
    cells = _generate(tmp_path, waterfall_terms=terms)
    assert cells["H60"] == 0.25


def test_growth_ladder_is_scenario_driven(tmp_path, stub_template):
    """Deliberate behavior change (scoped-backlog rule 4): the flat
    0%-then-3% ladder on all six rows becomes the resolved banding —
    revenue rows on the revenue CAGR, expense rows on `exp_growth`."""
    cells = _generate(tmp_path, scenario_results=_scenarios(
        rev_cagr_yr1_3=0.04, rev_cagr_yr4_5=0.035, exp_growth=0.028))

    for row in (102, 103, 104):                      # revenue rows
        assert cells[f"C{row}"] == 0                 # year 1: in-place
        assert cells[f"D{row}"] == cells[f"E{row}"] == 0.04    # years 2-3
        assert cells[f"F{row}"] == cells[f"G{row}"] == cells[f"H{row}"] == 0.035
    for row in (105, 106, 107):                      # expense rows
        assert cells[f"C{row}"] == 0
        for col in "DEFGH":
            assert cells[f"{col}{row}"] == 0.028


def test_stabilized_vacancy_complements_the_scenario(tmp_path, stub_template):
    """Was a standing 0.10 behind a dead if/else — the one value in the
    block that matched neither the workbook's own default nor the model."""
    cells = _generate(tmp_path, scenario_results=_scenarios(stabilized_occ=0.88))
    assert cells["I139"] == 0.12
    assert cells["G139"] == pytest.approx(0.08)   # 1 - 0.92 physical


def test_credit_loss_and_bank_fees_come_from_config(tmp_path, stub_template):
    inputs = cfg.XLSM_TEMPLATE_INPUTS
    cells = _generate(tmp_path)
    assert cells["G140"] == inputs["credit_loss_in_place"]
    assert cells["I140"] == inputs["credit_loss_stabilized"]

    # v1.3 states the bank/merchant fee per UNIT, where config states it
    # as a % of EGR. Same assumption, converted — not the percentage
    # dropped into a dollar cell.
    units = 100                       # the fixture's one unit type
    egr = FINANCIALS["income_analysis"]["egr"]
    assert cells["G148"] == round(inputs["bank_fee_pct_in_place"] * egr / units, 2)
    assert cells["I148"] == round(
        inputs["bank_fee_pct_stabilized"] * egr / units, 2)


def test_bank_fee_row_is_left_alone_without_an_egr(tmp_path, stub_template,
                                                   caplog):
    """No EGR means no conversion, and a percentage written into a
    per-unit dollar cell would be off by four orders of magnitude while
    looking like a filled-in workbook."""
    with caplog.at_level("WARNING"):
        cells = _generate(tmp_path, financial_analysis={
            "adjusted_ttm_noi": {"analyst_adjusted_noi": 650_000},
            "expense_analysis": {"lines": []}})
    assert cells.get("G148") is None
    assert any("bank/merchant fee" in r.message for r in caplog.records)


def test_benchmark_band_drives_the_capital_reserve(tmp_path, stub_template):
    """The band is $/NRSF/yr; v1.3's row is $/unit/yr."""
    cells = _generate(tmp_path)
    band_psf = cfg.EXPENSE_BENCHMARKS["cap_reserve"][0]
    assert cells["I157"] == round(band_psf * 50_000 / 100, 2)
    assert cells["G157"] == 0        # no in-place reserve line in a CIM


def test_operating_expenses_are_written_per_unit_not_per_sf(tmp_path,
                                                            stub_template):
    """**The conversion that separates v1.3 from v1.2.**

    v1.2's rows multiplied by square footage; v1.3's multiply by the
    unit count (`J143 = (...)*$C$126`). Writing the old $/SF figure into
    the new cell overstates the expense by SF-per-unit — about 100x on
    this fixture — and every error-check in the workbook still reports
    "OK", so nothing downstream would catch it. Asserted against the
    dollars, not against a ratio, so it cannot pass by cancellation.
    """
    financials = dict(FINANCIALS)
    financials["expense_analysis"] = {"lines": [
        {"benchmark_key": "payroll", "cim_value": 60_000,
         "adjusted_value": 75_000}]}
    cells = _generate(tmp_path, financial_analysis=financials)

    assert cells["G144"] == 600.0     # 60,000 / 100 units
    assert cells["I144"] == 750.0     # 75,000 / 100 units
    # The $/SF number this replaced, for the record: 60,000 / 50,000 NRSF.
    assert cells["G144"] != 1.2


def test_opex_block_is_left_alone_when_the_cim_has_no_unit_count(
        tmp_path, stub_template, caplog):
    """No unit count means no per-unit basis. The block keeps the
    template's own defaults — zero-filling would print a property with
    no operating expenses, which is a worse claim than a visibly generic
    one."""
    with caplog.at_level("WARNING"):
        cells = _generate(tmp_path, cim=_CIM(unit_mix=[], total_units=None))
    assert cells.get("G144") is None
    assert any("operating-expense" in r.message for r in caplog.records)


def test_other_income_is_converted_to_the_per_unit_period_basis(
        tmp_path, stub_template):
    """v1.3: `J132 = (...)*$C$126*(12/$E$128)`.

    The annual total must be divided by units AND by the periods-per-
    year the workbook's own E128 implies, or other income lands
    (units x 12/E128) times too large. E128 is read from the workbook
    rather than assumed, which is what makes the product tie back.
    """
    cells = _generate(tmp_path, cim=_CIM(other_income=26_000))
    periods = 12 / REAL_E128_PERIOD
    assert cells["G132"] == round(26_000 / (100 * periods), 2)
    assert cells["I132"] == cells["G132"]
    # And it ties: the workbook's own formula reproduces the input.
    assert cells["G132"] * 100 * periods == pytest.approx(26_000, rel=1e-3)
    # The template's other example lines are cleared, not left standing.
    assert cells["G133"] == 0 and cells["I136"] == 0


def test_the_mgmt_fee_follows_the_resolved_target_not_the_band(
        tmp_path, stub_template):
    """This assertion used to read
    `cfg.EXPENSE_BENCHMARKS["mgmt_fee_pct"][1]`, and it still PASSES that
    way today by coincidence — `MGMT_FEE_TARGET_PCT` is now 6%, which is
    the band's high end. Same number, two routes, and the test could not
    tell them apart.

    It matters because the routes diverge the moment a deal sets a
    per-deal target: the memo and the .xlsx would move and this workbook
    — a real deliverable — would keep writing 6%, two output files
    asserting different management fees on the same deal with nothing on
    screen saying so. So the target is moved away from the band high here
    to prove which one the cell actually follows.
    """
    cells = _generate(tmp_path, mgmt_fee_target_pct=0.02)
    assert cells["G150"] == 0.02
    assert cells["I150"] == 0.02
    assert cells["G150"] != cfg.EXPENSE_BENCHMARKS["mgmt_fee_pct"][1]

    # Unset, it falls back to the config default (which IS the band high
    # today — asserted against the target, not the band).
    cells = _generate(tmp_path)
    assert cells["G150"] == cfg.MGMT_FEE_TARGET_PCT

    # A CIM that states its own rate still wins over either.
    cells = _generate(tmp_path, cim=_CIM(mgmt_fee_pct=0.045),
                      mgmt_fee_target_pct=0.02)
    assert cells["G150"] == 0.045


def test_capex_timing_comes_from_config(tmp_path, stub_template):
    inputs = cfg.XLSM_TEMPLATE_INPUTS
    cells = _generate(tmp_path, cim=_CIM(capex_estimate=250_000), capex=250_000)
    assert cells["K25"] == 250_000
    assert cells["E25"] == inputs["capex_start_month"]
    assert cells["F25"] == inputs["capex_duration_months"]


def test_missing_occupancy_warns_and_uses_the_config_assumption(
        tmp_path, stub_template, caplog):
    """The audit's complaint was the silence, not the 0.90."""
    with caplog.at_level("WARNING"):
        cells = _generate(tmp_path, cim=_CIM(physical_occupancy=None))

    assumed = cfg.XLSM_TEMPLATE_INPUTS["assumed_physical_occupancy"]
    assert cells["G139"] == pytest.approx(1 - assumed)
    assert any("Physical occupancy missing" in r.message
               for r in caplog.records)


def test_thin_deal_does_not_invent_a_cap_rate(tmp_path, stub_template, caplog):
    """The 6.5% entry-cap fallback is gone. With no scenario, no NOI and
    no price there is nothing to write, so K181 keeps the workbook's own
    entry+50bp formula — which is only safe because there is no resolved
    exit cap for it to contradict."""
    thin = _CIM(asking_price=None)
    with caplog.at_level("WARNING"):
        cells = _generate(tmp_path, cim=thin, scenario_results=None,
                          financial_analysis={})

    assert cells["K173"] is None                     # never written
    assert cells["K174"] == REAL_K174_FORMULA        # formula untouched
    assert any("no entry cap" in r.message for r in caplog.records)


def test_resolved_exit_cap_overwrites_the_entry_plus_50bp_formula(
        tmp_path, stub_template):
    cells = _generate(tmp_path)
    assert cells["K173"] == 0.0625
    assert cells["K174"] == 0.0685


def test_defaults_resolve_without_any_terms_passed(tmp_path, stub_template):
    """`run.py` passes no terms at all. Config must be the whole resolved
    set on that path, not a crash and not a literal."""
    cells = _generate(tmp_path)
    assert cells["I74"] == cfg.DEBT_TERMS["rate"]
    assert cells["H249"] == cfg.PREF_RATE_LEVERED
    assert cells["H60"] == cfg.GP_COINVEST_PCT
    assert cells["G245"] == cfg.AM_FEE_PCT


# ── The divergence disclosures (settled 2026-08-10) ──────────────────

def test_divergence_disclosures_land_in_the_workbook(tmp_path,
                                                     stub_template):
    """Both structural divergences are stamped INTO the workbook, not
    only into the module docstring — the docstring is the one place the
    analyst reading the .xlsm will never look."""
    cells = _generate(tmp_path)
    assert (cells[template_writer._DISCLOSURE_PREF_CELL]
            == template_writer._PREF_DISCLOSURE)
    assert (cells[template_writer._DISCLOSURE_AM_FEE_CELL]
            == template_writer._AM_FEE_DISCLOSURE)
    # And the TRUE rate stands beside the disclosure: the operator
    # rejected the gross-up (2026-08-10), reaffirming the recorded
    # stance — dollars that tie by printing a rate the fund does not
    # charge trade a visible discrepancy for a hidden one.
    assert cells["G245"] == cfg.AM_FEE_PCT


def _disclosure_texts(ws):
    return [c.value for row in ws.iter_rows() for c in row
            if isinstance(c.value, str)
            and c.value in (template_writer._PREF_DISCLOSURE,
                            template_writer._AM_FEE_DISCLOSURE)]


def _write_into(tmp_path, prepare):
    """Run the disclosure writer over a sheet `prepare` has shaped."""
    wb = openpyxl.Workbook()
    ws = wb.active
    prepare(ws)
    template_writer._write_divergence_disclosures(ws)
    return ws


def test_disclosures_survive_a_merged_band_at_the_preferred_rows(tmp_path):
    """The failure the parity test alone could not prevent, now harmless.

    A merged label band covering B263/B264 makes those cells read empty
    while B264 is a read-only MergedCell — so the old writer raised
    AttributeError mid-run and NO workbook was produced. The disclosure
    now moves below the band instead.
    """
    ws = _write_into(tmp_path, lambda w: w.merge_cells("B263:D264"))
    assert _disclosure_texts(ws) == [template_writer._PREF_DISCLOSURE,
                                     template_writer._AM_FEE_DISCLOSURE]
    # And it did NOT write into the merged band.
    assert ws["B263"].value is None and ws["B264"].value is None


def test_disclosures_never_overwrite_an_occupied_cell(tmp_path):
    """The other half: a label already sitting where we guessed."""
    def occupy(w):
        w["B263"] = "Waterfall Notes"
        w["B264"] = "=SUM(H259:H261)"

    ws = _write_into(tmp_path, occupy)
    assert ws["B263"].value == "Waterfall Notes"
    assert ws["B264"].value == "=SUM(H259:H261)"
    assert _disclosure_texts(ws) == [template_writer._PREF_DISCLOSURE,
                                     template_writer._AM_FEE_DISCLOSURE]


def test_no_room_warns_and_writes_nothing_rather_than_corrupting(
        tmp_path, caplog):
    """The last resort must still produce a workbook. A missing note with
    a loud warning beats a destroyed deliverable — and beats a crash that
    leaves the analyst with no file at all."""
    def fill(w):
        for offset in range(template_writer._DISCLOSURE_SEARCH_ROWS):
            w.cell(row=template_writer._DISCLOSURE_FIRST_ROW + offset,
                   column=template_writer._DISCLOSURE_COL).value = "taken"

    with caplog.at_level("WARNING"):
        ws = _write_into(tmp_path, fill)
    assert _disclosure_texts(ws) == []
    assert any("divergence disclosures" in r.message for r in caplog.records)


def test_disclosures_state_mechanism_and_direction():
    """The disclosure is only a disclosure while it names the mechanism
    and the direction. A future rewording may shorten it; it may not
    hollow it."""
    pref = template_writer._PREF_DISCLOSURE
    fee = template_writer._AM_FEE_DISCLOSURE
    assert "IRR hurdle" in pref and "accru" in pref
    assert "LP equity" in fee and "invested equity" in fee
    assert "light" in fee            # the direction of the gap
    assert "true rate" in fee        # what G254 actually carries


# ── The real template, when it happens to be present ─────────────────

def _is_merged(ws, coordinate: str) -> bool:
    """True if `coordinate` falls inside ANY merged range.

    Covers both halves of the failure: the anchor (a writable Cell that
    would overwrite a multi-cell label) and every other member (a
    MergedCell whose `.value` is read-only, so the write raises).
    """
    cell = ws[coordinate]
    return any(cell.row >= rng.min_row and cell.row <= rng.max_row
               and cell.column >= rng.min_col and cell.column <= rng.max_col
               for rng in ws.merged_cells.ranges)


def test_the_merge_detector_catches_both_halves_of_the_failure():
    """The tripwire is only worth having if it fails when it should, and
    `.value is None` provably does not: a merged band reads empty from
    every cell it covers. Pins the detector against the two shapes —
    anchor and member — that an emptiness check waves through."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("B263:D264")
    try:
        assert ws["B263"].value is None and ws["B264"].value is None
        assert _is_merged(ws, "B263")      # anchor: silent overwrite
        assert _is_merged(ws, "B264")      # member: the write raises
        assert not _is_merged(ws, "B270")  # a genuinely free cell
    finally:
        wb.close()



# The path the WRITER resolves, not a second guess at it: the module
# honors UW_TEMPLATE_PATH (render.yaml sets it to /data), and a parity
# test hardcoded to the project root would skip — reporting "no template
# here" — on exactly the machine that has one somewhere else. Read at
# import, before any fixture monkeypatches it.
REAL_TEMPLATE = Path(template_writer.TEMPLATE_PATH)


@pytest.mark.skipif(not REAL_TEMPLATE.exists(),
                    reason="proprietary template is gitignored; absent in CI")
def test_real_template_still_has_the_cells_the_stub_claims(monkeypatch):
    """Guards the stub against drift. If the real workbook's formulas or
    row mapping ever move, the synthetic tests above are asserting
    against fiction — and this is the only test that can notice."""
    wb = openpyxl.load_workbook(REAL_TEMPLATE, keep_vba=True)
    try:
        ws = wb["Underwriting"]
        assert ws["K174"].value == REAL_K174_FORMULA
        # The v1.3 shape markers the writer refuses to proceed without.
        for address, label in template_writer._SHAPE_MARKERS.items():
            assert ws[address].value == label, (
                f"{address} reads {ws[address].value!r} — this template "
                f"is not the v1.3 the cell map was calibrated against")
        # Only H252 chains; H250/H251 are their own literals, which is
        # why the writer writes the pref into every hurdle row.
        assert ws["H252"].value == REAL_H252_FORMULA
        assert ws["H250"].value == REAL_SHIPPED_HURDLE
        assert ws["H251"].value == REAL_SHIPPED_HURDLE
        # Growth ladder row labels, so the revenue/expense split is real.
        assert ws["B102"].value == "In-Place Rent"
        assert ws["B104"].value == "Other Income"
        assert ws["B105"].value == "OpEx (Excl. Taxes)"
        assert ws["B107"].value == "CapEx"
        # H65 is loan / Total Uses, which is what makes sources_uses["ltv"]
        # the right thing to write there.
        assert ws["K65"].value == "=H65*$K$56"
        # The AM fee's base really is LP equity, not invested equity —
        # the residual mismatch the module docstring stamps.
        assert ws["H245"].value == (
            '=IF(G244="% of EGR",G245*K141/12,K61*G245/12)')
        assert ws["K61"].value == "=($K$56-$K$67)*H61"
        # **The unit basis.** These formulas are the whole reason the
        # writer converts to per-unit dollars: C126 is the unit count,
        # E128 the other-income period divisor. If either changes shape
        # again, every expense in the workbook is wrong by a factor
        # nothing else would report.
        assert ws["J143"].value == "=((1-$E$126)*G143+$E$126*I143)*$C$126"
        assert ws["J132"].value == (
            "=((1-$E$126)*G132+$E$126*I132)*$C$126*(12/$E$128)")
        assert ws["C126"].value == "=SUM(C112:C125)"
        assert isinstance(ws["E128"].value, (int, float))
        # The PREFERRED disclosure rows should be blank and unmerged in
        # the real workbook. This is now a drift report, not a load-
        # bearing guarantee: `_free_disclosure_cells` verifies its target
        # at write time, so a wrong guess costs a lower row rather than a
        # corrupted or missing workbook. Still asserted, because the
        # first machine to open the real file is the only one that can
        # tell us the constant points somewhere sensible.
        for cell in (template_writer._DISCLOSURE_PREF_CELL,
                     template_writer._DISCLOSURE_AM_FEE_CELL):
            assert ws[cell].value is None, (
                f"{cell} is occupied in the real template — the writer "
                f"will skip past it; move _DISCLOSURE_FIRST_ROW")
            # Emptiness alone would NOT be enough (review finding): a
            # cell inside a merged band reads None whatever the band
            # displays, so this checks the shape the value cannot show.
            assert not _is_merged(ws, cell), (
                f"{cell} participates in a merged range in the real "
                f"template — the writer will skip past it; move "
                f"_DISCLOSURE_FIRST_ROW to a free area")
    finally:
        wb.close()
