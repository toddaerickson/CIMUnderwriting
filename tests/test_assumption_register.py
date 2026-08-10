"""Item T Category 6 — the assumption register.

Category 4's lesson, which this file inherits: a register that records
perfectly and renders nowhere reproduces every number in the pipeline and
delivers none of the item. So the tests that matter here assert a SURFACE
or a WIRE, and each names the deliberate break that turns it red.

The two completeness guards are the other half. A register is a claim
about what it contains, and a claim nobody enforces stops being true the
first month somebody adds a config key — so membership FAILS BY DEFAULT
and has to be argued into `NOT_IN_REGISTER` with a reason.
"""

import pytest

import config as cfg
from analysis import assumptions as A
from tests.test_characterization import stabilized_deal
from tests.test_config_single_source import _memo_text, _run


# ── Completeness guards ─────────────────────────────────────────────

def test_every_settings_editable_key_is_in_the_register_or_declared_out():
    """The guard the whole item rests on.

    `override_key_registry()` is derived LIVE from config.py, so a new
    editable constant appears there the moment it is declared — and then
    fails here until somebody either registers it or writes down why it
    is exempt. That default matters more than the current pass: a
    completeness claim maintained by memory is one nobody can rely on.
    """
    from webapp.forms import override_key_registry

    have = {row.key for row in A.collect()}
    missing = []
    for key in override_key_registry():
        if key in have:
            continue
        if any(key.startswith(prefix) for prefix in A.NOT_IN_REGISTER):
            continue
        missing.append(key)
    assert not missing, (
        "settings-editable keys that reach no register row and are not "
        f"declared in NOT_IN_REGISTER: {sorted(missing)}")


def test_the_exempt_keys_are_exempt_for_a_stated_reason():
    """`NOT_IN_REGISTER` is a decision on the record, not a skip list.

    An entry with an empty reason is how an exemption stops being a
    decision and becomes a place to hide keys.
    """
    assert A.NOT_IN_REGISTER, "an empty exemption table means the guard above proves nothing"
    for prefix, reason in A.NOT_IN_REGISTER.items():
        assert prefix.endswith("."), f"{prefix} must be a dotted-key prefix"
        assert len(reason.split()) >= 8, f"{prefix} needs a real reason, got {reason!r}"


def test_the_market_cap_table_is_exempt_but_its_resolved_anchor_is_not():
    """The exemption is only honest because the number that MOVED is
    still reported. Exempting the table and reporting nothing would be
    the omission the exemption pretends to justify."""
    rows = {r.key: r for r in A.collect(
        market_cap={"market_cap": 0.0625, "source": "table",
                    "asset_class": "Self-Storage", "age_band": "2000s"})}
    assert not any(k.startswith("MARKET_CAP_RATES.")
                   and k != "MARKET_CAP_RATES.resolved" for k in rows)
    anchor = rows["MARKET_CAP_RATES.resolved"]
    assert anchor.value == 0.0625
    assert "Self-Storage" in anchor.detail and "2000s" in anchor.detail


def test_the_levered_constants_are_registered_though_no_guard_demands_them():
    """`DEBT_TERMS` and friends are per-deal only (item E3b), so
    `override_key_registry()` does not list them and the guard above is
    silent about them. Silence is not permission to omit a number that
    moves every levered figure in the memo."""
    have = {row.key for row in A.collect()}
    for key in cfg.DEBT_TERMS:
        assert f"DEBT_TERMS.{key}" in have, key
    for key in cfg.WATERFALL_TERMS:
        if key == "gp_coinvest_pct":
            continue
        assert f"WATERFALL_TERMS.{key}" in have, key
    for key in ("AM_FEE_PCT", "AM_FEE_BASE", "DEFAULT_HOLD_YEARS",
                "SOLVER_TARGET_IRR", "SOLVER_TARGET_LP_NET_IRR",
                "MGMT_FEE_TARGET_PCT", "IRR_STRONG_THRESHOLD"):
        assert key in have, key
    for name in ("SOLVER_BOUNDS", "SENSITIVITY_GRID", "TRANSACTION_COSTS",
                 "EXPENSE_RATIO", "POPULATION_TIERS", "OCCUPANCY_TIERS"):
        for key in getattr(cfg, name):
            assert f"{name}.{key}" in have, f"{name}.{key}"


def test_the_exit_noi_convention_is_registered_though_no_guard_demands_it():
    """`EXIT_NOI_CONVENTION` is a top-level constant and deliberately not
    settings-editable, so `override_key_registry()` never lists it and
    the membership guard above is silent about it. Same rule as the
    levered constants: silence is not permission to omit the convention
    that prices every exit in the run — static DCF and value-add both."""
    rows = {row.key: row for row in A.collect()}
    assert "EXIT_NOI_CONVENTION" in rows
    row = rows["EXIT_NOI_CONVENTION"]
    assert row.value == cfg.EXIT_NOI_CONVENTION == "trailing"
    assert row.provenance == A.CONFIG


def test_every_numeric_cim_form_field_is_registered():
    """A new box on the assumptions page moves an output the moment an
    analyst types in it. If `CIM_FIELDS` does not know about it, the
    appendix reports a run it did not describe."""
    from webapp.forms import CIM_FLOAT_FIELDS, CIM_INT_FIELDS, CIM_PCT_FIELDS

    declared = {field for field, _label, _unit in A.CIM_FIELDS}
    on_form = set(CIM_INT_FIELDS + CIM_FLOAT_FIELDS + CIM_PCT_FIELDS)
    assert not (on_form - declared), (
        f"analyst-editable CIM inputs missing from the register: "
        f"{sorted(on_form - declared)}")
    assert not (declared - on_form), (
        f"registered CIM fields with no input box: {sorted(declared - on_form)}")


def test_the_guard_fails_when_a_registered_key_disappears(monkeypatch):
    """The mutation that proves the guard above is load-bearing.

    Without this, a guard that silently matched everything would read as
    a passing completeness proof — which is precisely the failure
    `test_no_module_still_sizes_a_property_at_one_square_foot`
    overclaimed its way into during Category 4.
    """
    from webapp.forms import override_key_registry

    real = A.collect

    def crippled(**kwargs):
        return [r for r in real(**kwargs) if r.key != "GATES.min_irr_5yr"]

    monkeypatch.setattr(A, "collect", crippled)
    have = {row.key for row in A.collect()}
    missing = [k for k in override_key_registry()
               if k not in have
               and not any(k.startswith(p) for p in A.NOT_IN_REGISTER)]
    assert missing == ["GATES.min_irr_5yr"]


# ── Provenance ──────────────────────────────────────────────────────

def test_a_settings_row_is_labelled_settings_and_carries_what_it_displaced():
    """The defect that opened this category: an overridden gate rendered
    identically to the shipped one."""
    rows = {r.key: r for r in A.collect(
        config_deltas={"GATES.min_irr_5yr": 0.08},
        config_defaults={"GATES.min_irr_5yr": 0.10})}
    row = rows["GATES.min_irr_5yr"]
    assert row.provenance == A.SETTINGS
    assert row.was == 0.10
    assert row.chosen is True

    untouched = rows["GATES.min_yield_on_cost"]
    assert untouched.provenance == A.CONFIG
    assert untouched.was is None


def test_a_deal_override_beats_a_settings_row_and_only_one_row_survives():
    """Precedence is the model's own. Two rows for one assumption is how
    a reader ends up auditing a number the engine never used."""
    rows = [r for r in A.collect(
        deal_overrides={"hold_years": 7},
        config_deltas={"GATES.min_irr_5yr": 0.08},
        config_defaults={"GATES.min_irr_5yr": 0.10},
        hold_years=7) if r.key == "DEFAULT_HOLD_YEARS"]
    assert len(rows) == 1
    assert rows[0].provenance == A.DEAL
    assert rows[0].value == 7
    assert rows[0].was == cfg.DEFAULT_HOLD_YEARS


def test_a_per_key_deal_override_does_not_relabel_its_neighbour():
    """Transaction costs merge PER PARAMETER rather than replacing
    wholesale, so overriding one must not stamp the other as
    analyst-entered — the same key-level rule `webapp.services` applies
    when it pops superseded rows out of the applied stamp."""
    rows = {r.key: r for r in A.collect(
        deal_overrides={"transaction_costs": {"disposition_cost_pct": 0.02}},
        transaction_costs={"disposition_cost_pct": 0.02})}
    assert rows["TRANSACTION_COSTS.disposition_cost_pct"].provenance == A.DEAL
    assert rows["TRANSACTION_COSTS.acquisition_closing_pct"].provenance == A.CONFIG


def test_an_analyst_corrected_cim_number_shows_what_the_cim_said(mock_cim_data):
    """The row that makes a correction auditable rather than merely
    disclosed. Without `was`, "NRSF 60,000 — entered for this deal" hides
    that the broker's own number was different."""
    mock_cim_data.nrsf = 60_000
    rows = {r.key: r for r in A.collect(
        cim_data=mock_cim_data, cim_snapshot={"nrsf": 58_400})}
    row = rows["cim.nrsf"]
    assert row.provenance == A.DEAL
    assert row.value == 60_000 and row.was == 58_400
    # An untouched field stays the CIM's, not the analyst's.
    assert rows["cim.ttm_noi"].provenance == A.CIM


def test_an_absent_cim_field_contributes_no_row(mock_cim_data):
    """It moved no output. A register of numbers the run did not use
    teaches a reader to skim the ones it did — and a fallback that stood
    in for it is already a row of its own."""
    mock_cim_data.market_rent_psf = None
    keys = {r.key for r in A.collect(cim_data=mock_cim_data)}
    assert "cim.market_rent_psf" not in keys


def test_a_fill_becomes_the_fallback_row_for_the_field_it_filled():
    """The two registers join here: exactly one row per deal input, whose
    provenance is `cim`, `deal` or `fallback`."""
    from analysis import fills

    log = fills.to_dicts([fills.Fill(
        field="market_rent_psf", value_used=1.25,
        source_key=fills.MARKET_RENT_ABSENT, unit=fills.UNIT_PSF_MO,
        label="No market rent stated; in-place rent used.")])
    rows = {r.key: r for r in A.collect(fill_log=log)}
    row = rows["fill.market_rent_psf"]
    assert row.provenance == A.FALLBACK
    assert A.format_value(row) == "$1.25/SF/mo"
    assert row.chosen is True


def test_an_analyst_entered_exit_cap_is_a_deal_row_a_table_lookup_is_not():
    """`resolve_market_cap` already publishes which it was; this reads
    that flag rather than re-deciding it."""
    analyst = {r.key: r for r in A.collect(
        market_cap={"market_cap": 0.055, "source": "analyst"})}
    assert analyst["MARKET_CAP_RATES.resolved"].provenance == A.DEAL
    table = {r.key: r for r in A.collect(
        market_cap={"market_cap": 0.06, "source": "table",
                    "asset_class": "Self-Storage", "age_band": "2010s"})}
    assert table["MARKET_CAP_RATES.resolved"].provenance == A.CONFIG


# ── Values follow the model, not a second copy of it ────────────────

def test_the_expense_bands_are_the_regional_ones_the_run_charged_against(
        mock_cim_data):
    """`financials.py` underwrites against `get_regional_benchmarks(state)`.
    Reporting the national table would print a band the run never used —
    the same class of defect as Category 4's expense-ratio fill logging a
    share the projection had clamped away.
    """
    national = cfg.EXPENSE_BENCHMARKS["payroll"]
    regional = cfg.get_regional_benchmarks("TX")["payroll"]
    assert regional != national, (
        "fixture no longer proves anything — TX stopped adjusting payroll")

    rows = {r.key: r for r in A.collect(cim_data=mock_cim_data)}
    assert tuple(rows["EXPENSE_BENCHMARKS.payroll"].value) == tuple(regional)
    assert "TX" in rows["EXPENSE_BENCHMARKS.payroll"].detail


def test_a_zero_solver_target_is_reported_not_swallowed():
    """The truthiness trap item T Category 3 already had to fix once, in
    the engine. A register with its own `if target:` would reintroduce it
    one file away from where it was closed."""
    rows = {r.key: r for r in A.collect(
        deal_overrides={"solver_target_irr": 0.0}, solver_target_irr=0.0)}
    row = rows["SOLVER_TARGET_IRR"]
    assert row.value == 0.0
    assert row.provenance == A.DEAL


def test_the_registered_value_follows_a_moved_config_default(monkeypatch):
    """Mutation proof that the register READS config rather than
    restating it — the failure mode item T exists to close, appearing in
    the module that reports on it."""
    monkeypatch.setitem(cfg.GATES, "min_irr_5yr", 0.13)
    rows = {r.key: r for r in A.collect()}
    assert rows["GATES.min_irr_5yr"].value == 0.13


def test_the_scenario_rows_follow_a_per_deal_scenario_set():
    """The form stores sections keyed by `ScenarioType.value`; config
    keys them by the enum. Both spellings reach `collect`, and guessing
    wrong reports the config default for a run that used the analyst's."""
    from registry import ScenarioType

    custom = {"base": {"yr1_noi_bump": 0.09}}
    rows = {r.key: r for r in A.collect(scenarios=custom)}
    row = rows["SCENARIO_DEFAULTS.base.yr1_noi_bump"]
    assert row.value == 0.09 and row.provenance == A.DEAL
    assert row.was == cfg.SCENARIO_DEFAULTS[ScenarioType.BASE]["yr1_noi_bump"]
    # A parameter the analyst did not touch stays the model's.
    assert rows["SCENARIO_DEFAULTS.base.exp_growth"].provenance == A.CONFIG


# ── The register's own contract ─────────────────────────────────────

def test_the_provenance_vocabulary_cannot_drift_from_its_labels():
    """`PROVENANCE_KEYS` is derived from `PROVENANCE_LABELS` for the same
    reason `fills.SOURCE_KEYS` is: two structures listing one vocabulary
    is the duplicated-constant defect, in the module that closes it."""
    assert A.PROVENANCE_KEYS == tuple(A.PROVENANCE_LABELS)
    assert set(A.CHOSEN) <= set(A.PROVENANCE_KEYS)
    assert A.CONFIG not in A.CHOSEN and A.CIM not in A.CHOSEN


def test_every_row_a_real_run_emits_carries_a_declared_provenance(
        mock_cim_data):
    rows = A.collect(cim_data=mock_cim_data)
    assert rows
    for row in rows:
        assert row.provenance in A.PROVENANCE_KEYS, row.key
        assert row.group in A.GROUP_ORDER, row.key
        assert row.label, row.key


def test_no_two_rows_claim_the_same_assumption(mock_cim_data):
    keys = [r.key for r in A.collect(
        cim_data=mock_cim_data,
        market_cap={"market_cap": 0.06, "source": "table"},
        deal_overrides={"hold_years": 7}, hold_years=7)]
    assert len(keys) == len(set(keys))


def test_the_register_round_trips_and_tolerates_a_row_from_an_older_run():
    """Adding a column must not 500 the results page for every deal
    already in the database — the contract `fills.from_dicts` holds, held
    here for the same reason."""
    rows = A.collect()
    back = A.from_dicts(A.to_dicts(rows))
    assert len(back) == len(rows)
    assert back[0] == rows[0]

    older = A.from_dicts([{"key": "GATES.min_irr_5yr", "provenance": "config",
                           "value": 0.1}])
    assert len(older) == 1 and older[0].label == "GATES.min_irr_5yr"
    # A row with no provenance is not a row; it is a shape that happens
    # to have keys.
    assert A.from_dicts([{"key": "x"}]) == []


def test_a_band_renders_as_one_value_not_two_columns():
    band = A.Assumption(key="EXPENSE_BENCHMARKS.insurance", label="Insurance",
                        group=A.G_EXPENSES, value=(0.12, 0.25),
                        provenance=A.CONFIG, unit=A.UNIT_PSF_YR)
    assert A.format_value(band) == "$0.12/SF/yr – $0.25/SF/yr"
    # A stored band comes back from JSON as a list and must render the same.
    assert A.format_value(A.from_dicts(A.to_dicts([band]))[0]) == \
        "$0.12/SF/yr – $0.25/SF/yr"


def test_the_summary_counts_every_provenance_including_the_empty_ones():
    """A count silently absent reads as a count of none anyway, and only
    one of those two is true on purpose."""
    counts = A.summarize(A.collect(
        config_deltas={"GATES.min_irr_5yr": 0.08},
        config_defaults={"GATES.min_irr_5yr": 0.10}))
    assert set(A.PROVENANCE_KEYS) <= set(counts)
    assert counts[A.SETTINGS] == 1
    assert counts[A.DEAL] == 0
    assert counts["chosen"] == 1
    assert counts["total"] == sum(counts[k] for k in A.PROVENANCE_KEYS)


@pytest.mark.parametrize("value,unit,expected", [
    (0.0625, A.UNIT_PCT, "6.25%"),      # a coupon `.1%` would round to 6.2%
    (0.0025, A.UNIT_PCT, "0.25%"),      # a 25bp grid step, likewise
    (0.75, A.UNIT_PCT, "75.0%"),        # unchanged from before this item
    (0.06, A.UNIT_PCT, "6.0%"),         # the value Category 4's test pins
])
def test_the_percent_unit_stopped_restating_the_number_it_discloses(
        value, unit, expected):
    """A disclosure that rounds its own subject is not a disclosure. The
    two unchanged cases are the point: precision was ADDED where `.1%`
    lost information and nowhere else, so no existing surface moved."""
    from analysis.fills import format_number
    assert format_number(value, unit) == expected


# ── The wires: a register that renders nowhere is not the item ──────

def test_the_register_reaches_the_memo_appendix(tmp_path):
    """MUTATION: drop `assumption_register=` from `generate_memo`'s call
    site in `engine.py` and Appendix B vanishes while every number in the
    document stays identical.

    This is Category 4's lesson restated: the feature IS a report, so a
    register that records perfectly and renders nowhere reproduces the
    whole pipeline and delivers none of the item. Every assertion here is
    about the DOCUMENT.
    """
    result = _run(stabilized_deal(), tmp_path)

    assert result.assumption_register, "the engine assembled no register"
    text = _memo_text(result.memo_path)
    assert "Appendix B. Assumption Register" in text
    assert "B.1 Assumptions not taken from the model defaults" in text
    assert "B.2 Full register" in text
    # Not a sample of it — the whole register, which is the claim the
    # appendix makes in its own opening sentence.
    for row in A.from_dicts(result.assumption_register):
        assert row.label in text, f"{row.key} registered but never rendered"


def test_a_settings_override_reaches_the_appendix_with_what_it_displaced(
        tmp_path, monkeypatch):
    """The end-to-end version of this category's opening defect: an
    overridden gate used to render as a bare number.

    The delta is passed the way the worker passes it, so this exercises
    the engine parameter and the appendix together. MUTATION: drop
    `config_deltas`/`config_defaults` from either the engine signature or
    the `collect()` call and the row silently reverts to "model default".
    """
    monkeypatch.setitem(cfg.GATES, "min_irr_5yr", 0.08)
    result = _run(stabilized_deal(), tmp_path,
                  config_deltas={"GATES.min_irr_5yr": 0.08},
                  config_defaults={"GATES.min_irr_5yr": 0.10})

    row = next(r for r in A.from_dicts(result.assumption_register)
               if r.key == "GATES.min_irr_5yr")
    assert row.provenance == A.SETTINGS and row.was == 0.10

    text = _memo_text(result.memo_path)
    assert "settings override" in text
    # B.1 exists to make this readable without scanning 140 rows: the
    # value used, and the shipped value it replaced, on one line.
    assert "8.0%" in text and "10.0%" in text


def test_the_register_reaches_the_workbook_inputs_tab(tmp_path):
    """MUTATION: drop `assumption_register=` from `generate_excel`'s call
    site, or the `_write_assumption_register` line in `_build_inputs_tab`."""
    from openpyxl import load_workbook

    result = _run(stabilized_deal(), tmp_path)
    ws = load_workbook(result.excel_path)["Inputs"]
    cells = {c.value for row in ws.iter_rows() for c in row if c.value}

    assert "Assumption Register" in cells
    rows = A.from_dicts(result.assumption_register)
    for row in rows:
        assert row.label in cells, f"{row.key} missing from the workbook"
    # The provenance column is the point of the block; a register of
    # values without it is the Inputs tab it already had.
    assert A.PROVENANCE_LABELS[A.CONFIG] in cells


def test_a_cli_register_never_claims_a_settings_or_deal_override(tmp_path):
    """The CLI has no ConfigOverride table and no assumptions page, so
    those two provenances are unreachable there — and the register must
    say so rather than mislabel a default.

    Also the Category 4 audit finding in its own right: the CLI is a
    SEPARATE orchestration, so every wire runs twice or one entry point
    ships a document that hides what the other discloses. That defect
    shipped once already, with the fill log.

    MUTATION: drop `assumption_register=ctx.assumption_register` from
    either writer call in `run.stage_output`.
    """
    import run as cli
    from context import AnalysisContext

    class _NoComps:
        def query_expense_benchmarks(self, **kw):
            return None

        def query_rent_comps(self, **kw):
            return None

        def save_analysis(self, **kw):
            return None

    ctx = AnalysisContext(pdf_path=str(tmp_path / "deal.pdf"))
    ctx.cim_data = stabilized_deal()
    comp_db = _NoComps()
    cli.stage_analyze(ctx, comp_db)
    cli.stage_valuate(ctx)
    cli.stage_gates_and_risks(ctx)

    assert ctx.assumption_register, "the CLI never assembled a register"
    rows = A.from_dicts(ctx.assumption_register)
    claimed = {r.provenance for r in rows}
    assert A.SETTINGS not in claimed and A.DEAL not in claimed
    assert {A.CONFIG, A.CIM} <= claimed

    cli.stage_output(ctx, comp_db)
    text = _memo_text(ctx.memo_path)
    assert "Appendix B. Assumption Register" in text
    for row in rows:
        assert row.label in text


@pytest.mark.django_db
def test_the_worker_actually_persists_the_register(tmp_path):
    """The one wire from `run_analysis` to the database.

    Written the way Category 4's equivalent had to be rewritten after an
    audit: a test that hand-builds `result_json` passes against a worker
    that drops the key on the floor, and that is not hypothetical — item
    E3a shipped exactly that defect with the levered lens. So this runs
    the REAL worker and reads the REAL row.

    MUTATION: delete `"assumption_register": result.assumption_register`
    from the payload in `webapp/services.py`.
    """
    from webapp.models import AnalysisRun, Deal
    from webapp.services import _analysis_worker, cim_to_dict

    deal = Deal.objects.create(
        deal_id="persist-register", property_name="Persist Register",
        deal_dir=str(tmp_path), cim_json=cim_to_dict(stabilized_deal()))
    run = AnalysisRun.objects.create(deal=deal, status="running")
    _analysis_worker(run.pk)

    run.refresh_from_db()
    assert run.status == "done", run.error
    stored = (run.result_json or {}).get("assumption_register")
    assert stored, "the run recorded no register — the key never landed"
    assert all(isinstance(row, dict) for row in stored)
    rows = A.from_dicts(stored)
    assert len(rows) == len(stored), "a stored row failed to round-trip"
    assert {r.provenance for r in rows} <= set(A.PROVENANCE_KEYS)


@pytest.mark.django_db
def test_a_stored_settings_row_is_labelled_settings_by_the_real_worker(
        tmp_path):
    """The provenance wire end to end, through the code that resolves it.

    `config_defaults` has to be captured BEFORE `_patched_config` mutates
    the dicts — inside the lock the live value IS the override, so a
    register that asked config what it used to be would report
    "settings override, was 8%, now 8%". This is the test that fails if
    that capture moves.
    """
    import datetime as dt

    from webapp.models import AnalysisRun, ConfigOverride, Deal
    from webapp.services import _analysis_worker, cim_to_dict

    ConfigOverride.objects.create(key="GATES.min_irr_5yr", value=0.08,
                                 effective_date=dt.date(2026, 1, 1))
    deal = Deal.objects.create(
        deal_id="settings-provenance", property_name="Settings Provenance",
        deal_dir=str(tmp_path), cim_json=cim_to_dict(stabilized_deal()))
    run = AnalysisRun.objects.create(deal=deal, status="running")
    _analysis_worker(run.pk)

    run.refresh_from_db()
    assert run.status == "done", run.error
    row = next(r for r in A.from_dicts(run.result_json["assumption_register"])
               if r.key == "GATES.min_irr_5yr")
    assert row.provenance == A.SETTINGS
    assert row.value == 0.08
    assert row.was == 0.10, "the displaced value was read after the patch"


@pytest.mark.django_db
def test_an_analyst_edit_is_labelled_deal_by_the_real_worker(tmp_path):
    """The other provenance the CLI cannot produce, through the real
    worker: `Deal.cim_json` is the pristine extraction and the override
    is applied on top of it, so the register can report both."""
    from webapp.models import AnalysisRun, Deal
    from webapp.services import _analysis_worker, cim_to_dict

    cim = stabilized_deal()
    pristine = cim_to_dict(cim)
    deal = Deal.objects.create(
        deal_id="deal-provenance", property_name="Deal Provenance",
        deal_dir=str(tmp_path), cim_json=pristine,
        assumption_overrides={"cim_overrides": {"asking_price": 4_250_000}})
    run = AnalysisRun.objects.create(deal=deal, status="running")
    _analysis_worker(run.pk)

    run.refresh_from_db()
    assert run.status == "done", run.error
    row = next(r for r in A.from_dicts(run.result_json["assumption_register"])
               if r.key == "cim.asking_price")
    assert row.provenance == A.DEAL
    assert row.value == 4_250_000
    assert row.was == pristine["asking_price"]


def test_the_register_changes_no_number_it_reports(tmp_path):
    """The claim a reporting layer has to make, tested rather than
    asserted in a commit message: the same deal run with and without any
    provenance produces identical returns, gates and offers.

    The characterization snapshots make the same point across the whole
    pipeline; this makes it locally, so a future edit that lets the
    register reach into a resolver fails HERE, next to the code.
    """
    a, b = tmp_path / "a", tmp_path / "b"
    a.mkdir()
    b.mkdir()
    plain = _run(stabilized_deal(), a)
    stamped = _run(stabilized_deal(), b,
                   config_deltas={}, config_defaults={},
                   deal_overrides={}, cim_snapshot={})

    assert plain.scenario_results == stamped.scenario_results
    assert plain.gate_results == stamped.gate_results
    assert plain.max_offer == stamped.max_offer
    assert plain.sensitivity == stamped.sensitivity


@pytest.mark.django_db
def test_the_register_reaches_the_results_page(tmp_path, client,
                                               django_user_model, monkeypatch):
    """MUTATION: remove `ctx.update(results_ctx.register_context(r))` from
    `webapp/views.py` and this goes red while every other tab assertion in
    the suite still passes.

    The panel is closed by default, which is a rendering decision, not a
    disclosure one — the rows are in the DOM either way, and that is what
    this asserts.
    """
    from webapp.models import AnalysisRun, Deal

    django_user_model.objects.create_user(username="u", password="p")
    client.login(username="u", password="p")

    monkeypatch.setitem(cfg.GATES, "min_irr_5yr", 0.08)
    result = _run(stabilized_deal(), tmp_path,
                  config_deltas={"GATES.min_irr_5yr": 0.08},
                  config_defaults={"GATES.min_irr_5yr": 0.10})
    deal = Deal.objects.create(deal_id="register", property_name="Register")
    AnalysisRun.objects.create(
        deal=deal, status="done",
        result_json={"assumption_register": result.assumption_register,
                     "assumption_fill_log": [], "checks": [],
                     "check_summary": {}})

    # Unescaped, because a label like "Repairs & Maintenance" reaches the
    # DOM as `&amp;` — asserting against the raw body would test Django's
    # autoescaping rather than whether the row rendered.
    import html

    body = html.unescape(
        client.get(f"/deals/{deal.pk}/?tab=summary").content.decode())
    assert "Assumption Register" in body
    assert "settings override" in body
    for row in A.from_dicts(result.assumption_register):
        assert row.label in body, f"{row.key} registered but not on the page"


@pytest.mark.django_db
def test_the_results_page_marks_the_rows_a_human_chose(tmp_path, monkeypatch):
    """The register's whole value on screen is telling the two apart. A
    panel that renders 130 rows identically has disclosed the numbers and
    hidden the finding."""
    from webapp.results import register_context

    monkeypatch.setitem(cfg.GATES, "min_irr_5yr", 0.08)
    result = _run(stabilized_deal(), tmp_path,
                  config_deltas={"GATES.min_irr_5yr": 0.08},
                  config_defaults={"GATES.min_irr_5yr": 0.10})
    ctx = register_context({"assumption_register": result.assumption_register})
    assert ctx["register_counts"]["settings"] == 1
    assert [r["label"] for r in ctx["register_chosen"]], "no chosen rows split out"
    assert all(r["chosen"] for r in ctx["register_chosen"])
    # The displaced value renders through the SAME formatter as the value,
    # so a percent cannot print as 0.1 in one column and 10% in the other.
    gate = next(r for r in ctx["register_chosen"]
                if r["key"] == "GATES.min_irr_5yr")
    assert gate["value"] == "8.0%" and gate["was"] == "10.0%"


@pytest.mark.parametrize("field", ["year_built", "year_expanded"])
def test_a_vintage_renders_as_a_year_not_a_quantity(field, mock_cim_data):
    """Caught by reading the rendered memo, not by a test — `UNIT_COUNT`
    groups thousands, so a 2015 vintage printed as "2,015", which reads as
    a quantity of something. A year is a label, not a magnitude."""
    setattr(mock_cim_data, field, 2015)
    row = next(r for r in A.collect(cim_data=mock_cim_data)
               if r.key == f"cim.{field}")
    assert A.format_value(row) == "2015"


def test_the_appendix_lead_sentence_agrees_with_a_single_fill(tmp_path):
    """The count sentence is generated prose, so its grammar has to hold
    for the singular case too — the stabilized fixture fills exactly one
    input, which is the case a plural-only sentence gets wrong."""
    result = _run(stabilized_deal(), tmp_path)
    text = _memo_text(result.memo_path)
    assert "the CIM did not state a value" in text
    assert "did not state them" not in text
