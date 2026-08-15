"""Phase 4: AnalysisRun, background analysis, results pages, downloads."""
import json
import os

import pytest


@pytest.fixture
def operator(client, django_user_model):
    user = django_user_model.objects.create_user(username="op", password="x")
    client.force_login(user)
    return user


@pytest.fixture
def deals_dir(tmp_path, settings):
    d = tmp_path / "deals"
    d.mkdir()
    settings.CIM_DEALS_DIR = str(d)
    return d


def _sample_cim():
    from extract.parser import CIMData, UnitType
    return CIMData(
        property_name="Expo Storage", city="Belton", state="TX",
        nrsf=45000.0, total_units=350, physical_occupancy=0.92,
        economic_occupancy=0.78, asking_price=3_500_000.0,
        ttm_noi=250_000.0, ttm_egr=420_000.0, acreage=5.2,
        unit_mix=[UnitType(size_label="10x10", sf=100.0, count=100, rate=95.0),
                  UnitType(size_label="10x20", sf=200.0, count=50, rate=165.0,
                           climate_controlled=True)],
    )


def _make_extracted_deal(deals_dir, slug="expo"):
    """Deal with a completed extraction snapshot, ready to run."""
    from webapp.models import Deal
    from webapp.services import cim_to_dict
    folder = deals_dir / slug
    (folder / "inputs").mkdir(parents=True)
    (folder / "inputs" / "expo.pdf").write_bytes(b"%PDF-1.4 fake")
    cim = _sample_cim()
    return Deal.objects.create(
        deal_id=slug, property_name="Expo Storage", city="Belton", state="TX",
        deal_dir=str(folder), input_files=["expo.pdf"],
        extract_status="done", cim_json=cim_to_dict(cim),
        extraction_report=cim.extraction_report())


@pytest.mark.django_db
def test_analysis_run_defaults_and_ordering(deals_dir):
    from webapp.models import AnalysisRun
    deal = _make_extracted_deal(deals_dir)
    first = AnalysisRun.objects.create(deal=deal)
    second = AnalysisRun.objects.create(deal=deal)
    assert first.status == "running"
    assert first.result_json is None
    assert first.progress_total == 9
    assert deal.runs.first().pk == second.pk  # newest first


def test_json_safe_scrubs_nan_numpy_enum_tuple():
    import numpy as np
    from registry import ScenarioType
    from webapp.services import json_safe

    payload = {
        ScenarioType.BASE: {"irr": np.float64(0.123), "moic": float("nan")},
        "grid": [(np.float64(1.0), float("inf"))],
        "count": np.int64(7),
        "label": "ok",
        "scen": ScenarioType.BEAR,
    }
    out = json_safe(payload)
    # must round-trip through strict JSON (what Postgres JSONB enforces)
    restored = json.loads(json.dumps(out, allow_nan=False))
    assert restored["base"]["irr"] == pytest.approx(0.123)
    assert restored["base"]["moic"] is None
    assert restored["grid"] == [[1.0, None]]
    assert restored["count"] == 7
    assert restored["scen"] == "bear"


def test_json_safe_stringifies_unknown_objects():
    from webapp.services import json_safe

    class Weird:
        def __str__(self):
            return "weird"

    assert json_safe({"x": Weird()}) == {"x": "weird"}


def test_patched_replacement_cost_mutates_in_place_and_restores():
    from analysis.physical import REPLACEMENT_COST as bound
    from webapp.services import _patched_config

    original = bound["ss_driveup_per_sf"]
    with _patched_config({"REPLACEMENT_COST": {"ss_driveup_per_sf": [100, 120],
                                               "not_a_real_key": [1, 2]}}):
        assert tuple(bound["ss_driveup_per_sf"]) == (100, 120)
        assert "not_a_real_key" not in bound
    assert bound["ss_driveup_per_sf"] == original


def test_patched_replacement_cost_restores_on_exception():
    from analysis.physical import REPLACEMENT_COST as bound
    from webapp.services import _patched_config

    original = bound["ss_driveup_per_sf"]
    with pytest.raises(RuntimeError):
        with _patched_config({"REPLACEMENT_COST": {"ss_driveup_per_sf": [100, 120]}}):
            raise RuntimeError("boom")
    assert bound["ss_driveup_per_sf"] == original


def test_run_analysis_accepts_solver_target_irr():
    import inspect
    from engine import run_analysis
    params = inspect.signature(run_analysis).parameters
    assert "solver_target_irr" in params
    assert params["solver_target_irr"].default is None


def test_solver_honors_target_irr():
    from model.solver import solve_max_price
    out = solve_max_price(adjusted_ttm_noi=250_000.0, capex=0,
                          target_irr=0.12, expense_ratio=0.40)
    assert out["converged"]
    assert out["achieved_irr"] == pytest.approx(0.12, abs=0.005)


def test_engine_end_to_end_with_overrides(tmp_path, monkeypatch):
    """Real pipeline (no PDF, no network): overrides reach the solver and
    the writers produce files. Comp DB redirected to a scratch path —
    data.comp_db binds COMP_DB_PATH at import, so patch that module's name."""
    monkeypatch.setattr("data.comp_db.COMP_DB_PATH", str(tmp_path / "comps.db"))
    from engine import AnalysisResult, run_analysis
    from webapp.services import _patched_config

    result = AnalysisResult(pdf_path=str(tmp_path / "expo.pdf"))
    result.cim_data = _sample_cim()
    with _patched_config({"REPLACEMENT_COST": {"ss_driveup_per_sf": [100, 120]}}):
        result = run_analysis(result, output_dir=str(tmp_path),
                              solver_target_irr=0.12)
    assert result.max_offer["achieved_irr"] == pytest.approx(0.12, abs=0.005)
    assert os.path.isfile(result.memo_path)
    assert os.path.isfile(result.excel_path)
    assert result.gate_summary["recommendation"]

    # Item E4: the levered max offer is a SECOND LENS, so the unlevered
    # answer above must be untouched by its presence — and the two targets
    # must not bleed into each other. `solver_target_irr=0.12` is the
    # UNLEVERED target; the levered solver keeps its own 15% LP net
    # target, because one number cannot be both and forwarding the
    # unlevered one would silently re-price the levered answer whenever
    # an analyst edited the unlevered one.
    import config as cfg
    levered_offer = result.levered_max_offer
    assert levered_offer["max_price"] > 0
    assert levered_offer["target_irr"] == cfg.SOLVER_TARGET_LP_NET_IRR
    assert levered_offer["target_irr"] != 0.12
    assert levered_offer["lp_net_irr"] == pytest.approx(
        cfg.SOLVER_TARGET_LP_NET_IRR, abs=0.005)
    # Priced on the SAME deal terms the results page shows, so the loan
    # it reports is one this deal could actually raise.
    assert levered_offer["senior_debt"] > 0
    assert levered_offer["total_equity"] > 0
    assert levered_offer["assumption_stamp"]

    # Anti-drift guard: webapp/results.py (returns_context) and
    # output/excel_writer.py both read these exact sensitivity keys from
    # the real engine output — pin the shape so a fixture can never
    # silently diverge from what model/returns_model.py._build_sensitivity
    # actually emits.
    assert set(result.sensitivity.keys()) == {
        "price_labels", "price_values", "cap_labels", "cap_values",
        "irr_grid", "base_price", "base_exit_cap",
    }

    # The sample CIM (92% physical / 78% economic occupancy) trips the
    # econ/phys spread risk, so real risk items are produced here — pin
    # the keys and title-case severities that webapp/results.py's
    # risks_context depends on.
    risks = result.risk_analysis["risks"]
    assert risks
    for item in risks:
        assert {"category", "risk", "description", "severity",
                "mitigation"} <= item.keys()
        assert item["severity"] in {"High", "Medium", "Low"}


def test_engine_end_to_end_writes_the_investor_summary(tmp_path, monkeypatch):
    """Drives the REAL `run_analysis` through to the investor summary.

    Same reasoning as the template test below: `generate_investor_summary`
    is called inside a bare `except Exception` that appends to
    `result.errors`, so a renamed kwarg or a missing result key would be
    swallowed with the suite still green. `tests/test_investor_summary.py`
    calls the writer directly and cannot see the engine's wiring at all.
    """
    from docx import Document

    monkeypatch.setattr("data.comp_db.COMP_DB_PATH", str(tmp_path / "comps.db"))
    from engine import AnalysisResult, run_analysis

    result = AnalysisResult(pdf_path=str(tmp_path / "expo.pdf"))
    result.cim_data = _sample_cim()
    result = run_analysis(result, output_dir=str(tmp_path))

    # Match the message this writer emits, not a loose substring: the
    # errors list also carries the template writer's "Template not
    # found", whose PATH contains the word summary on some checkouts.
    assert not [e for e in result.errors
                if e.startswith("Investor summary")], result.errors
    assert os.path.isfile(result.investor_summary_path)

    body = "\n".join(p.text for p in Document(result.investor_summary_path).paragraphs)
    # The legend is what makes an un-cleared copy visibly not an offer.
    assert "not an offer to sell" in body


def test_engine_end_to_end_writes_the_template_with_resolved_terms(
        tmp_path, monkeypatch):
    """Drives the REAL `run_analysis` through to the XLSM writer.

    `generate_template` is called inside a bare `except Exception` that
    appends to `result.errors`, so a broken kwarg, a renamed parameter or
    an unbound `resolved_debt_terms` would be swallowed and the rest of
    the suite would stay green — the same blind spot that let the
    floating-rate bug (#28) and the market-cap double-resolve (#31) ship.
    Every other template test calls `generate_template` directly and so
    cannot see the engine's wiring at all. This one asserts the file is
    produced, that nothing was swallowed, and that the terms which
    arrived are the ones the run resolved rather than config defaults.
    """
    import openpyxl
    from output import template_writer
    from tests.test_template_writer import build_stub_template

    monkeypatch.setattr("data.comp_db.COMP_DB_PATH", str(tmp_path / "comps.db"))
    monkeypatch.setattr(template_writer, "TEMPLATE_PATH",
                        str(build_stub_template(tmp_path / "stub.xlsm")))
    from engine import AnalysisResult, run_analysis

    result = AnalysisResult(pdf_path=str(tmp_path / "expo.pdf"))
    result.cim_data = _sample_cim()
    result = run_analysis(result, output_dir=str(tmp_path),
                          debt_terms={"rate": 0.077, "term_years": 7},
                          waterfall_terms={"pref_rate": 0.09},
                          am_fee_pct=0.015)

    assert not [e for e in result.errors if "Template" in e], result.errors
    assert os.path.isfile(result.template_path)

    wb = openpyxl.load_workbook(result.template_path, keep_vba=True)
    try:
        ws = wb["Underwriting"]
        # The per-deal overrides reached the workbook, not config's
        # 6.25% / 10yr / 8% pref / 1% fee.
        assert ws["I74"].value == 0.077
        assert ws["F74"].value == 84
        assert ws["H249"].value == 0.09
        assert ws["G245"].value == 0.015
        # Leverage is on by default (E3a), so the run sized a loan and
        # the workbook is not all-equity.
        assert ws["H65"].value == pytest.approx(
            result.sources_uses["ltv"], abs=1e-6)
        assert ws["H65"].value > 0
    finally:
        wb.close()


def test_run_analysis_enrich_true_refills_missing_demographics(tmp_path,
                                                               monkeypatch):
    """enrich=True re-runs Census enrichment when a demographic field is
    still None — the post-assumptions second chance a corrected address
    never got — and surfaces the enrichment's own errors on the result."""
    monkeypatch.setattr("data.comp_db.COMP_DB_PATH", str(tmp_path / "comps.db"))
    from engine import AnalysisResult, run_analysis
    from extract.enrichment import EnrichmentResult

    calls = {"n": 0}

    def fake_enrich(cim_data, census_api_key=None, comp_db=None):
        calls["n"] += 1
        cim_data.population_3mi = 80_000
        res = EnrichmentResult(fields_enriched=1, geocode_success=True)
        res.errors.append("Cannot fetch demographics without geocoded address")
        return res

    monkeypatch.setattr("extract.enrichment.enrich_cim_data", fake_enrich)

    result = AnalysisResult(pdf_path=str(tmp_path / "expo.pdf"))
    result.cim_data = _sample_cim()  # demographics unset → gate fires
    result = run_analysis(result, output_dir=str(tmp_path), enrich=True)

    assert calls["n"] == 1
    assert result.cim_data.population_3mi == 80_000
    assert result.enrichment.fields_enriched == 1
    assert ("Enrichment: Cannot fetch demographics without geocoded address"
            in result.errors)


def test_run_analysis_enrichment_never_called_when_complete_or_off(
        tmp_path, monkeypatch):
    """The Census call is skipped when every enrichable field is already
    present (no pointless API traffic), and never happens on the default
    enrich=False path (CLI and tests stay network-free)."""
    monkeypatch.setattr("data.comp_db.COMP_DB_PATH", str(tmp_path / "comps.db"))
    from engine import ENRICHABLE_FIELDS, AnalysisResult, run_analysis

    def boom(*args, **kwargs):
        raise AssertionError("enrichment must not be called")
    monkeypatch.setattr("extract.enrichment.enrich_cim_data", boom)

    cim = _sample_cim()
    for f in ENRICHABLE_FIELDS:
        setattr(cim, f, 60_000)
    result = AnalysisResult(pdf_path=str(tmp_path / "expo.pdf"))
    result.cim_data = cim
    run_analysis(result, output_dir=str(tmp_path), enrich=True)  # complete → skip

    result2 = AnalysisResult(pdf_path=str(tmp_path / "expo2.pdf"))
    result2.cim_data = _sample_cim()  # incomplete, but enrich defaults False
    run_analysis(result2, output_dir=str(tmp_path))


@pytest.fixture
def fake_run(monkeypatch):
    """Stand-in for engine.run_analysis: fills the result fields the
    worker consumes, writes the three output files, captures kwargs."""
    calls = {}

    def _fake(result, progress=None, output_dir=None, custom_scenarios=None,
              custom_va_scenarios=None, solver_target_irr=None, enrich=False,
              expense_line_overrides=None, hold_years=None,
              transaction_costs=None, capital_structure=None,
              market_cap_rate=None, market_cap=None,
              debt_terms=None, waterfall_terms=None, am_fee_pct=None,
              mgmt_fee_target_pct=None, config_deltas=None,
              config_defaults=None, deal_overrides=None, cim_snapshot=None):
        calls["cim_data"] = result.cim_data
        calls["config_deltas"] = config_deltas
        calls["config_defaults"] = config_defaults
        calls["deal_overrides"] = deal_overrides
        calls["cim_snapshot"] = cim_snapshot
        calls["output_dir"] = output_dir
        calls["custom_scenarios"] = custom_scenarios
        calls["custom_va_scenarios"] = custom_va_scenarios
        calls["solver_target_irr"] = solver_target_irr
        calls["enrich"] = enrich
        calls["expense_line_overrides"] = expense_line_overrides
        calls["debt_terms"] = debt_terms
        calls["waterfall_terms"] = waterfall_terms
        calls["am_fee_pct"] = am_fee_pct
        if progress:
            progress(9, 9, "Generating memo & model...")
        name = result.cim_data.property_name.replace(" ", "_")
        for attr, suffix in [("memo_path", "_memo.docx"),
                             ("excel_path", "_model.xlsx"),
                             ("template_path", "_uw.xlsm"),
                             ("investor_summary_path", "_summary.docx")]:
            path = os.path.join(output_dir, f"{name}{suffix}")
            with open(path, "wb") as f:
                f.write(b"fake-office-bytes")
            setattr(result, attr, path)
        result.gate_results = [
            {"gate": 1, "name": "Population (3-mi ≥ 50K)", "threshold": "≥ 50,000",
             "actual": "62,000", "result": "PASS", "note": "", "source": None},
            {"gate": 2, "name": "No Unproven Demand", "threshold": "phys ≥ 75%",
             "actual": "92%", "result": "PASS", "note": "", "source": None},
        ]
        result.gate_summary = {"passed": 2, "failed": 0, "tbd": 0, "total": 2,
                               "recommendation": "PURSUE",
                               "failed_gates": [], "tbd_gates": []}
        result.scenario_results = {
            "bear": {"irr": 0.06, "moic": 1.3, "yield_on_cost": 0.065},
            "base": {"irr": float("nan"), "moic": 1.6, "yield_on_cost": 0.075},
            "bull": {"irr": 0.14, "moic": 1.9, "yield_on_cost": 0.085},
        }
        result.sensitivity = {"price_labels": ["-5.0%", "+0.0%", "+5.0%"],
                              "price_values": [3_325_000.0, 3_500_000.0, 3_675_000.0],
                              "cap_labels": ["5.50%", "6.00%", "6.50%"],
                              "cap_values": [0.055, 0.06, 0.065],
                              "irr_grid": [[0.11, 0.10, 0.09],
                                           [0.10, 0.09, 0.08],
                                           [0.09, 0.08, 0.07]],
                              "base_price": 3_500_000.0,
                              "base_exit_cap": 0.06}
        result.va_results = {
            "base": {"irr": 0.13, "moic": 1.7, "yield_on_cost": 0.08,
                     "development_spread": 0.02, "stabilized_noi": 300_000.0}}
        # Built by the real function, not hand-rolled: a fixture carrying
        # a stale schema is how a display test stays green past a change
        # that broke the page.
        from model.debt import build_debt_schedule, resolve_debt_terms
        from model.returns_model import build_sources_uses
        _debt = build_debt_schedule(3_500_000, 250_000, resolve_debt_terms(),
                                    hold_years=5)
        result.sources_uses = build_sources_uses(
            price=3_500_000, capex=100_000, acquisition_cost=35_000,
            reserve=50_000, financing_costs=_debt["financing_costs"],
            senior_debt=_debt["loan"], gp_coinvest_pct=0.10)
        # The levered lens (item E3a), built by the real functions for the
        # same reason the stack above is: a hand-rolled fixture carrying a
        # stale schema is how a persistence test stays green past a change
        # that broke it.
        from analysis.valuation import project_cash_flows
        from model.levered import build_levered_returns
        from model.waterfall import resolve_waterfall_terms
        # `exit_cap` is an argument, not a scenario parameter (PR #31): it
        # is derived from a market anchor. Pinned here so this fixture's
        # levered figures do not move with the market table.
        _projection = project_cash_flows(
            250_000, 3_500_000, 100_000,
            {"yr1_noi_bump": 0.0, "stabilized_occ": 0.88,
             "rev_cagr_yr1_3": 0.03, "rev_cagr_yr4_5": 0.03,
             "exp_growth": 0.03},
            hold_years=5, expense_ratio=0.40, reserve=50_000,
            exit_cap=0.08)
        result.debt = _debt
        result.levered = {"base": build_levered_returns(
            _projection, sources_uses=result.sources_uses, debt=_debt,
            waterfall_terms=resolve_waterfall_terms())}
        result.max_offer = {"max_price": 3_100_000.0, "achieved_irr": 0.10,
                            "converged": True}
        result.va_max_offer = {"max_price": 3_300_000.0, "achieved_irr": 0.10,
                               "converged": True}
        # Item E4, built by the REAL solver for the same reason the stack
        # and the levered lens above are: this payload is what the
        # persistence test asserts round-trips through JSONB, and a
        # hand-rolled dict would keep that test green past a change that
        # broke the actual shape. The two dicts above predate that rule
        # and are only ever read for `max_price`.
        from model.solver import solve_max_price_levered
        result.levered_max_offer = solve_max_price_levered(
            adjusted_ttm_noi=250_000, capex=100_000, expense_ratio=0.40,
            reserve=50_000)
        result.financial_analysis = {
            "adjusted_ttm_noi": {"cim_ttm_noi": 250_000.0,
                                 "analyst_adjusted_noi": 230_000.0},
            "adjustments": ["Property tax adjusted to benchmark",
                            {"category": "Insurance", "flag": "understated"}],
            "expense_ratio_check": {"opex_revenue_ratio": 0.42},
        }
        result.risk_analysis = {"risks": [
            {"category": "Market", "risk": "ECRI bridge", "severity": "High",
             "description": "Street rates falling", "mitigation": "Verify trend"}]}
        result.adjusted_noi = 230_000.0
        result.expense_ratio = 0.42
        result.checks = [
            {"id": "expense_line_floor", "label": "Expense line floors",
             "severity": "advisory", "status": "fail",
             "message": "Expense lines that cannot be taken at face value — "
                        "below half the benchmark floor: Property Taxes at "
                        "$0.00/SF vs a $1.20/SF floor.",
             "values": {}, "source": "financial_analysis"},
            {"id": "egr_le_gpr", "label": "EGR ≤ GPR", "severity": "blocking",
             "status": "pass", "message": "EGR sits below GPR.",
             "values": {}, "source": "ttm_gpr, ttm_egr"},
            {"id": "price_vs_replacement", "label": "Price vs replacement cost",
             "severity": "advisory", "status": "skipped",
             "message": "Asking price or replacement cost unavailable.",
             "values": {}, "source": "physical_analysis"},
        ]
        result.check_summary = {"total": 3, "passed": 1, "failed": 1,
                                "skipped": 1, "blocking_failed": 0,
                                "advisory_failed": 1}
        result.errors = ["Template generation failed: test-only"]
        return result

    monkeypatch.setattr("webapp.services.run_analysis", _fake)
    return calls


def _start_run(deal):
    from webapp import services
    from webapp.models import AnalysisRun
    run = AnalysisRun.objects.create(deal=deal)
    services.start_analysis(run)
    run.refresh_from_db()
    return run


@pytest.mark.django_db
def test_worker_success_updates_run_and_deal(deals_dir, fake_run):
    deal = _make_extracted_deal(deals_dir)
    deal.assumption_overrides = {
        "cim_overrides": {"asking_price": 3_400_000.0},
        "scenario_overrides": {"base": {"exit_cap": 0.06}},
        "va_scenario_overrides": {"base": {"target_occupancy": 0.9}},
        "replacement_cost_overrides": {"ss_driveup_per_sf": [100, 120]},
        "solver_target_irr": 0.12,
        "expense_line_overrides": {"payroll": 12600.0},
    }
    deal.save()
    run = _start_run(deal)

    assert run.status == "done"
    assert run.finished_at is not None
    assert run.memo_filename == "Expo_Storage_memo.docx"
    assert run.excel_filename == "Expo_Storage_model.xlsx"
    assert run.template_filename == "Expo_Storage_uw.xlsm"
    # NaN scrubbed for Postgres JSONB
    assert run.result_json["scenario_results"]["base"]["irr"] is None
    assert run.result_json["gate_summary"]["recommendation"] == "PURSUE"
    assert run.result_json["errors"] == ["Template generation failed: test-only"]
    # overrides all reached the engine
    assert fake_run["cim_data"].asking_price == 3_400_000.0
    assert fake_run["custom_scenarios"] == {"base": {"exit_cap": 0.06}}
    assert fake_run["custom_va_scenarios"] == {"base": {"target_occupancy": 0.9}}
    assert fake_run["solver_target_irr"] == 0.12
    assert fake_run["output_dir"] == deal.deal_dir
    # The web path must re-run enrichment (post-assumptions Census pass)
    assert fake_run["enrich"] is True
    assert fake_run["expense_line_overrides"] == {"payroll": 12600.0}

    deal.refresh_from_db()
    assert deal.recommendation == "PURSUE"
    assert deal.estimated_fair_value == 3_300_000.0  # VA max offer preferred
    assert deal.analysis_date is not None
    assert deal.memo_filename == "Expo_Storage_memo.docx"
    assert deal.excel_filename == "Expo_Storage_model.xlsx"


@pytest.mark.django_db
def test_worker_writes_meta_with_row_deal_id(deals_dir, fake_run):
    # Slug came from the FILENAME (Phase 3 decision #2); property name
    # differs. deal_meta.json must carry the row's slug so import_deals
    # round-trips onto the same row instead of forking a duplicate.
    deal = _make_extracted_deal(deals_dir, slug="expo-cim-v2")
    _start_run(deal)
    with open(os.path.join(deal.deal_dir, "deal_meta.json")) as f:
        meta = json.load(f)
    assert meta["deal_id"] == "expo-cim-v2"
    assert meta["property_name"] == "Expo Storage"
    assert meta["memo_path"] == "Expo_Storage_memo.docx"
    assert meta["input_files"] == ["expo.pdf"]


@pytest.mark.django_db
def test_worker_failure_records_error(deals_dir, monkeypatch):
    def boom(result, **kwargs):
        raise RuntimeError("solver exploded")

    monkeypatch.setattr("webapp.services.run_analysis", boom)
    deal = _make_extracted_deal(deals_dir)
    run = _start_run(deal)
    assert run.status == "failed"
    assert "solver exploded" in run.error
    deal.refresh_from_db()
    assert deal.recommendation == "N/A"  # deal row untouched on failure


@pytest.mark.django_db
def test_worker_progress_updates_row(deals_dir, fake_run):
    deal = _make_extracted_deal(deals_dir)
    run = _start_run(deal)
    assert run.progress_step == 9
    assert run.progress_msg == "Generating memo & model..."


@pytest.mark.django_db
def test_worker_deal_refresh_failure_rolls_back_run_done_flip(
        deals_dir, fake_run, monkeypatch):
    """A failure AFTER the AnalysisRun 'done' flip is queued (while
    building/applying deal_updates) must not leave the row stuck
    status='failed' while still carrying done-looking result_json /
    finished_at / filenames — that combination reads as done to any
    poller that checks status first. The done-flip and the Deal
    refresh must be all-or-nothing."""
    def boom(cim):
        raise RuntimeError("asset-type blew up")

    monkeypatch.setattr("webapp.services.detect_asset_type", boom)
    deal = _make_extracted_deal(deals_dir)
    run = _start_run(deal)

    assert run.status == "failed"
    assert "asset-type blew up" in run.error
    # Rolled back, not just re-flagged: no done-looking payload survives.
    assert run.result_json is None
    assert run.memo_filename == ""
    assert run.excel_filename == ""
    assert run.template_filename == ""
    deal.refresh_from_db()
    assert deal.recommendation == "N/A"  # deal row untouched too


@pytest.mark.django_db
def test_deal_run_starts_and_redirects(client, operator, deals_dir, fake_run):
    deal = _make_extracted_deal(deals_dir)
    resp = client.post(f"/deals/{deal.pk}/run/")
    assert resp.status_code == 302
    assert resp.url == f"/deals/{deal.pk}/"
    run = deal.runs.first()
    assert run.status == "done"  # sync mode ran inline


@pytest.mark.django_db
def test_deal_run_refuses_without_snapshot(client, operator, deals_dir):
    from webapp.models import Deal
    imported = Deal.objects.create(deal_id="legacy", property_name="Legacy")
    resp = client.post(f"/deals/{imported.pk}/run/")
    assert resp.status_code == 302
    assert imported.runs.count() == 0


@pytest.mark.django_db
def test_deal_run_refuses_while_running(client, operator, deals_dir, monkeypatch):
    from webapp import services
    from webapp.models import AnalysisRun
    deal = _make_extracted_deal(deals_dir)
    AnalysisRun.objects.create(deal=deal)  # status=running, fresh stamp
    monkeypatch.setattr(services, "start_analysis",
                        lambda run: pytest.fail("must not start a second run"))
    client.post(f"/deals/{deal.pk}/run/")
    assert deal.runs.count() == 1


@pytest.mark.django_db
def test_run_status_running_polls(client, operator, deals_dir):
    from webapp.models import AnalysisRun
    deal = _make_extracted_deal(deals_dir)
    AnalysisRun.objects.create(deal=deal, progress_step=3,
                               progress_msg="Analyzing market...")
    resp = client.get(f"/deals/{deal.pk}/run-status/")
    assert resp.status_code == 200
    assert b"hx-trigger" in resp.content
    assert b"Analyzing market..." in resp.content


@pytest.mark.django_db
def test_run_status_done_redirects(client, operator, deals_dir):
    from webapp.models import AnalysisRun
    deal = _make_extracted_deal(deals_dir)
    AnalysisRun.objects.create(deal=deal, status="done")
    resp = client.get(f"/deals/{deal.pk}/run-status/")
    assert resp.headers["HX-Redirect"] == f"/deals/{deal.pk}/"


@pytest.mark.django_db
def test_run_status_failed_and_timeout_stop_polling(client, operator, deals_dir):
    import datetime as dt

    from django.utils import timezone

    from webapp.models import AnalysisRun
    deal = _make_extracted_deal(deals_dir)
    run = AnalysisRun.objects.create(deal=deal, status="failed",
                                     error="solver exploded")
    resp = client.get(f"/deals/{deal.pk}/run-status/")
    assert b"hx-trigger" not in resp.content
    assert b"solver exploded" in resp.content
    # timeout: still "running" but created too long ago
    run.status = "running"
    run.error = ""
    run.save()
    AnalysisRun.objects.filter(pk=run.pk).update(
        created_at=timezone.now() - dt.timedelta(seconds=999))
    resp = client.get(f"/deals/{deal.pk}/run-status/")
    assert b"hx-trigger" not in resp.content
    assert b"timed out" in resp.content


@pytest.mark.django_db
def test_assumptions_save_and_run(client, operator, deals_dir, fake_run):
    deal = _make_extracted_deal(deals_dir)
    resp = client.post(f"/deals/{deal.pk}/assumptions/",
                       {"asking_price": "3400000", "run": "1"})
    assert resp.status_code == 302
    assert resp.url == f"/deals/{deal.pk}/"
    deal.refresh_from_db()
    assert deal.assumption_overrides["cim_overrides"]["asking_price"] == 3_400_000.0
    assert deal.runs.first().status == "done"


def test_results_formatters():
    from webapp.results import fmt_money, fmt_pct, fmt_x
    assert fmt_pct(0.123) == "12.3%"
    assert fmt_pct(None) == "N/A"
    assert fmt_money(3_100_000.0) == "$3,100,000"
    assert fmt_money(None) == "N/A"
    assert fmt_x(1.62) == "1.62x"
    assert fmt_x(None) == "N/A"


def _run_deal(client, deals_dir):
    deal = _make_extracted_deal(deals_dir)
    client.post(f"/deals/{deal.pk}/run/")
    return deal


@pytest.mark.django_db
def test_detail_summary_tab(client, operator, deals_dir, fake_run):
    deal = _run_deal(client, deals_dir)
    resp = client.get(f"/deals/{deal.pk}/")
    content = resp.content.decode()
    assert resp.status_code == 200
    assert "PURSUE" in content
    assert "Population (3-mi ≥ 50K)" in content       # gate row
    assert "Template generation failed: test-only" in content  # run warnings
    assert "Belton, TX" in content                     # property caption


@pytest.mark.django_db
def test_detail_returns_tab(client, operator, deals_dir, fake_run):
    deal = _run_deal(client, deals_dir)
    content = client.get(f"/deals/{deal.pk}/?tab=returns").content.decode()
    assert "14.0%" in content            # bull IRR
    assert "N/A" in content              # NaN base IRR scrubbed → N/A
    assert "$3,300,000" in content       # VA max offer
    assert "$3,325,000" in content       # sensitivity price row (price_values)
    assert "6.00%" in content            # sensitivity cap column (cap_values, 2dp)


@pytest.mark.django_db
def test_detail_financials_and_risks_tabs(client, operator, deals_dir, fake_run):
    deal = _run_deal(client, deals_dir)
    fin = client.get(f"/deals/{deal.pk}/?tab=financials").content.decode()
    assert "$230,000" in fin                       # analyst-adjusted NOI
    assert "Property tax adjusted to benchmark" in fin
    assert "Insurance: understated" in fin         # dict adjustment normalized
    risks = client.get(f"/deals/{deal.pk}/?tab=risks").content.decode()
    assert "ECRI bridge" in risks
    assert "Street rates falling" in risks             # description → detail column
    assert "bg-red-100 text-red-800" in risks           # High severity → red badge tone
    assert ">High</span>" in risks                      # badge text stays title-case


@pytest.mark.django_db
def test_detail_no_runs_shows_run_button(client, operator, deals_dir):
    deal = _make_extracted_deal(deals_dir)
    content = client.get(f"/deals/{deal.pk}/").content.decode()
    assert "Run Analysis" in content
    assert "tab=returns" not in content   # no tabs before first done run


@pytest.mark.django_db
def test_detail_imported_deal_has_no_run_button(client, operator):
    from webapp.models import Deal
    imported = Deal.objects.create(deal_id="legacy", property_name="Legacy",
                                   recommendation="PURSUE")
    content = client.get(f"/deals/{imported.pk}/").content.decode()
    assert "Run Analysis" not in content
    assert "no extraction snapshot" in content.lower()


@pytest.mark.django_db
def test_detail_bad_tab_falls_back_to_summary(client, operator, deals_dir, fake_run):
    deal = _run_deal(client, deals_dir)
    resp = client.get(f"/deals/{deal.pk}/?tab=nope")
    assert resp.status_code == 200
    assert b"Go / No-Go Gates" in resp.content


@pytest.mark.django_db
def test_run_payload_carries_the_check_register(client, operator, deals_dir,
                                                fake_run):
    """The register is stored WITH the run: a finding must stay attached to
    the numbers it was raised against, not be recomputed later against
    whatever the deal looks like by then."""
    deal = _run_deal(client, deals_dir)
    r = deal.runs.filter(status="done").first().result_json
    assert [c["id"] for c in r["checks"]] == [
        "expense_line_floor", "egr_le_gpr", "price_vs_replacement"]
    assert r["check_summary"]["advisory_failed"] == 1


@pytest.mark.django_db
def test_run_payload_carries_the_capital_stack(client, operator, deals_dir,
                                               fake_run):
    deal = _run_deal(client, deals_dir)
    su = deal.runs.filter(status="done").first().result_json["sources_uses"]
    # 3,685,000 of non-financing uses, plus the origination fee on the
    # loan item E3a now sizes for every deal. Derived rather than
    # hardcoded so a change to the config debt terms updates the
    # expectation instead of failing this test for the wrong reason.
    assert su["financing_costs"] > 0
    assert su["total_uses"] == pytest.approx(3_685_000 + su["financing_costs"])
    assert su["total_sources"] == pytest.approx(su["total_uses"])
    # Debt displaces equity; it does not add to uses.
    assert su["total_equity"] == pytest.approx(
        su["total_uses"] - su["senior_debt"])
    assert [u["key"] for u in su["uses"]] == [
        "price", "acquisition_cost", "capex", "reserve", "financing_costs"]


@pytest.mark.django_db
def test_run_payload_carries_the_levered_lens(client, operator, deals_dir,
                                              fake_run):
    """Item E3a's whole output. The first draft computed the levered lens
    on every deal and then dropped it: `result.debt` and `result.levered`
    were set in the engine, never added to the persisted payload, and
    discarded when the worker returned. Both audit agents caught it
    independently. Nothing surfaces these yet — E3b does that — but a
    figure that is not stored with its run cannot be surfaced later
    without recomputing it against whatever config says then, which is a
    different number wearing this run's date."""
    deal = _run_deal(client, deals_dir)
    payload = deal.runs.filter(status="done").first().result_json

    debt = payload["debt"]
    assert debt["loan"] > 0
    assert debt["binding_constraint"]
    # A dict, not the frozen dataclass stringified by json_safe.
    assert isinstance(debt["terms"], dict)
    assert debt["terms"]["rate"] > 0

    levered = payload["levered"]["base"]
    assert levered["lp_net_irr"] is not None
    assert levered["am_fee_pct"] > 0
    assert levered["am_fee_base"] == "lp_equity"
    assert len(levered["years"]) == 5
    assert levered["distributions"][0] == 0
    # The stamp travels with the number it qualifies.
    am_row = next(r for r in levered["assumption_stamp"]
                  if r["key"] == "am_fee_treatment")
    assert am_row["base"] == "lp_equity"
    assert "1.00%" in am_row["label"]

    # Item E4's payload, held to the same round-trip contract. This goes
    # through a REAL run and a real DB fetch, because the failure it
    # guards is invisible in memory: `json_safe` falls back to `str(obj)`
    # on anything it does not recognise, so a frozen dataclass reaching
    # JSONB persists as "DebtTerms(rate=0.0625, ...)" — unqueryable, and
    # a consumer reading ["terms"]["rate"] gets "string indices must be
    # integers". That is exactly how E3a lost `debt["terms"]`, and a new
    # persisted payload is where it would happen again.
    offer = payload["levered_max_offer"]
    assert offer["max_price"] > 0
    assert offer["lp_net_irr"] is not None
    assert offer["senior_debt"] > 0
    assert offer["total_equity"] > 0
    # Solved to the LP NET target, NOT the unlevered one.
    import config as cfg
    assert offer["target_irr"] == cfg.SOLVER_TARGET_LP_NET_IRR
    # Plain JSON types all the way down — no stringified dataclass.
    assert isinstance(offer["transaction_costs"], dict)
    assert isinstance(offer["assumption_stamp"], list)
    assert isinstance(offer["binding_constraint"], str)
    # The stamp survives the round trip, so the stored price can be
    # displayed later without recomputing what it assumed.
    assert {r["key"] for r in offer["assumption_stamp"]} == {
        "pref_compounding", "accrual_base", "ordering", "am_fee_treatment",
        "promote_basis", "catch_up"}
    # The unlevered max offer is untouched beside it — second lens, not a
    # replacement.
    assert payload["max_offer"]["max_price"] > 0


@pytest.mark.django_db
def test_capital_structure_is_stamped_even_at_the_defaults(
        client, operator, deals_dir, fake_run):
    """Same rule item B set for the hold and the cost percentages: a run
    that sat on the defaults must SAY so, or it is indistinguishable from
    a run that predates the setting existing."""
    import config as cfg

    deal = _run_deal(client, deals_dir)
    stamped = deal.runs.filter(status="done").first(
        ).applied_overrides["assumptions"]["capital_structure"]
    assert stamped == {
        "capex_basis": cfg.DEFAULT_CAPEX_BASIS,
        "operating_reserve": cfg.DEFAULT_OPERATING_RESERVE,
        "operating_reserve_basis": cfg.DEFAULT_OPERATING_RESERVE_BASIS,
        "gp_coinvest_pct": cfg.GP_COINVEST_PCT,
    }


@pytest.mark.django_db
def test_summary_tab_renders_the_capital_block(client, operator, deals_dir,
                                               fake_run):
    deal = _run_deal(client, deals_dir)
    content = client.get(f"/deals/{deal.pk}/?tab=summary").content.decode()

    su = deal.runs.filter(status="done").first().result_json["sources_uses"]

    assert "Capital" in content
    assert "Purchase Price" in content
    assert "LP Equity" in content
    # Rendered from the run's own stack rather than hardcoded: item E3a
    # sizes a loan on every deal, so Total Uses now carries an
    # origination fee and Equity is Uses LESS the loan.
    assert f"${su['total_uses']:,.0f}" in content
    assert f"Equity Required: ${su['total_equity']:,.0f}" in content
    assert f"GP 10%: ${su['gp_equity']:,.0f}" in content
    # In balance, so the warning chip must NOT be on the page.
    assert "OUT OF BALANCE" not in content


@pytest.mark.django_db
def test_summary_tab_flags_an_unbalanced_capital_stack(
        client, operator, deals_dir, fake_run, monkeypatch):
    """The one thing this block exists to catch has to be visible when it
    happens — a silent mismatch is worse than no block at all."""
    deal = _run_deal(client, deals_dir)
    run = deal.runs.filter(status="done").first()
    payload = run.result_json
    payload["sources_uses"]["balanced"] = False
    payload["sources_uses"]["delta"] = 1_234.0
    run.result_json = payload
    run.save(update_fields=["result_json"])
    content = client.get(f"/deals/{deal.pk}/?tab=summary").content.decode()
    assert "OUT OF BALANCE by $1,234" in content


@pytest.mark.django_db
def test_summary_tab_renders_the_check_register(client, operator, deals_dir,
                                                fake_run):
    deal = _run_deal(client, deals_dir)
    content = client.get(f"/deals/{deal.pk}/?tab=summary").content.decode()

    assert "Model Checks" in content
    assert "1 flagged" in content
    # The finding itself, not just the heading — a hardcoded "0 flagged"
    # summary line would otherwise leave every other assertion green.
    assert "below half the benchmark floor: Property Taxes" in content
    # Findings sort above the passes and the not-testable rows.
    assert (content.index("Expense line floors")
            < content.index("EGR ≤ GPR")
            < content.index("Price vs replacement cost"))
    # Register is open because something is flagged.
    assert "<details open>" in content


@pytest.mark.django_db
def test_download_endpoints_serve_run_outputs(client, operator, deals_dir, fake_run):
    deal = _run_deal(client, deals_dir)
    for kind, filename, mime in [
        ("memo", "Expo_Storage_memo.docx",
         "application/vnd.openxmlformats-officedocument.wordprocessingml.document"),
        ("excel", "Expo_Storage_model.xlsx",
         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ("template", "Expo_Storage_uw.xlsm",
         "application/vnd.ms-excel.sheet.macroEnabled.12"),
        ("investor_summary", "Expo_Storage_summary.docx",
         "application/vnd.openxmlformats-officedocument.wordprocessingml.document"),
    ]:
        resp = client.get(f"/deals/{deal.pk}/download/{kind}/")
        assert resp.status_code == 200
        assert filename in resp.headers["Content-Disposition"]
        assert resp.headers["Content-Type"] == mime


@pytest.mark.django_db
def test_download_legacy_deal_falls_back_to_deal_row(client, operator, deals_dir):
    from webapp.models import Deal
    folder = deals_dir / "legacy"
    folder.mkdir()
    (folder / "memo.docx").write_bytes(b"legacy-memo")
    legacy = Deal.objects.create(deal_id="legacy", property_name="Legacy",
                                 deal_dir=str(folder), memo_filename="memo.docx")
    resp = client.get(f"/deals/{legacy.pk}/download/memo/")
    assert resp.status_code == 200
    resp = client.get(f"/deals/{legacy.pk}/download/template/")
    assert resp.status_code == 404


@pytest.mark.django_db
def test_download_rejects_bad_kind_missing_file_and_escape(client, operator, deals_dir):
    from webapp.models import Deal
    deal = _make_extracted_deal(deals_dir)
    assert client.get(f"/deals/{deal.pk}/download/nope/").status_code == 404
    assert client.get(f"/deals/{deal.pk}/download/memo/").status_code == 404  # no file yet
    # a poisoned stored filename must not escape the deal folder
    outside = deals_dir.parent / "secret.docx"
    outside.write_bytes(b"secret")
    deal.memo_filename = "../../secret.docx"
    deal.save()
    assert client.get(f"/deals/{deal.pk}/download/memo/").status_code == 404


@pytest.mark.django_db
def test_deal_list_links_detail(client, operator, deals_dir, fake_run):
    deal = _run_deal(client, deals_dir)
    resp = client.get("/deals/")
    assert f'href="/deals/{deal.pk}/"'.encode() in resp.content


@pytest.mark.django_db
def test_the_assumptions_page_renders_the_unit_stamps(client, operator, deals_dir):
    """The unit guard in webapp.forms is only real if the page actually
    emits the stamps it reads."""
    deal = _make_extracted_deal(deals_dir)
    content = client.get(f"/deals/{deal.pk}/assumptions/").content.decode()
    assert 'name="capex_unit_stamp" value="amount"' in content
    assert 'name="reserve_unit_stamp" value="amount"' in content
    # And the CapEx basis selector rides in the CapEx driver row itself.
    assert 'aria-label="CapEx basis"' in content


@pytest.mark.django_db
def test_a_unit_change_is_refused_once_then_saves(client, operator, deals_dir):
    """End-to-end through the real page: the refusal has to reach the
    analyst, and the second attempt has to go through — a guard that
    cannot be satisfied is worse than no guard."""
    deal = _make_extracted_deal(deals_dir)
    url = f"/deals/{deal.pk}/assumptions/"
    post = {"asking_price": "3500000", "nrsf": "45000", "total_units": "350",
            "ttm_noi": "250000", "ttm_egr": "420000", "state": "TX",
            "physical_occupancy": "92", "economic_occupancy": "78",
            "capex_estimate": "2", "capex_basis": "pct_price",
            "capex_unit_stamp": "amount", "reserve_unit_stamp": "amount"}

    first = client.post(url, post)
    assert first.status_code == 422
    assert "will now be read as % of price" in first.content.decode()
    deal.refresh_from_db()
    assert deal.assumption_overrides in (None, {})

    # The page re-renders stamping the NEW selection, so resubmitting the
    # figure the analyst just confirmed is accepted.
    second = client.post(url, {**post, "capex_unit_stamp": "pct_price"})
    assert second.status_code == 302
    deal.refresh_from_db()
    assert deal.assumption_overrides["capital_structure"]["capex_basis"] == "pct_price"
    assert deal.assumption_overrides["cim_overrides"]["capex_estimate"] == 0.02


@pytest.mark.django_db
def test_investor_summary_button_appears_only_once_a_run_produced_one(
        client, operator, deals_dir, fake_run):
    """The button is conditional on the RUN's filename, like the .xlsm
    one. Generation is wrapped in try/except, so a deal whose summary
    failed must not offer a link that 404s."""
    deal = _run_deal(client, deals_dir)
    content = client.get(f"/deals/{deal.pk}/").content.decode()
    assert "Investor Summary (.docx)" in content

    run = deal.runs.filter(status="done").first()
    assert run.investor_summary_filename == "Expo_Storage_summary.docx"

    run.investor_summary_filename = ""
    run.save(update_fields=["investor_summary_filename"])
    content = client.get(f"/deals/{deal.pk}/").content.decode()
    assert "Investor Summary (.docx)" not in content
    # ...and the endpoint itself refuses rather than serving a stale file.
    assert client.get(
        f"/deals/{deal.pk}/download/investor_summary/").status_code == 404


# ── Item G: the distribution gate on the download surface ────────────

@pytest.mark.django_db
def test_the_investor_summary_button_warns_while_gc_has_not_cleared(
        client, deals_dir, monkeypatch, django_user_model):
    """MUTATION: drop `"gc_cleared": cfg.INVESTOR_SUMMARY_GC_CLEARED` from
    the deal_detail context, or the `{% if not gc_cleared %}` block from
    the template.

    The gate lived in a backlog paragraph and a code comment — the two
    places the analyst clicking this button will never look.
    """
    import config as cfg
    from webapp.models import AnalysisRun

    monkeypatch.setattr(cfg, "INVESTOR_SUMMARY_GC_CLEARED", False)
    django_user_model.objects.create_user(username="u", password="p")
    client.login(username="u", password="p")

    deal = _make_extracted_deal(deals_dir)
    AnalysisRun.objects.create(deal=deal, status="done", result_json={},
                               investor_summary_filename="s.docx")

    body = client.get(f"/deals/{deal.pk}/").content.decode()
    assert "has not been cleared by counsel" in body
    assert "internal only" in body
    # The other three downloads are unconditional; only this one is gated.
    assert "Returns Model (.xlsx)" in body


@pytest.mark.django_db
def test_clearing_the_gate_removes_the_download_caveat(
        client, deals_dir, monkeypatch, django_user_model):
    """The flag has to be a flag on this surface too, or the operator
    clears it in config and the page keeps crying wolf."""
    import config as cfg
    from webapp.models import AnalysisRun

    monkeypatch.setattr(cfg, "INVESTOR_SUMMARY_GC_CLEARED", True)
    django_user_model.objects.create_user(username="u", password="p")
    client.login(username="u", password="p")

    deal = _make_extracted_deal(deals_dir)
    AnalysisRun.objects.create(deal=deal, status="done", result_json={},
                               investor_summary_filename="s.docx")

    body = client.get(f"/deals/{deal.pk}/").content.decode()
    assert "has not been cleared by counsel" not in body
    assert "internal only" not in body
    assert "Investor Summary (.docx)" in body


# ── Runs stored before a section existed (QA findings C4 / C5-b) ────
# A run from 2026-07-28 rendered a Summary tab carrying two of its six
# blocks and a Returns tab with no levered lens at all. Nothing was
# broken: every absent block postdated the run, and each is correctly
# gated on its own payload key. What was missing was any way for the
# reader to tell that apart from a rendering bug — a browser QA pass
# filed it as one, twice, at "major".


def _legacy_payload():
    """A run stored before `checks`, `sources_uses`, the fill log, the
    register and the levered lens existed. No `payload_version` key at
    all, which is what every run written before the stamp looks like."""
    return {"gate_summary": {"recommendation": "DECLINE",
                             "gates": [], "pass_count": 0, "total": 7},
            "scenario_results": {"base": {"irr": 0.12, "moic": 1.6}}}


def _legacy_deal(client, deals_dir, **payload_extra):
    from webapp.models import AnalysisRun
    deal = _make_extracted_deal(deals_dir)
    AnalysisRun.objects.create(deal=deal, status="done",
                               result_json=dict(_legacy_payload(),
                                                **payload_extra))
    return deal


@pytest.mark.django_db
def test_a_stale_run_says_which_sections_it_predates(client, operator,
                                                     deals_dir):
    """MUTATION: drop the `ctx.update(legacy_context(...))` line from
    `deal_detail`, or the `{% if run_is_legacy %}` block from the
    template."""
    deal = _legacy_deal(client, deals_dir)
    body = client.get(f"/deals/{deal.pk}/").content.decode()

    assert "predates some of the sections on this page" in body
    # Naming them is the point. "Some sections are missing" leaves the
    # reader exactly where the QA pass was — unable to tell an absent
    # block from a broken one.
    for label in ("Model Checks", "Capital (Sources &amp; Uses)",
                  "Assumptions Filled", "Assumption Register",
                  "Levered Returns (LP Net)"):
        assert label in body, label


@pytest.mark.django_db
def test_the_stale_notice_rides_every_tab(client, operator, deals_dir):
    """The register went missing on Summary and the levered lens on
    Returns, so a notice rendering only on Summary would leave the
    Returns reader with the same unexplained gap. It sits outside the
    per-tab branches for that reason."""
    deal = _legacy_deal(client, deals_dir)
    for tab in ("summary", "returns", "financials", "risks"):
        body = client.get(f"/deals/{deal.pk}/?tab={tab}").content.decode()
        assert "predates some of the sections" in body, tab


@pytest.mark.django_db
def test_a_current_run_missing_a_block_is_not_called_stale(client, operator,
                                                           deals_dir):
    """The distinction the whole version stamp exists for, and the only
    reason `legacy_context` is not just an emptiness test.

    A modern run with no `levered` block priced no loan, and the Returns
    tab says so in its own words. Calling that stale would fire the
    banner on a healthy deal, and a caveat that fires on healthy deals
    stops being read — which is the failure decision 8 names for
    `coerced_region`."""
    from webapp.results import RESULT_PAYLOAD_VERSION

    deal = _legacy_deal(client, deals_dir,
                        payload_version=RESULT_PAYLOAD_VERSION)
    body = client.get(f"/deals/{deal.pk}/").content.decode()
    assert "predates some of the sections" not in body


def test_the_notice_is_silent_when_nothing_is_actually_absent():
    """An unversioned run carrying every block gets no banner: a version
    number is bookkeeping the reader cannot act on, and there is nothing
    on the page for them to go looking for."""
    from webapp.results import legacy_context

    full = {"checks": [{}], "sources_uses": {"uses": []},
            "assumption_fill_log": [{}], "assumption_register": [{}],
            "levered": {"base": {}}}
    ctx = legacy_context(full)
    assert ctx["run_is_legacy"] is False


def test_an_unversioned_run_reports_only_the_blocks_it_lacks():
    """`legacy_missing` is the blocks actually absent, not the whole
    catalogue — a notice naming five sections on a run missing one
    sends the reader looking for four that are on the page."""
    from webapp.results import legacy_context

    ctx = legacy_context({"checks": [{}], "sources_uses": {"uses": []},
                          "assumption_fill_log": [{}],
                          "assumption_register": [{}]})
    assert ctx["run_is_legacy"] is True
    assert ctx["legacy_missing"] == ["Levered Returns (LP Net)"]


@pytest.mark.django_db
def test_the_worker_stamps_the_version_on_what_it_writes(client, operator,
                                                         deals_dir, fake_run):
    """MUTATION: drop `payload_version` from the payload dict in
    `webapp.services`. Every NEW run then reads as stale, and a banner
    that exists to explain four missing blocks starts appearing on deals
    that have all six."""
    from webapp.results import RESULT_PAYLOAD_VERSION

    deal = _run_deal(client, deals_dir)
    run = deal.runs.filter(status="done").first()
    assert run.result_json["payload_version"] == RESULT_PAYLOAD_VERSION
    # ...and the page it produces is therefore never called stale.
    assert "predates some of the sections" not in \
        client.get(f"/deals/{deal.pk}/").content.decode()
