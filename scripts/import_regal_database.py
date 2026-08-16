#!/usr/bin/env python3
"""
Import the BCRE/Regal self-storage market database into the comp database.

The source workbook ("BCRE REGAL FULL SELF STORAGE DATABASE.xlsx") is a
Texas-focused market database compiled by Regal from County Appraisal District
(CAD) records. This importer maps its CAD sheets into the ``properties`` table
of the SQLite schema that ``data/comp_db.py`` maintains, as a PROPERTY
reference — NRSF, year built, acreage, location — never as rents.

**The workbook's data is ~2013-2016 and every row says so.** ``analysis_date``
carries the SOURCE's vintage (see ``SHEET_VINTAGE``), not the import wall-clock
time. That is the whole point: nothing downstream filters comps by age, and
``get_comp_summary`` orders by ``analysis_date DESC``, so a decade-old CAD row
stamped with today's date sorts ABOVE a real analysis from last month and reads
as current. It is imported as what it is or not at all.

What is imported
    properties   - facilities carrying a usable NRSF + location, from the four
                   richest CAD sheets (Travis/TCadSqFt, Bexar/BcadNRSF,
                   Williamson/Wcad, Harris/hcad).
    data_sources - provenance rows (tier=3,
                   source_detail="BCRE Regal DB (<sheet>, vintage <year>)").

What is deliberately NOT imported
    unit_mix     - the workbook's unit-rate columns are ~2015 ASKING rents.
                   Seeding them made 158 of 166 TX rent comps one decade-old
                   portfolio presenting as current, which moved a $1.20/SF
                   subject's Tier-3 rent gap from +12.1% to +9.0%. Rents are a
                   live underwriting input; stale ones are worse than none,
                   because ``query_rent_comps`` returning None falls back to
                   the CIM's own stated market rent. Re-adding rents means
                   CURRENT data and a vintage filter in the query, not this
                   workbook.
    price_per_sf - the CAD figure is a TAX-APPRAISED value per SF. The column
                   it would land in is filled by ``save_analysis`` with broker
                   ASKING price / NRSF. Blending the two silently corrupts any
                   future price benchmark, so the appraised value is recorded
                   in ``data_sources`` (with its vintage) and the column is
                   left NULL.

Rows are tagged ``pdf_filename = "[regal:<sheet>:<source_id>].xlsx"`` so they
can be upserted idempotently on re-run and swept clean with ``--reset`` without
touching rows that came from real CIM analyses. Note the trailing ``.xlsx``:
a sweep written against the bracketed form alone matches nothing.

Usage
    python scripts/import_regal_database.py                # upsert into data/cim_comps.db
    python scripts/import_regal_database.py --file X.xlsx  # a different workbook
    python scripts/import_regal_database.py --db /path.db  # a different comp DB
    python scripts/import_regal_database.py --dry-run      # report only, no writes
    python scripts/import_regal_database.py --reset        # drop prior regal rows first
"""

import argparse
import os
import re
import sqlite3
import sys

# Make repo-root modules importable when run as `python scripts/....py`.
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

try:
    import openpyxl
except ImportError:  # pragma: no cover - dependency guard
    sys.exit("Missing dependency: pip install openpyxl")

from config import COMP_DB_PATH  # noqa: E402


class ImportError_(Exception):
    """A sheet did not import as configured. Raised rather than reported."""


# ── Scalar coercion ────────────────────────────────────────────────────

_ERROR_TOKENS = {"none", "nan", "null", "n/a", "site", "unknown",
                 "#num!", "#ref!", "#value!", "#div/0!", "#n/a"}


def to_float(v):
    """Best-effort float conversion; returns None for blanks / CAD error cells."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return f if f == f else None  # drop NaN
    s = str(v).strip().replace(",", "").replace("$", "").replace("%", "")
    if not s or s.lower() in _ERROR_TOKENS:
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return f if f == f else None


def to_int(v):
    f = to_float(v)
    if f is None:
        return None
    return int(round(f))


def split_latlon(v):
    """Parse '30.659, -97.874' style cells into (lat, lon) or (None, None)."""
    if v is None:
        return None, None
    parts = re.split(r"[,\s]+", str(v).strip())
    nums = [n for n in (to_float(p) for p in parts) if n is not None]
    if len(nums) >= 2:
        return nums[0], nums[1]
    return None, None


# ── Sheet layouts (CAD sheets rich enough to become properties) ────────

REGAL_PREFIX = "[regal:"
# Match imported rows by their literal "[regal:" prefix. SQLite LIKE treats
# '[' as an ordinary literal here (not a character-class wildcard), but we
# compare with substr() anyway so the sweep never depends on LIKE bracket
# semantics that differ across SQLite builds.
REGAL_PREFIX_LEN = len(REGAL_PREFIX)  # len("[regal:") == 7

PROPERTY_SHEETS = {
    "TCadSqFt": {  # Travis County (Austin)
        "key": "PropertyID",
        "header_rows": 1,
        "fields": {
            "name": "BusinessName",
            "address": "StreetAddress",
            "city": "City",
            "state": "State",
            "zip_code": "ZipCode",
            "year_built": "YearBuilt",
            "acreage": "Acres",
            "nrsf": "NRSF",
            "price_per_sf": "AppValPerSqFt",
            "lat": "Latitude",
            "lon": "Longitude",
        },
    },
    "BcadNRSF": {  # Bexar County (San Antonio)
        "key": "PropertyID",
        "header_rows": 3,
        "fields": {
            "name": "DBA",
            "address": "Property Address",
            "zip_code": "Zip Code",
            "year_built": "Year Built",
            "acreage": "Land, Acres",
            "nrsf": "NRSF",
            "lat": "LAT",
            "lon": "LONG",
        },
        "state": "TX",
    },
    "Wcad": {  # Williamson County
        "key": "PropertyQuickRefID",
        "header_rows": 1,
        "fields": {
            "name": "DBA",
            "name_fallback": "OwnerName",
            "address": "SitusAddress",
            "year_built": "YearBuilt",
            "acreage": "Acres",
            "nrsf": "Bldg SF",
            "appr_val": "2015AppVal",
            "latlon": "Lat, Long",
        },
        "state": "TX",
    },
    "hcad": {  # Harris County (Houston)
        "key": "AccountNumber",
        "header_rows": 1,
        "fields": {
            "name": "DBA",
            "name_fallback": "Owner Name",
            "address": "PropertyAddress",
            "zip_code": "Zip",
            "year_built": "Built",
            "nrsf": "BldgSqFt",
            "price_per_sf": "AppValPSF",
            "appr_val": "MktValHCAD",
        },
        "state": "TX",
    },
}

# ── Source vintage ─────────────────────────────────────────────────────
#
# The date each sheet's data actually describes, read off the workbook's own
# headers — never invented, and never newer than the evidence.
#
#   TCadSqFt  the newest dated appraisal column on the sheet is "2015AppVal"
#             (it also carries 2013AppVal / 2014AppVal / 2014LandAppVal).
#   Wcad      the mapped value column IS "2015AppVal". Its TaxYear column
#             reads 2016 and PropertyValueTaxYear reads 2015; the value we
#             import is the 2015 one.
#   BcadNRSF  no year appears anywhere in its three header rows.
#   hcad      no year either — ApprValHCAD / MktValHCAD / AppValPSF are all
#             undated.
#
# For the two sheets that state no year, we use the workbook's newest dated
# evidence (Wcad's TaxYear 2016) as an UPPER BOUND. That is deliberately the
# least-stale defensible reading: claiming older would overstate the staleness
# as confidently as claiming newer would understate it. Both are recorded as
# bounds, not measurements, in the provenance string.
SHEET_VINTAGE = {
    "TCadSqFt": ("2015-01-01", "2015 appraisal column"),
    "BcadNRSF": ("2016-01-01", "not stated on sheet; bounded by workbook (Wcad TaxYear 2016)"),
    "Wcad": ("2015-01-01", "2015AppVal column"),
    "hcad": ("2016-01-01", "not stated on sheet; bounded by workbook (Wcad TaxYear 2016)"),
}


def _header_index(header_rows):
    """Map lowercased header label -> column index, across ALL header rows.

    Reading row 1 only was a live bug: BcadNRSF declares ``header_rows: 3`` and
    puts its field labels in row 1, but a sheet whose labels sat in row 2 or 3
    would resolve every column to None and import as ZERO properties with no
    error. First label to claim a column wins, so a merged multi-level header
    keeps its top-level name.
    """
    index = {}
    for header_row in header_rows:
        for i, v in enumerate(header_row):
            if v in (None, ""):
                continue
            name = str(v).strip().lower()
            if name and name not in index:
                index[name] = i
    return index



# ── Database helpers ───────────────────────────────────────────────────

def _insert_property(conn, sheet, key, row, header, sheet_cfg):
    """Upsert one facility. Returns property_id, or None when it is skipped."""
    def col(field):
        name = sheet_cfg["fields"].get(field)
        return header.get(name.lower()) if name else None

    def cell(field):
        i = col(field)
        return row[i] if i is not None and i < len(row) else None

    nrsf = to_float(cell("nrsf"))
    if not nrsf or nrsf <= 0:
        return None  # only import facilities with a usable square-footage

    name = cell("name") or cell("name_fallback") or sheet
    state = cell("state") or sheet_cfg.get("state", "TX")
    state = str(state).strip().upper()[:2] if state else "TX"

    lat = lon = None
    if col("latlon") is not None:
        lat, lon = split_latlon(cell("latlon"))
    if col("lat") is not None:
        lat = to_float(cell("lat")) or lat
    if col("lon") is not None:
        lon = to_float(cell("lon")) or lon

    # A CAD appraised value per SF. It is NOT an asking price, and it does not
    # go in properties.price_per_sf — see the module docstring. Recorded as
    # provenance only, so the figure survives without pretending to be a comp.
    appraised_per_sf = to_float(cell("price_per_sf"))
    appr_val = to_float(cell("appr_val"))
    if appraised_per_sf is None and appr_val and nrsf:
        appraised_per_sf = appr_val / nrsf

    vintage_date, vintage_basis = SHEET_VINTAGE[sheet]
    provenance = f"BCRE Regal DB ({sheet}, vintage {vintage_date[:4]} — {vintage_basis})"

    filename = f"{REGAL_PREFIX}{sheet}:{key}].xlsx"

    # Clear the prior version of this imported row (idempotent upsert).
    old_id = conn.execute(
        "SELECT id FROM properties WHERE pdf_filename = ?", (filename,)
    ).fetchone()
    old_id = old_id[0] if old_id else None
    if old_id:
        for tbl in ("data_sources", "unit_mix", "expense_lines"):
            conn.execute(f"DELETE FROM {tbl} WHERE property_id = ?", (old_id,))
        conn.execute("DELETE FROM properties WHERE id = ?", (old_id,))

    conn.execute(
        """
        INSERT INTO properties (
            property_name, address, city, state, msa, zip_code, lat, lon,
            year_built, nrsf, total_units, cc_pct, acreage, occupancy,
            asking_price, price_per_sf,
            population_1mi, population_3mi, population_5mi, median_hhi_3mi,
            ttm_gpr, ttm_egr, ttm_noi, adjusted_noi,
            revenue_per_sf, noi_per_sf, opex_ratio,
            market_rent_psf, in_place_rent_psf,
            pdf_filename, analysis_date
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            str(name), cell("address"), cell("city"), state, None,
            cell("zip_code"), lat, lon,
            to_int(cell("year_built")), nrsf, None, None,
            to_float(cell("acreage")), None,
            None, None,          # asking_price, price_per_sf: never from a CAD appraisal
            None, None, None, None,
            None, None, None, None,
            None, None, None,
            None, None,
            filename,
            vintage_date,        # the SOURCE's vintage, not the import time
        ),
    )
    property_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    # Provenance rows for the fields that actually moved a number.
    sources = [
        ("nrsf", nrsf),
        ("year_built", to_int(cell("year_built"))),
        ("acreage", to_float(cell("acreage"))),
        # Named for what it is. NOT "price_per_sf" — that name is what invited
        # the confusion with the asking-price column in the first place.
        ("cad_appraised_per_sf", appraised_per_sf),
        ("state", state),
    ]
    if lat is not None and lon is not None:
        sources.append(("lat_lon", f"{lat},{lon}"))
    for field, value in sources:
        if value is None or value == "":
            continue
        conn.execute(
            "INSERT INTO data_sources (property_id, field_name, tier, source_detail, value_used) "
            "VALUES (?, ?, ?, ?, ?)",
            (property_id, field, 3, provenance, str(value)),
        )

    return property_id


def _clear_regal_rows(conn):
    rows = conn.execute(
        "SELECT id FROM properties "
        "WHERE substr(pdf_filename, 1, ?) = ?",
        (REGAL_PREFIX_LEN, REGAL_PREFIX),
    ).fetchall()
    ids = [r[0] for r in rows]
    for pid in ids:
        for tbl in ("data_sources", "unit_mix", "expense_lines"):
            conn.execute(f"DELETE FROM {tbl} WHERE property_id = ?", (pid,))
    conn.execute(
        "DELETE FROM properties WHERE substr(pdf_filename, 1, ?) = ?",
        (REGAL_PREFIX_LEN, REGAL_PREFIX),
    )
    return len(ids)



# ── Main ───────────────────────────────────────────────────────────────

def run_import(workbook_path, db_path, dry_run=False, reset=False):
    """Import the workbook's CAD sheets. ALL-OR-NOTHING.

    The sweep and the reload are one transaction, committed once at the end.
    Previously ``--reset`` committed before the import started and every row
    committed individually, so any mid-run failure left the DB swept and
    half-loaded — with no rollback copy, since the comp DB is gitignored.
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    conn = sqlite3.connect(db_path)
    try:
        stats = {"properties": 0, "skipped_no_nrsf": 0, "skipped_no_key": 0,
                 "sheets": {}, "missing_sheets": []}
        conn.execute("BEGIN")
        if reset:
            stats["cleared"] = _clear_regal_rows(conn)

        for sheet_name, cfg in PROPERTY_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                stats["missing_sheets"].append(sheet_name)
                continue
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            # Read ALL header rows, then index across them. Consuming rows 2..N
            # without reading them is what let a mis-declared header silently
            # import a whole sheet as zero properties.
            header_rows = [r for r in (next(rows, None) for _ in range(cfg["header_rows"]))
                           if r is not None]
            header = _header_index(header_rows)
            key_col = header.get(str(cfg["key"]).lower())
            if key_col is None:
                raise ImportError_(
                    f"{sheet_name}: key column {cfg['key']!r} not found in the first "
                    f"{cfg['header_rows']} row(s). Every row would be skipped and the "
                    f"sheet would import as zero properties."
                )
            sheet_props = sheet_no_key = sheet_no_nrsf = 0
            for row in rows:
                if key_col >= len(row) or row[key_col] in (None, ""):
                    sheet_no_key += 1
                    continue
                if _insert_property(conn, sheet_name, row[key_col], row, header, cfg) is None:
                    sheet_no_nrsf += 1
                    continue
                sheet_props += 1
            if sheet_props == 0:
                raise ImportError_(
                    f"{sheet_name}: imported 0 properties "
                    f"({sheet_no_key} rows without a key, {sheet_no_nrsf} without usable NRSF). "
                    f"A configured sheet contributing nothing is a mapping failure, not an "
                    f"empty sheet — refusing rather than reporting success."
                )
            stats["properties"] += sheet_props
            stats["skipped_no_key"] += sheet_no_key
            stats["skipped_no_nrsf"] += sheet_no_nrsf
            stats["sheets"][sheet_name] = {
                "properties": sheet_props,
                "vintage": SHEET_VINTAGE[sheet_name][0][:4],
                "skipped_no_key": sheet_no_key,
                "skipped_no_nrsf": sheet_no_nrsf,
            }

        if dry_run:
            conn.rollback()
        else:
            conn.commit()
        return stats
    except BaseException:
        conn.rollback()
        raise
    finally:
        conn.close()


def main(argv=None):
    parser = argparse.ArgumentParser(description="Import the BCRE/Regal self-storage database.")
    parser.add_argument("--file", default="BCRE REGAL FULL SELF STORAGE DATABASE.xlsx",
                        help="Path to the workbook (default: repo-root BCRE REGAL DB).")
    parser.add_argument("--db", default=COMP_DB_PATH, help="Path to the comp SQLite DB.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Report what would be imported without writing.")
    parser.add_argument("--reset", action="store_true",
                        help="Drop prior [regal: rows before importing.")
    args = parser.parse_args(argv)

    if not os.path.exists(args.file):
        sys.exit(f"Workbook not found: {args.file}")

    # A dry run needs no scratch copy now that the whole import is one
    # transaction: it runs against the real DB and rolls back. The old
    # copy-to-".dryrun" path snapshotted only the main file and not the WAL
    # (so it could read stale), and its os.remove was not in a finally (so a
    # failure orphaned the copy).
    try:
        stats = run_import(args.file, args.db, dry_run=args.dry_run, reset=args.reset)
    except ImportError_ as exc:
        sys.exit(f"Import refused (nothing was written): {exc}")

    print(f"Regal DB import into: {args.db}")
    if stats.get("cleared"):
        print(f"  cleared prior regal rows: {stats['cleared']}")
    if stats["missing_sheets"]:
        print(f"  sheets absent from workbook: {', '.join(stats['missing_sheets'])}")
    for sheet, s in stats["sheets"].items():
        print(f"  {sheet:10s} properties={s['properties']:<5d} vintage={s['vintage']}  "
              f"skipped(no key)={s['skipped_no_key']:<5d} skipped(no NRSF)={s['skipped_no_nrsf']}")
    print(f"  TOTAL properties={stats['properties']} "
          f"skipped_no_key={stats['skipped_no_key']} "
          f"skipped_no_nrsf={stats['skipped_no_nrsf']}")
    print("  unit_mix rows written: 0 (rents are never seeded from this workbook)")
    print("  dry-run (rolled back)" if args.dry_run else "  committed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())

