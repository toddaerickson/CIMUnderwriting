"""
Generates the SS Returns Model as a .xlsx file using openpyxl.

Tabs:
  1. Inputs — purchase price, NRSF, scenario assumptions (yellow = editable)
  2. Base Case — 5-year P&L, exit calc, IRR, MOIC
  3. Bear Case — same structure
  4. Bull Case — same structure
  5. Sensitivity — IRR sensitivity table (price × exit cap)
  6. Max Offer — solved max price and derivation
  7. Sources & Uses — capital stack (model/returns_model.py)
  8. Levered Returns — the sized loan, LP net returns and the assumption
     stamp (model/levered.py). Absent when a deal priced no loan.
  9. Checks — the model error-check register (analysis/checks.py)
"""

import os

import config as cfg
from analysis.valuation import describe_market_cap
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from registry import ScenarioType
from openpyxl.utils import get_column_letter


# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
LABEL_FONT = Font(name="Calibri", size=11)
VALUE_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
PCT_FORMAT = "0.0%"
#: Cap rates carry three decimals. The obsolescence drift is 5–10 bp/yr, so
#: PCT_FORMAT's single decimal would round the derived component away and
#: make the printed build-up fail to add up.
CAP_FORMAT = "0.000%"
CURRENCY_FORMAT = '#,##0'
CURRENCY_FULL = '$#,##0'
MULTIPLE_FORMAT = '0.00"x"'
THIN_BORDER = Border(
    bottom=Side(style="thin", color="999999"),
)


def generate_excel(property_name: str, cim_data, financial_analysis: dict,
                   scenario_results: dict, sensitivity: dict,
                   max_offer: dict, va_results: dict = None,
                   va_max_offer: dict = None, checks: list = None,
                   sources_uses: dict = None, levered: dict = None,
                   debt: dict = None,
                   output_dir: str = ".") -> str:
    """
    Generate the SS Returns Model .xlsx.

    Returns: path to generated file.
    """
    wb = Workbook()
    safe_name = _safe_filename(property_name or "Unknown_Property")

    # Tab 1: Inputs
    _build_inputs_tab(wb.active, cim_data, financial_analysis,
                      scenario_results)
    wb.active.title = "Inputs"

    # Tabs 2-4: Scenario cases (static)
    for scen_name in ScenarioType:
        ws = wb.create_sheet(title=f"{scen_name.title()} Case")
        scen = scenario_results.get(scen_name, {})
        _build_scenario_tab(ws, scen_name, scen, cim_data)

    # Tab 5: Value-Add (if applicable)
    if va_results:
        ws_va = wb.create_sheet(title="Value-Add")
        _build_value_add_tab(ws_va, va_results, va_max_offer or {}, cim_data)

    # Tab 6: Sensitivity
    ws_sens = wb.create_sheet(title="Sensitivity")
    _build_sensitivity_tab(ws_sens, sensitivity)

    # Tab 7: Max Offer
    ws_max = wb.create_sheet(title="Max Offer")
    _build_max_offer_tab(ws_max, max_offer, cim_data)

    # Tab 8: Sources & Uses — what the deal costs and where the money
    # comes from. Placed before Checks so the workbook reads deal →
    # returns → capital → integrity.
    if sources_uses:
        _build_sources_uses_tab(wb.create_sheet(title="Sources & Uses"),
                                sources_uses)

    # Tab 9: Levered Returns — the second lens. After Sources & Uses
    # because it reads off that capital stack, and before Checks so the
    # workbook keeps reading deal → returns → capital → levered →
    # integrity. Absent on a deal that priced no loan.
    if levered:
        _build_levered_tab(wb.create_sheet(title="Levered Returns"),
                           levered, debt or {}, scenario_results)

    # Tab 10: Checks — the whole register, not just the findings, so the
    # workbook says what was verified as well as what failed.
    if checks:
        _build_checks_tab(wb.create_sheet(title="Checks"), checks)

    filename = f"SS_Returns_Model_{safe_name}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath


# ── Tab Builders ────────────────────────────────────────────────────

def _dotted(d: dict, key):
    """Read `a.b` out of a nested result dict, missing levels → None.
    Lets a comparison table name a component of `exit_cap_detail` without
    the table's rows growing a second shape. A callable key is applied to
    the whole scenario dict, for the one row that is rendered rather than
    read (the market-cap provenance sentence)."""
    if callable(key):
        return key(d)
    cur = d
    for part in str(key).split("."):
        if not isinstance(cur, dict):
            return None
        cur = cur.get(part)
    return cur


def _exit_cap_rows(scen: dict) -> list:
    """(label, value, format) rows that make ONE scenario's exit cap
    retraceable — the anchor it started from, the two modifiers, and, when
    the exit ≥ entry floor fired, the derived rate the applied one
    replaced. Falls back to the bare applied rate for a stored run that
    predates the derivation."""
    d = scen.get("exit_cap_detail") or {}
    if not d:
        return [("Exit Cap Rate", scen.get("exit_cap"), PCT_FORMAT)]
    rows = [
        ("Market Cap Rate", d.get("market_cap"), CAP_FORMAT),
        ("  Scenario Spread (bp)", d.get("scenario_spread_bps"), '#,##0.0'),
        (f"  Obsolescence Drift ({d.get('drift_bps_per_year')} bp/yr × "
         f"{d.get('hold_years')} yrs)", d.get("drift_total_bps"), '#,##0.0'),
    ]
    if scen.get("exit_cap_coerced"):
        rows.append(("  Derived Exit Cap (before floor)",
                     scen.get("requested_exit_cap"), CAP_FORMAT))
        rows.append(("Exit Cap Rate (raised to entry cap)",
                     scen.get("exit_cap"), CAP_FORMAT))
    else:
        rows.append(("Exit Cap Rate", scen.get("exit_cap"), CAP_FORMAT))
    return rows


def _build_inputs_tab(ws, cim_data, fin, scenario_results=None):
    """Build the Inputs tab with editable assumption cells."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    row = 1
    row = _write_section_header(ws, row, "Property Information", cols=2)

    inputs = [
        ("Property Name", cim_data.property_name or "TBD", None),
        ("Address", cim_data.address or "TBD", None),
        ("City, State", f"{cim_data.city or 'TBD'}, {cim_data.state or 'TBD'}", None),
        ("Asking Price", cim_data.asking_price, CURRENCY_FULL),
        ("NRSF", cim_data.nrsf, CURRENCY_FORMAT),
        ("Total Units", cim_data.total_units, CURRENCY_FORMAT),
        ("Physical Occupancy", cim_data.physical_occupancy, PCT_FORMAT),
        ("Economic Occupancy", cim_data.economic_occupancy, PCT_FORMAT),
        ("CC %", cim_data.cc_pct, PCT_FORMAT),
        ("Year Built", cim_data.year_built, None),
        ("Price / SF", cim_data.price_per_sf, '$#,##0.00'),
    ]

    for label, val, fmt in inputs:
        row = _write_input_row(ws, row, label, val, fmt, editable=True)

    row += 1
    row = _write_section_header(ws, row, "Financial Summary", cols=2)

    adj_noi = fin.get("adjusted_ttm_noi", {})
    fin_inputs = [
        ("TTM NOI (CIM)", cim_data.ttm_noi, CURRENCY_FULL),
        ("Analyst-Adjusted TTM NOI", adj_noi.get("analyst_adjusted_noi"), CURRENCY_FULL),
        ("TTM Total Revenue", cim_data.ttm_total_revenue, CURRENCY_FULL),
        ("TTM Total Expenses", cim_data.ttm_total_expenses, CURRENCY_FULL),
        ("TTM GPR", cim_data.ttm_gpr, CURRENCY_FULL),
        ("TTM EGR", cim_data.ttm_egr, CURRENCY_FULL),
        ("Other Income", cim_data.other_income, CURRENCY_FULL),
    ]

    for label, val, fmt in fin_inputs:
        row = _write_input_row(ws, row, label, val, fmt)

    row += 1
    row = _write_section_header(ws, row, "Scenario Assumptions", cols=2)

    # The params this run ACTUALLY used, read off the scenario results —
    # not a second read of config.SCENARIO_DEFAULTS. A per-deal override
    # or a ConfigOverride row makes those two different documents, and
    # this is the tab a reader treats as the record of what was assumed.
    # Config is the fallback only when no scenario ran at all.
    for scen_name in ScenarioType:
        scen = (scenario_results or {}).get(scen_name) or {}
        params = scen.get("params") or cfg.SCENARIO_DEFAULTS[scen_name]
        ws.cell(row=row, column=1, value=f"── {scen_name.title()} Case ──").font = BOLD_FONT
        row += 1
        for key, val in params.items():
            label = key.replace("_", " ").title()
            fmt = PCT_FORMAT if isinstance(val, float) and val < 1 else None
            row = _write_input_row(ws, row, f"  {label}", val, fmt, editable=True)
        row += 1

    row = _write_exit_cap_derivation(ws, row, scenario_results)


def _write_exit_cap_derivation(ws, row, scenario_results):
    """The exit cap build-up: market anchor, then each scenario's parts.

    The exit cap was a per-scenario constant an analyst could type; it is
    now derived, so an applied rate printed on its own would be less
    auditable than what it replaced. This prints the formula's terms.
    Deliberately NOT editable-yellow: editing a derived cell in the
    workbook detaches it from the anchor without recomputing anything.
    """
    detail = {}
    for scen in (scenario_results or {}).values():
        if isinstance(scen, dict) and scen.get("exit_cap_detail"):
            detail = scen["exit_cap_detail"]
            break
    if not detail:
        return row

    row += 1
    row = _write_section_header(ws, row, "Exit Cap Derivation", cols=2)
    band = detail.get("age_band") or "—"
    if detail.get("age_band_known") is False:
        band = f"{band} (year built unknown — fallback)"
    for label, val, fmt in (
            ("Asset Class", detail.get("asset_class"), None),
            ("Age Band", band, None),
            ("Market Cap Rate", detail.get("market_cap"), CAP_FORMAT),
            ("Market Cap Source", describe_market_cap(detail), None),
            # The table rate is reported even when an analyst overrode it,
            # and the as-of dates THAT, not the applied rate — labelled so
            # an override cannot read as carrying a table vintage.
            ("Table Market Cap", detail.get("table_market_cap"), CAP_FORMAT),
            ("Table As Of", detail.get("as_of"), None)):
        row = _write_input_row(ws, row, label, val, fmt)

    row += 1
    ws.cell(row=row, column=1,
            value="Exit cap = market cap + scenario spread "
                  "+ obsolescence drift × hold").font = LABEL_FONT
    row += 1
    for scen_name in ScenarioType:
        scen = (scenario_results or {}).get(scen_name) or {}
        d = scen.get("exit_cap_detail") or {}
        if not d:
            continue
        row = _write_input_row(
            ws, row, f"  {scen_name.title()} Spread (bp)",
            d.get("scenario_spread_bps"), '#,##0.0')
        row = _write_input_row(
            ws, row, f"  {scen_name.title()} Drift ({d.get('drift_bps_per_year')} "
                     f"bp/yr × {d.get('hold_years')} yrs)",
            d.get("drift_total_bps"), '#,##0.0')
        row = _write_input_row(
            ws, row, f"  {scen_name.title()} Derived Exit Cap",
            scen.get("requested_exit_cap"), "0.000%")
        if scen.get("exit_cap_coerced"):
            row = _write_input_row(
                ws, row, f"  {scen_name.title()} Applied (raised to entry cap)",
                scen.get("exit_cap"), "0.000%")
    return row


def _build_scenario_tab(ws, scen_name: str, scen: dict, cim_data):
    """Build a single scenario tab with the hold-period P&L."""
    noi_proj = scen.get("noi_projection", [])
    years = len(noi_proj) or scen.get("hold_years") or cfg.DEFAULT_HOLD_YEARS

    ws.column_dimensions["A"].width = 28
    for i in range(2, 2 + max(years, 7)):
        ws.column_dimensions[get_column_letter(i)].width = 16

    row = 1
    row = _write_section_header(
        ws, row,
        f"{scen_name.title()} Case — {years}-Year Unlevered Returns",
        cols=max(years + 2, 7))

    # Key metrics. Basis is cost-inclusive: the acquisition line below is
    # already inside Total Basis, shown so the build-up is traceable.
    metrics = [
        ("Total Basis", scen.get("total_basis"), CURRENCY_FULL),
        ("Asking Price", scen.get("asking_price"), CURRENCY_FULL),
        ("CapEx", scen.get("capex"), CURRENCY_FULL),
        ("Acquisition Closing Costs", scen.get("acquisition_cost"), CURRENCY_FULL),
        ("Hold Period (yrs)", years, None),
        ("Entry Cap Rate", scen.get("entry_cap"), PCT_FORMAT),
        # CAP_FORMAT, not PCT_FORMAT: the Exit & Returns block below prints
        # this same value to three places, and one sheet showing "6.6%" here
        # and "6.625%" there reads as two different caps.
        ("Exit Cap Rate", scen.get("exit_cap"), CAP_FORMAT),
    ]
    for label, val, fmt in metrics:
        row = _write_input_row(ws, row, label, val, fmt)

    row += 1

    # Year headers
    rev_proj = scen.get("revenue_projection", [])
    exp_proj = scen.get("expense_projection", [])

    ws.cell(row=row, column=1, value="").font = BOLD_FONT
    for yr in range(years):
        col = yr + 2
        cell = ws.cell(row=row, column=col, value=f"Year {yr + 1}")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Revenue projection
    if rev_proj:
        ws.cell(row=row, column=1, value="Revenue").font = BOLD_FONT
        for i, val in enumerate(rev_proj[:years]):
            ws.cell(row=row, column=i + 2, value=val).number_format = CURRENCY_FULL
        row += 1

    # Expense projection
    if exp_proj:
        ws.cell(row=row, column=1, value="Expenses").font = BOLD_FONT
        for i, val in enumerate(exp_proj[:years]):
            ws.cell(row=row, column=i + 2, value=val).number_format = CURRENCY_FULL
        row += 1

    # NOI projection
    ws.cell(row=row, column=1, value="Net Operating Income").font = BOLD_FONT
    for i, val in enumerate(noi_proj[:years]):
        cell = ws.cell(row=row, column=i + 2, value=val)
        cell.number_format = CURRENCY_FULL
        cell.font = BOLD_FONT
    row += 1

    # NOI per SF
    noi_per_sf = scen.get("noi_per_sf", [])
    if noi_per_sf:
        ws.cell(row=row, column=1, value="NOI / SF").font = LABEL_FONT
        for i, val in enumerate(noi_per_sf[:years]):
            ws.cell(row=row, column=i + 2, value=val).number_format = '$#,##0.00'
        row += 1

    row += 1

    # Exit & Returns
    row = _write_section_header(ws, row, "Exit & Returns", cols=2)
    exit_items = [
        (f"Year {years} NOI", noi_proj[-1] if noi_proj else None, CURRENCY_FULL),
        *_exit_cap_rows(scen),
        ("Exit Value (gross)", scen.get("exit_value"), CURRENCY_FULL),
        ("Disposition Costs", scen.get("disposition_cost"), CURRENCY_FULL),
        ("Net Sale Proceeds", scen.get("net_exit_proceeds"), CURRENCY_FULL),
    ]
    for label, val, fmt in exit_items:
        row = _write_input_row(ws, row, label, val, fmt)

    row += 1
    return_items = [
        (f"{years}-Year Unlevered IRR", scen.get("irr"), PCT_FORMAT),
        (f"{years}-Year MOIC", scen.get("moic"), MULTIPLE_FORMAT),
        ("Year 1 Yield on Cost", scen.get("yield_on_cost"), PCT_FORMAT),
    ]
    for label, val, fmt in return_items:
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_a.font = BOLD_FONT
        cell_b = ws.cell(row=row, column=2, value=val)
        if fmt and val is not None:
            cell_b.number_format = fmt
        cell_b.font = BOLD_FONT
        row += 1

    row += 1

    # Cash flows
    row = _write_section_header(ws, row, "Cash Flow Summary", cols=7)
    cfs = scen.get("cash_flows", [])
    ws.cell(row=row, column=1, value="").font = BOLD_FONT
    cf_labels = ["Year 0 (Invest)"] + [f"Year {i+1}" for i in range(len(cfs) - 1)]
    for i, label in enumerate(cf_labels):
        cell = ws.cell(row=row, column=i + 1 + 1, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    # Extend column 1 header
    cell = ws.cell(row=row, column=1, value="")
    cell.fill = HEADER_FILL
    row += 1

    ws.cell(row=row, column=1, value="Cash Flow").font = BOLD_FONT
    for i, cf in enumerate(cfs):
        ws.cell(row=row, column=i + 2, value=cf).number_format = CURRENCY_FULL
    row += 1


def _build_sensitivity_tab(ws, sensitivity: dict):
    """Build IRR sensitivity table."""
    ws.column_dimensions["A"].width = 18
    for i in range(2, 15):
        ws.column_dimensions[get_column_letter(i)].width = 12

    row = 1
    row = _write_section_header(ws, row, "IRR Sensitivity: Purchase Price vs Exit Cap Rate", cols=10)
    row += 1

    price_labels = sensitivity.get("price_labels", [])
    price_values = sensitivity.get("price_values", [])
    cap_labels = sensitivity.get("cap_labels", [])
    grid = sensitivity.get("irr_grid", [])

    if not grid:
        ws.cell(row=row, column=1, value="Insufficient data for sensitivity analysis.")
        return

    # Header row: exit cap labels
    ws.cell(row=row, column=1, value="Price \\ Exit Cap").font = BOLD_FONT
    for j, cap_label in enumerate(cap_labels):
        cell = ws.cell(row=row, column=j + 2, value=cap_label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Data rows
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for i, price_label in enumerate(price_labels):
        price_val = price_values[i] if i < len(price_values) else 0
        ws.cell(row=row, column=1, value=f"{price_label} (${price_val:,.0f})").font = BOLD_FONT

        if i < len(grid):
            for j, irr in enumerate(grid[i]):
                cell = ws.cell(row=row, column=j + 2)
                if irr is not None:
                    cell.value = irr
                    cell.number_format = PCT_FORMAT
                    # Color code
                    if irr >= 0.12:
                        cell.fill = green_fill
                    elif irr >= 0.10:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill
                else:
                    cell.value = "N/A"
                cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Legend:").font = BOLD_FONT
    row += 1
    c = ws.cell(row=row, column=1, value="  ≥ 12% IRR")
    c.fill = green_fill
    row += 1
    c = ws.cell(row=row, column=1, value="  10-12% IRR")
    c.fill = yellow_fill
    row += 1
    c = ws.cell(row=row, column=1, value="  < 10% IRR")
    c.fill = red_fill


def _build_max_offer_tab(ws, max_offer: dict, cim_data):
    """Build Max Offer derivation tab."""
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    row = 1
    row = _write_section_header(ws, row, "Maximum Offer Price Derivation", cols=2)
    row += 1

    items = [
        ("Target IRR", max_offer.get("target_irr"), PCT_FORMAT),
        ("Solver Converged", "Yes" if max_offer.get("converged") else "No", None),
        ("Iterations", max_offer.get("iterations"), None),
        ("", None, None),
        ("Maximum Purchase Price", max_offer.get("max_price"), CURRENCY_FULL),
        ("Implied Entry Cap Rate", max_offer.get("implied_entry_cap"), PCT_FORMAT),
        ("CapEx Budget", max_offer.get("capex"), CURRENCY_FULL),
        ("Acquisition Closing Costs", max_offer.get("acquisition_cost"), CURRENCY_FULL),
        ("Total Basis at Max Price", max_offer.get("total_basis"), CURRENCY_FULL),
        ("Achieved IRR", max_offer.get("achieved_irr"), PCT_FORMAT),
    ]

    for label, val, fmt in items:
        if not label:
            row += 1
            continue
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_a.font = BOLD_FONT if "Maximum" in label or "Achieved" in label else LABEL_FONT
        cell_b = ws.cell(row=row, column=2, value=val)
        if fmt and val is not None and not isinstance(val, str):
            cell_b.number_format = fmt
        cell_b.font = BOLD_FONT if "Maximum" in label or "Achieved" in label else VALUE_FONT
        row += 1

    # Comparison to asking
    if cim_data.asking_price and max_offer.get("max_price"):
        row += 1
        asking = cim_data.asking_price
        mp = max_offer["max_price"]
        discount = (asking - mp) / asking if asking else 0

        row = _write_section_header(ws, row, "Comparison to Asking", cols=2)
        comp_items = [
            ("Asking Price", asking, CURRENCY_FULL),
            ("Max Offer Price", mp, CURRENCY_FULL),
            ("Discount to Asking", discount, PCT_FORMAT),
            ("Dollar Difference", asking - mp, CURRENCY_FULL),
        ]
        for label, val, fmt in comp_items:
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            cell = ws.cell(row=row, column=2, value=val)
            if fmt and val is not None:
                cell.number_format = fmt
            row += 1


def _build_value_add_tab(ws, va_results: dict, va_max_offer: dict, cim_data):
    """Build Value-Add tab with annual summary across all three scenarios."""
    ws.column_dimensions["A"].width = 28
    for i in range(2, 9):
        ws.column_dimensions[get_column_letter(i)].width = 16

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    row = 1
    row = _write_section_header(ws, row, "Value-Add Model — Scenario Comparison", cols=4)
    row += 1

    # Key assumptions
    base = va_results.get(ScenarioType.BASE, {})
    va_hold = (base.get("hold_years") or len(base.get("annual_noi") or [])
               or cfg.DEFAULT_HOLD_YEARS)
    row = _write_section_header(ws, row, "Deal Overview", cols=4)
    overview = [
        ("Asking Price", base.get("asking_price"), CURRENCY_FULL),
        ("CapEx", base.get("capex"), CURRENCY_FULL),
        ("Acquisition Closing Costs", base.get("acquisition_cost"), CURRENCY_FULL),
        ("Total Basis", base.get("total_basis"), CURRENCY_FULL),
        ("Hold Period (yrs)", va_hold, None),
        ("Current Occupancy", base.get("current_occupancy"), PCT_FORMAT),
        ("In-Place Rent/SF/Mo", base.get("in_place_rent_psf"), '$#,##0.00'),
        ("Market Rent/SF/Mo", base.get("market_rent_psf"), '$#,##0.00'),
    ]
    for label, val, fmt in overview:
        row = _write_input_row(ws, row, label, val, fmt)
    row += 1

    # Scenario comparison header
    ws.cell(row=row, column=1, value="").font = BOLD_FONT
    for j, scen_name in enumerate(("Bear", "Base", "Bull")):
        cell = ws.cell(row=row, column=j + 2, value=scen_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Scenario metrics
    metric_rows = [
        ("Target Occupancy", "target_occupancy", PCT_FORMAT),
        ("Months to Stabilize", "months_to_stabilize", '#,##0'),
        ("Target Rent/SF/Mo", "target_rent_psf", '$#,##0.00'),
        ("Stabilized NOI", "stabilized_noi", CURRENCY_FULL),
        ("Entry Cap Rate", "entry_cap", PCT_FORMAT),
        # The VA engine used to read its own exit cap off a second config
        # triple 100 bp tighter than the static one. It now shares the
        # resolver, so the anchor printed here is the same anchor the
        # Inputs tab prints — showing it is what makes that checkable.
        ("Market Cap Rate", "exit_cap_detail.market_cap", CAP_FORMAT),
        # Provenance here too, not just on the Inputs tab and the memo.
        # Value-add IS the target deal profile, so this is the surface an
        # analyst reads most — and an override that shows as a bare rate
        # is the same missing trace the static side was fixed for.
        ("  Market Cap Source",
         lambda s: describe_market_cap(s.get("exit_cap_detail") or {}), None),
        ("  Scenario Spread (bp)",
         "exit_cap_detail.scenario_spread_bps", '#,##0.0'),
        ("  Obsolescence Drift (bp)",
         "exit_cap_detail.drift_total_bps", '#,##0.0'),
        ("Exit Cap Rate", "exit_cap", CAP_FORMAT),
        ("Exit Value (gross)", "exit_value", CURRENCY_FULL),
        ("Disposition Costs", "disposition_cost", CURRENCY_FULL),
        ("Net Sale Proceeds", "net_exit_proceeds", CURRENCY_FULL),
        ("", None, None),
        (f"{va_hold}-Year Unlevered IRR", "irr", PCT_FORMAT),
        (f"{va_hold}-Year MOIC", "moic", MULTIPLE_FORMAT),
        ("Stabilized Yield/Cost", "yield_on_cost", PCT_FORMAT),
        ("Development Spread", "development_spread", PCT_FORMAT),
    ]

    for label, key, fmt in metric_rows:
        if not label:
            row += 1
            continue
        is_return = key in ("irr", "moic", "yield_on_cost", "development_spread")
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT if is_return else LABEL_FONT
        for j, scen_name in enumerate(ScenarioType):
            scen = va_results.get(scen_name, {})
            val = _dotted(scen, key)
            cell = ws.cell(row=row, column=j + 2, value=val)
            if fmt and val is not None:
                cell.number_format = fmt
            cell.font = BOLD_FONT if is_return else VALUE_FONT
            cell.alignment = Alignment(horizontal="center")
        row += 1

    row += 1

    # Annual NOI projection for base case
    row = _write_section_header(ws, row, "Base Case — Annual Projection", cols=7)
    annual_noi = base.get("annual_noi", [])
    annual_rev = base.get("annual_revenue", [])
    annual_exp = base.get("annual_expenses", [])
    # Show the whole hold — this used to cap at 5, so a 10-year VA run
    # silently lost years 6-10 off the projection table.
    years = len(annual_noi)

    ws.cell(row=row, column=1, value="").font = BOLD_FONT
    for yr in range(years):
        cell = ws.cell(row=row, column=yr + 2, value=f"Year {yr + 1}")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    if annual_rev:
        ws.cell(row=row, column=1, value="Revenue").font = LABEL_FONT
        for i, val in enumerate(annual_rev[:years]):
            ws.cell(row=row, column=i + 2, value=val).number_format = CURRENCY_FULL
        row += 1

    if annual_exp:
        ws.cell(row=row, column=1, value="Expenses").font = LABEL_FONT
        for i, val in enumerate(annual_exp[:years]):
            ws.cell(row=row, column=i + 2, value=val).number_format = CURRENCY_FULL
        row += 1

    ws.cell(row=row, column=1, value="Net Operating Income").font = BOLD_FONT
    for i, val in enumerate(annual_noi[:years]):
        cell = ws.cell(row=row, column=i + 2, value=val)
        cell.number_format = CURRENCY_FULL
        cell.font = BOLD_FONT
    row += 1

    noi_per_sf = base.get("noi_per_sf", [])
    if noi_per_sf:
        ws.cell(row=row, column=1, value="NOI / SF").font = LABEL_FONT
        for i, val in enumerate(noi_per_sf[:years]):
            ws.cell(row=row, column=i + 2, value=val).number_format = '$#,##0.00'
        row += 1

    row += 1

    # Cash flows
    row = _write_section_header(ws, row, "Base Case — Cash Flow Summary", cols=7)
    cfs = base.get("cash_flows", [])
    ws.cell(row=row, column=1, value="").font = BOLD_FONT
    cf_labels = ["Year 0 (Invest)"] + [f"Year {i+1}" for i in range(len(cfs) - 1)]
    for i, label in enumerate(cf_labels):
        cell = ws.cell(row=row, column=i + 2, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    cell = ws.cell(row=row, column=1, value="")
    cell.fill = HEADER_FILL
    row += 1

    ws.cell(row=row, column=1, value="Cash Flow").font = BOLD_FONT
    for i, cf in enumerate(cfs):
        ws.cell(row=row, column=i + 2, value=cf).number_format = CURRENCY_FULL
    row += 1

    row += 1

    # VA Max offer
    if va_max_offer and va_max_offer.get("max_price"):
        row = _write_section_header(ws, row, "Value-Add Max Offer Price", cols=2)
        va_items = [
            ("Max Price (10% VA IRR)", va_max_offer.get("max_price"), CURRENCY_FULL),
            ("Implied Entry Cap", va_max_offer.get("implied_entry_cap"), PCT_FORMAT),
            ("Achieved IRR", va_max_offer.get("achieved_irr"), PCT_FORMAT),
        ]
        for label, val, fmt in va_items:
            cell_a = ws.cell(row=row, column=1, value=label)
            cell_a.font = BOLD_FONT
            cell_b = ws.cell(row=row, column=2, value=val)
            if fmt and val is not None:
                cell_b.number_format = fmt
            cell_b.font = BOLD_FONT
            row += 1


def _build_sources_uses_tab(ws, su: dict):
    """Capital stack. Every line is read off `build_sources_uses` — the
    percentages are not recomputed here, so the sheet cannot disagree with
    the returns model about what the deal costs."""
    row = _write_section_header(ws, 1, "Uses of Funds", cols=3)
    total_uses = su.get("total_uses") or 0
    for line in su.get("uses") or []:
        amount = line.get("amount") or 0
        ws.cell(row=row, column=1, value=line.get("label")).font = LABEL_FONT
        cell = ws.cell(row=row, column=2, value=amount)
        cell.font = VALUE_FONT
        cell.number_format = CURRENCY_FULL
        share = ws.cell(row=row, column=3,
                        value=(amount / total_uses) if total_uses else None)
        share.font = VALUE_FONT
        share.number_format = PCT_FORMAT
        row += 1
    row = _write_total_row(ws, row, "Total Uses", total_uses)

    row += 1
    row = _write_section_header(ws, row, "Sources of Funds", cols=3)
    total_sources = su.get("total_sources") or 0
    for line in su.get("sources") or []:
        amount = line.get("amount") or 0
        ws.cell(row=row, column=1, value=line.get("label")).font = LABEL_FONT
        cell = ws.cell(row=row, column=2, value=amount)
        cell.font = VALUE_FONT
        cell.number_format = CURRENCY_FULL
        share = ws.cell(row=row, column=3,
                        value=(amount / total_sources) if total_sources else None)
        share.font = VALUE_FONT
        share.number_format = PCT_FORMAT
        row += 1
    row = _write_total_row(ws, row, "Total Sources", total_sources)

    row += 1
    row = _write_input_row(ws, row, "Total Equity", su.get("total_equity"),
                           CURRENCY_FULL)
    row = _write_input_row(ws, row, "Loan-to-Cost", su.get("ltv"), PCT_FORMAT)
    # Say the tie held (or did not) on the sheet itself. A capital stack
    # that silently disagrees with the DCF is the failure this block
    # exists to prevent, so it should not require opening another tab.
    row = _write_input_row(
        ws, row, "Sources − Uses",
        "In balance" if su.get("balanced")
        else f"OUT OF BALANCE by ${abs(su.get('delta') or 0):,.2f}")

    for col, width in ((1, 34), (2, 18), (3, 12)):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_total_row(ws, row: int, label: str, value) -> int:
    """Bold, ruled total line under a Sources/Uses block."""
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = BOLD_FONT
    label_cell.border = THIN_BORDER
    cell = ws.cell(row=row, column=2, value=value)
    cell.font = BOLD_FONT
    cell.number_format = CURRENCY_FULL
    cell.border = THIN_BORDER
    ws.cell(row=row, column=3).border = THIN_BORDER
    return row + 1


def _build_levered_tab(ws, levered: dict, debt: dict, scenario_results: dict):
    """The levered second lens (item E3b).

    Every figure is read off the persisted `levered` / `debt` payload.
    Nothing here re-runs the waterfall, so the workbook cannot disagree
    with the memo or the results page about the LP's net return.

    The assumption stamp is at the BOTTOM and is not optional: five of
    those inputs are open LPA questions and each one moves the IRR above.
    """
    base = levered.get("base") or {}
    terms = debt.get("terms") or {}
    scenarios = [s for s in ScenarioType]

    row = _write_section_header(ws, 1, "Senior Loan (sized once, base case)",
                                cols=2)
    from model.debt import binding_constraint_label, displayed_rate
    for label, value, fmt in (
            ("Loan Amount", debt.get("loan"), CURRENCY_FULL),
            ("Bound By", binding_constraint_label(debt), None),
            ("All-In Rate", displayed_rate(terms), CAP_FORMAT),
            ("Amortization (yrs)", terms.get("amort_years"), None),
            ("Interest-Only (mos)", terms.get("io_months"), None),
            ("Loan Term (yrs)", terms.get("term_years"), None),
            ("LTV", debt.get("ltv"), PCT_FORMAT),
            ("Year-1 DSCR", debt.get("dscr_year_1"), MULTIPLE_FORMAT),
            ("Debt Yield", debt.get("debt_yield"), PCT_FORMAT),
            ("Origination Fee", debt.get("origination_fee"), CURRENCY_FULL),
            ("Exit Fee", debt.get("exit_fee"), CURRENCY_FULL),
            ("Payoff Balance at Exit", debt.get("payoff_balance"),
             CURRENCY_FULL),
            ("Matures Before Exit",
             "YES — no refinancing modeled" if debt.get("matures_before_exit")
             else "No", None)):
        row = _write_input_row(ws, row, label, value, fmt)

    row += 1
    row = _write_section_header(ws, row, "LP Net Returns by Scenario",
                                cols=len(scenarios) + 1)
    for col, scen in enumerate(scenarios, start=2):
        cell = ws.cell(row=row, column=col, value=scen.title())
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
    ws.cell(row=row, column=1).border = THIN_BORDER
    row += 1
    for label, key, fmt in (
            ("LP Net IRR", "lp_net_irr", PCT_FORMAT),
            ("LP MOIC", "lp_moic", MULTIPLE_FORMAT),
            ("GP Promote", "gp_promote", CURRENCY_FULL),
            ("AM Fee (total)", "am_fee_total", CURRENCY_FULL),
            ("Capital Called After Close", "capital_calls_total",
             CURRENCY_FULL),
            ("Equity Required", "total_equity", CURRENCY_FULL)):
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        for col, scen in enumerate(scenarios, start=2):
            value = (levered.get(scen) or {}).get(key)
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = VALUE_FONT
            if value is not None:
                cell.number_format = fmt
        row += 1
    # Leverage is allowed to be dilutive; the sheet says so rather than
    # leaving a reader to treat it as an arithmetic error.
    dilutive = [scen.title() for scen in scenarios
                if (levered.get(scen) or {}).get("lp_net_irr") is not None
                and ((scenario_results or {}).get(scen) or {}).get("irr")
                is not None
                and (levered[scen]["lp_net_irr"]
                     < scenario_results[scen]["irr"])]
    if dilutive:
        row = _write_input_row(
            ws, row, "Leverage Is Dilutive In",
            f"{', '.join(dilutive)} — LP net below the unlevered IRR "
            f"(loan constant above yield on cost)")

    years = base.get("years") or []
    if years:
        row += 1
        row = _write_section_header(ws, row, "Base-Case Equity Cash Flow",
                                    cols=8)
        headers = ("Year", "NOI", "Debt Service", "AM Fee", "Levered CF",
                   "Distribution", "Capital Call", "DSCR")
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=title)
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
        row += 1
        for yr in years:
            values = (yr.get("year"), yr.get("noi"), yr.get("debt_service"),
                      yr.get("am_fee"), yr.get("levered_cf"),
                      yr.get("distribution"), yr.get("capital_call"),
                      yr.get("dscr"))
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = VALUE_FONT
                if col == 8:
                    cell.number_format = MULTIPLE_FORMAT
                elif col > 1:
                    cell.number_format = CURRENCY_FULL
            row += 1

    stamp = base.get("assumption_stamp") or []
    if stamp:
        row += 1
        row = _write_section_header(ws, row, "Levered Assumptions", cols=2)
        for entry in stamp:
            row = _write_input_row(ws, row, entry.get("question") or "",
                                   entry.get("label") or "")

    for col, width in ((1, 34), (2, 22), (3, 16), (4, 14), (5, 16),
                       (6, 16), (7, 14), (8, 10)):
        ws.column_dimensions[get_column_letter(col)].width = width


def _build_checks_tab(ws, checks: list):
    """Model error-check register. Findings sort to the top — a register
    read top-down should hit the problems first — but every check is
    listed, including the ones that were not testable, because "we did not
    look" and "we looked and it was fine" are different claims."""
    STATUS_FILL = {
        "fail": PatternFill(start_color="FFD9D9", end_color="FFD9D9",
                            fill_type="solid"),
        "pass": PatternFill(start_color="DDF3E4", end_color="DDF3E4",
                            fill_type="solid"),
    }
    ORDER = {"fail": 0, "pass": 1, "skipped": 2}

    row = _write_section_header(ws, 1, "Model Error-Check Register", cols=5)
    for col, title in enumerate(
            ("Check", "Severity", "Status", "Finding", "Source"), start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
    row += 1

    for c in sorted(checks, key=lambda c: (ORDER.get(c.get("status"), 3),
                                           c.get("label") or "")):
        status = c.get("status") or ""
        values = (c.get("label") or c.get("id") or "",
                  (c.get("severity") or "").title(), status.upper(),
                  c.get("message") or "", c.get("source") or "")
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = VALUE_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 4))
            if status in STATUS_FILL:
                cell.fill = STATUS_FILL[status]
        row += 1

    for col, width in ((1, 30), (2, 12), (3, 10), (4, 90), (5, 42)):
        ws.column_dimensions[get_column_letter(col)].width = width


# ── Helpers ─────────────────────────────────────────────────────────

def _write_section_header(ws, row: int, title: str, cols: int = 2) -> int:
    """Write a section header row and return next row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.cell(row=row, column=1, value=title)
    return row + 1


def _write_input_row(ws, row: int, label: str, value, fmt=None, editable=False) -> int:
    """Write a label/value row. Yellow fill if editable."""
    ws.cell(row=row, column=1, value=label).font = LABEL_FONT
    cell = ws.cell(row=row, column=2, value=value)
    cell.font = VALUE_FONT
    if fmt and value is not None and not isinstance(value, str):
        cell.number_format = fmt
    if editable:
        cell.fill = INPUT_FILL
    return row + 1


def _safe_filename(name: str) -> str:
    return "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in name).strip().replace(" ", "_")
