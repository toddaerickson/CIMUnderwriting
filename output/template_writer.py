"""
Template Writer — populate the XLSM underwriting template with CIM data
extracted by the analysis pipeline.

Writes only to INPUT cells (not formulas). The user opens the .xlsm in
Excel and formulas recalculate automatically.

Place your .xlsm template in the project root and set TEMPLATE_FILENAME below
(or override via the UW_TEMPLATE_PATH environment variable).

## The template never decides a value (item E3b)

Every number written below reads from the run's resolved assumption set
(config + ConfigOverride + per-deal overrides) or from the run's computed
results. It used to assert its own: 6.5% debt at 360-month amortization
against a model running 6.25% over 25 years, a 6%-GP / 20%-promote
waterfall read from environment variables, a flat 3% growth ladder, a
0.10 stabilized vacancy against a 0.88 stabilized-occupancy assumption.
Two deliverables quoting different terms on the same deal is the failure
mode the transparency audit raised, and after item E3a wired levered
returns on by default it shipped on every deal.

`tests/test_template_writer.py::test_no_numeric_literals_in_write_paths`
enforces this with an AST walk, not by inspection: no numeric literal may
appear on the VALUE side of a cell write. The walk FOLLOWS LOCAL NAMES,
so parking a literal in a variable one line above the write does not
evade it — that hole was found in review and the gate now has its own
test proving it fails when it should. Structural constants — a column
index, an at-closing zero, the template's own year-1 convention — are
named once below, which is what makes them reviewable.

## What does not reconcile — SETTLED, and disclosed in the workbook

The XLSM computes its own returns from these inputs, and two of its
conventions are not ours. Neither is reachable from an input cell, so
the writer stamps them into the workbook itself
(`_write_divergence_disclosures`) as well as here:

1. **The pref is an IRR hurdle** (H248 "IRR Hurdle"; the hurdle rows are
   H249-H252). `model.waterfall` runs an ACCRUAL account on
   contributed/unreturned capital. Same 8%, different construction, so
   the promote dollars differ. Writing `pref_rate` into the hurdle rows
   makes the two agree on the RATE, which is as far as an input cell reaches. A
   future edit of the TEMPLATE could swap the hurdle formula for an
   accrual — that is an XLSM edit, not a writer change; until someone
   makes it, the divergence is permanent and disclosed.
2. **The AM fee is charged on LP equity** (H245 = `K61*G245/12`, and
   K61 is LP equity). `config.AM_FEE_BASE` is `invested_equity` —
   GP + LP — so at a 10% GP co-invest the workbook's fee runs ~10%
   light. The dropdown has no invested-equity option. Grossing the rate
   up to 1.11% would make the dollars tie while printing a fee rate the
   fund does not charge, so the true rate is written and the gap is
   disclosed — REAFFIRMED by the operator 2026-08-10 when this residue
   was settled. (The gross-up could not even tie in general: the model's
   fee base rolls forward on a capital call — `model/levered.py` — where
   the workbook's K61 is fixed at close, so a single compensated rate
   reproduces the model's dollars only on deals that never call.)

Settled 2026-08-10; until then both lines read "item T's to reconcile"
and T closed without them. Do not "fix" either by editing a value into
agreement — that trades a visible discrepancy for a hidden one.
"""

import logging
import os
import shutil
from datetime import datetime

import openpyxl

import config as cfg
from analysis.financials import resolve_mgmt_fee_target
from output import safe_filename
from model.debt import MONTHS_PER_YEAR
from registry import ScenarioType

logger = logging.getLogger(__name__)

# Path to the blank template (.xlsm with macros)
# Override via environment variable or place your template in the project root.
TEMPLATE_FILENAME = "template_uw.xlsm"
TEMPLATE_PATH = os.environ.get(
    "UW_TEMPLATE_PATH",
    os.path.join(os.path.dirname(os.path.dirname(__file__)), TEMPLATE_FILENAME),
)

# ── Template version: Self-Storage-Acquisition-Model v1.3 ────────────
#
# **Every address below was re-derived from v1.3 by reading the file**
# (2026-08-11), not carried over. The map this replaced was written
# against v1.2 and does not survive the upgrade in either axis:
#
# - **Rows moved.** The waterfall sits ~10 rows higher (pref H258 →
#   H249), the sale-month cell moved D182 → D175, the loan-terms row 73
#   → 74, the deferred-maintenance row 30 → 25.
# - **UNITS CHANGED, which is the half that silently corrupts.** v1.2
#   took operating expenses as **$/SF/year**; v1.3 takes them as
#   **$/UNIT/year** (`J143 = (...)*$C$126`, and C126 is the unit count).
#   Other income likewise moved to a per-unit basis. Writing the old
#   $/SF numbers into the new cells would not fail — it would produce a
#   workbook whose expenses are wrong by a factor of (SF per unit),
#   roughly 100x, with every formula still returning "OK". So the
#   converters below are not cosmetic: `_per_unit` is the correctness
#   boundary between the two versions.
#
# There is no v1.2 compatibility path. Two live maps means a workbook
# whose numbers depend on which file the operator happened to drop in,
# and nothing on screen would say which. `_assert_template_shape`
# refuses to write a workbook that does not look like v1.3.

# Unit mix rows: 112-125 (14 slots in v1.3; v1.2 had 21 at 111-131)
UNIT_MIX_START_ROW = 112
UNIT_MIX_END_ROW = 125

# OpEx row mapping: benchmark_key → row. Every one of these rows takes
# **dollars per unit per year** — see `_per_unit`.
#
# Rows deliberately NOT written: 148 (bank/merchant fees — written
# separately, since config states it as a % of EGR and the cell wants
# dollars per unit), 149 (Security — no benchmark of ours corresponds,
# so the template's own default stands), 150 (Management Fee — a % of
# EGR in both models, written by `_write_opex` directly).
OPEX_ROW_MAP = {
    "repairs":      143,
    "payroll":      144,
    "ga":           145,
    "advertising":  146,   # v1.3 labels it "Marketing"
    "utilities":    147,
    "insurance":    151,
    "property_tax": 152,
}
#: Capital reserves (row 157) sit in the CAPITAL EXPENDITURES block, not
#: in OpEx, and come from the benchmark band rather than from a CIM
#: expense line — so `_write_capex` owns the row and it is deliberately
#: absent above. Two writers aiming at one row is how a value gets
#: silently overwritten by whichever runs last.
_CAP_RESERVE_ROW = 157

# Cells whose LABEL identifies v1.3, and what each must say. Checked
# before the first write: a v1.2 workbook (or any other file) fails
# here, loudly, instead of scattering correct-looking numbers into
# wrong cells. Labels rather than addresses alone because an address is
# present in both versions — it is what it SAYS that differs.
_SHAPE_MARKERS = {
    "B143": "Repairs and Maintenance",
    "B150": "Management Fee",
    "B157": "Capital Reserves",
    "B175": "Sale Month",
    "B249": "Pref. Return",
}

# ── Structural constants ─────────────────────────────────────────────
# Not assumptions — the template's own schema. Named so the AST gate can
# be absolute ("no numeric literal on the value side of a write") and so
# a reviewer can tell a layout fact from an underwriting opinion.

_COL_IN_PLACE = 7        # G — the in-place column on every paired row
_COL_STABILIZED = 9      # I — the stabilized column
_COL_PROMOTE = 9         # I — promote share on the tier rows

# Unit-mix columns.
_COL_UNIT_LABEL = 2          # B
_COL_UNIT_COUNT = 3          # C
_COL_UNIT_SF = 4             # D
_COL_UNIT_STABILIZED_PCT = 5  # E
_COL_UNIT_CLIMATE = 13       # M

# Rounding precision, in decimal places. Not assumptions — they keep
# float noise out of a cell and nothing else. Rates carry more digits
# than dollars because 6 dp on a cap rate is a hundredth of a basis
# point, where 2 dp on a dollar is a cent.
_DOLLARS_DP = 2
_PERCENT_DP = 4
_RATE_DP = 6

# Growth ladder: rows 102-107, years 1-6 across columns C..H.
_GROWTH_FIRST_COL = 3                    # C
_GROWTH_YEARS = 6                        # C..H
_REVENUE_GROWTH_ROWS = (102, 103, 104)   # in-place rent, stabilized rent, other income
_EXPENSE_GROWTH_ROWS = (105, 106, 107)   # OpEx ex-taxes, property taxes, CapEx
# B108 in the workbook reads "Growth is assumed to begin month 13", so the
# year-1 column is the template's in-place year and carries no growth.
# That is the template's convention, not our assumption — the app expresses
# year 1 through `yr1_noi_bump`, which has no cell here.
_GROWTH_BEGINS_YEAR = 2
_NO_GROWTH = 0
# `rev_cagr_yr1_3` applies through year 3 and `rev_cagr_yr4_5` year 4
# onward — the banding `analysis.valuation.project_cash_flows` uses.
_REV_CAGR_BAND_END_YEAR = 3

_STABILIZATION_BEGIN_MONTH = 1     # K102 — lease-up starts immediately

# Costs incurred at closing: start month 0, spread over 0 months.
_AT_CLOSING_START_MONTH = 0
_AT_CLOSING_DURATION = 0

# Blank unit-mix row: {column: value}. "% stabilized" of 1 means fully
# stabilized.
_UNIT_STABILIZED = 1
_UNIT_NOT_STABILIZED = 0
_NO_UNITS = 0
_NO_RENT = 0
_BLANK_UNIT_ROW = {
    _COL_UNIT_LABEL: "[Unit Type]",
    _COL_UNIT_COUNT: _NO_UNITS,
    _COL_UNIT_SF: _NO_UNITS,
    _COL_UNIT_STABILIZED_PCT: _UNIT_STABILIZED,
    _COL_IN_PLACE: _NO_RENT,
    _COL_STABILIZED: _NO_RENT,
    _COL_UNIT_CLIMATE: "Non-Climate",
}
# Other income: rows 132-136 (Admin Fees / Late Fees / Insurance /
# Merchandise / Uhaul). Each takes dollars PER UNIT per billing period,
# not an annual total — `J132 = (...)*$C$126*(12/$E$128)`. We carry one
# undifferentiated other-income figure, so the whole amount lands on the
# first row and the rest are zeroed; splitting a number the CIM never
# split would be the writer inventing a breakdown.
_OTHER_INCOME_ROWS = (132, 133, 134, 135, 136)
_OTHER_INCOME_TARGET_ROW = 132
#: E128 — "Concessions Avg. Length", which the other-income formulas
#: also use as their annualization divisor (12/E128). A template quirk,
#: read from the workbook rather than assumed, so the dollars tie.
_OTHER_INCOME_PERIOD_CELL = "E128"
_NO_INCOME = 0

# Absence-of-data fallbacks. A CIM that reports no CapEx, no other
# income and no expense line for a category gets zero, not a guess.
_NO_CAPEX = 0
_NO_EXPENSE = 0

# We model ONE senior loan. Mezz/junior paper is out of scope (scoped
# backlog item D), so H65 stays flat zero rather than reading a term.
_NO_JUNIOR_DEBT = 0
_NO_DEBT = 0

# The current owner's reserve is whatever the CIM shows, and CIMs do not
# show one; the stabilized reserve below it is the underwriting number.
_NO_IN_PLACE_CAPITAL_RESERVE = 0

# Waterfall block (rows 242-252 in v1.3). We charge no GP acquisition or
# disposition fee, so the promote structure is pref + promote only. (The
# disposition COST — the broker — is K175, a different cell; writing it
# here would double-count the sale. Recorded in `_write_reversion`.)
_NO_GP_FEE = 0
#: 2nd/3rd/4th tier promote shares (column I).
_PROMOTE_TIER_ROWS = (250, 251, 252)
#: Hurdle rates. **v1.3 does NOT chain them.** v1.2's H259-261 each read
#: `=+H258`, so writing the pref once set all four. Here H250 and H251
#: are their own literals (0.08 as shipped) and only H252 reads `=H251`,
#: so the pref must be written into all three or the workbook runs a
#: multi-hurdle structure `model.waterfall` does not implement.
_HURDLE_ROWS = (249, 250, 251)
_COL_HURDLE = 8          # H
_YES = "Yes"
# The only non-EGR option the H245 formula understands. See the module
# docstring: the label is honest about the workbook, and the workbook
# disagrees with `config.AM_FEE_BASE`.
_AM_FEE_BASIS_LABEL = "% of LP Equity"

_SUMMARY_NOTE_COL = 6              # F
_SUMMARY_STRENGTH_ROWS = range(6, 11)
_SUMMARY_WEAKNESS_ROWS = range(12, 17)

# Divergence disclosures — column B, starting just below the
# partnership block (which ends at the net-cash-flow total, row 263).
#
# **These addresses are a PREFERENCE, not an assumption.** The template
# is proprietary, gitignored and absent from CI and from the machine
# that chose them, so "B263 is blank" was a guess — and a guess is not
# something a deliverable may depend on. Two ways it goes wrong: a cell
# inside a merged label band raises on write (openpyxl MergedCells are
# read-only), killing the run before `wb.save` so NO workbook is
# produced at all; and an occupied cell would be silently overwritten.
#
# So the writer VERIFIES rather than trusts: `_free_disclosure_cells`
# takes the first free, unmerged, empty cells at or below the preferred
# row, and if it cannot find enough it warns and writes nothing rather
# than corrupting a workbook or crashing a run. The parity test still
# asserts the preferred cells are the ideal ones, so drift is reported —
# but the workbook no longer BREAKS when they are not.
#
# v1.3 note: B263 — the guess these constants were first written
# against — is OCCUPIED in v1.3 (it is the "Total" label of the GP-LP
# net cash-flow block). The verify-don't-trust design above is exactly
# why that cost nothing: the writer would have skipped down. The
# preferred row is corrected to 265 anyway, so the common case needs no
# searching.
_DISCLOSURE_COL = 2                    # B
_DISCLOSURE_FIRST_ROW = 265
#: How far below the preferred row to look before giving up. Bounded so
#: a disclosure can never wander into an unrelated part of the sheet.
_DISCLOSURE_SEARCH_ROWS = 12
_DISCLOSURE_PREF_CELL = "B265"
_DISCLOSURE_AM_FEE_CELL = "B266"

# The strings are static on purpose: interpolating the deal's own rates
# would put values in prose where no test reconciles them. Direction and
# mechanism are what the reader needs; the rates live in their cells.
_PREF_DISCLOSURE = (
    "Note: this workbook's pref (H248) is an IRR hurdle; the app's "
    "waterfall accrues the pref on unreturned capital. Same rate (H249), "
    "different construction, so promote dollars differ from the memo. "
    "Not reachable from an input cell; disclosed here instead."
)
_AM_FEE_DISCLOSURE = (
    "Note: H245 charges the AM fee on LP equity (K61); the app charges "
    "invested equity (GP + LP), so this workbook's fee runs light by "
    "roughly the GP co-invest share. G245 is the fund's true rate, not "
    "a grossed-up one (operator decision 2026-08-10)."
)

# Benchmark bands are (low, high) tuples. `_BAND_HIGH` went with the
# management-fee cell when it started reading `resolve_mgmt_fee_target`
# instead of the band's top end.
_BAND_LOW = 0


def generate_template(
    cim_data,
    financial_analysis: dict,
    scenario_results: dict = None,
    max_offer: dict = None,
    output_dir: str = ".",
    property_name: str = "",
    hold_years: int = None,
    transaction_costs: dict = None,
    capex: float = None,
    debt_terms=None,
    waterfall_terms=None,
    am_fee_pct: float = None,
    sources_uses: dict = None,
    mgmt_fee_target_pct: float = None,
) -> str:
    """
    Copy the XLSM template and populate input cells with CIM data.

    Args:
        cim_data: CIMData dataclass with extracted property data
        financial_analysis: dict from analyze_financials()
        scenario_results: dict with bear/base/bull scenario results
        max_offer: dict with max price solver results
        output_dir: directory to write the output file
        property_name: display name for the property
        hold_years: hold period; drives the template's sale month (D175)
        transaction_costs: override of config.TRANSACTION_COSTS; the
            disposition percentage drives the template's selling costs
            (K175), which the template ships at 5%
        capex: CapEx already resolved to DOLLARS (item H). None falls
            back to reading `cim_data.capex_estimate` as dollars, which
            is right for every caller that has no basis selector — but
            wrong for a rate, so the engine passes the resolved figure
        debt_terms: an already-RESOLVED `model.debt.DebtTerms`, not an
            override dict. None resolves `config.DEBT_TERMS`.
        waterfall_terms: an already-RESOLVED
            `model.waterfall.WaterfallTerms`, not an override dict. None
            resolves `config.WATERFALL_TERMS`.

            Both take resolved objects deliberately. Re-resolving an
            override dict here would need the deal's capital structure to
            get `gp_coinvest_pct` right, and getting it wrong is silent:
            a deal edited to 25% co-invest would print a Sources & Uses
            stack split 25/75 in the .xlsx and a 10/90 waterfall in the
            .xlsm. `resolve_waterfall_terms` argues the same point at
            greater length. The engine resolves once and hands the result
            to every writer.
        am_fee_pct: resolved annual management fee rate. None uses
            `config.AM_FEE_PCT`.
        sources_uses: the run's Sources & Uses block. Supplies the loan's
            share of total uses for the template's LTC cell (H65); without
            it the workbook is written all-equity.

    Returns:
        Path to the generated .xlsm file
    """
    from analysis.valuation import (resolve_hold_years,
                                    resolve_transaction_costs)
    from model.debt import resolve_debt_terms
    from model.waterfall import resolve_waterfall_terms

    hold_years = resolve_hold_years(hold_years)
    costs = resolve_transaction_costs(transaction_costs)
    # The CLI has no per-deal overrides, so config defaults are the whole
    # resolved set there. The web path always passes resolved objects.
    debt = debt_terms if debt_terms is not None else resolve_debt_terms()
    waterfall = (waterfall_terms if waterfall_terms is not None
                 else resolve_waterfall_terms())
    am_fee = cfg.AM_FEE_PCT if am_fee_pct is None else am_fee_pct
    params = _base_scenario_params(scenario_results)

    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")

    # Build output filename
    safe_name = safe_filename(property_name or cim_data.property_name or "Deal")
    out_path = os.path.join(output_dir, f"UW_{safe_name}.xlsm")

    # Copy template
    shutil.copy2(TEMPLATE_PATH, out_path)

    # Open with VBA preservation
    wb = openpyxl.load_workbook(out_path, keep_vba=True)
    ws = wb["Underwriting"]
    ws_summary = wb["Summary"]

    _assert_template_shape(ws)

    # Populate sections
    _write_property_description(ws, cim_data, hold_years)
    _write_investment_cf(ws, cim_data, costs, capex)
    _write_financing(ws, debt, sources_uses)
    _write_growth_rates(ws, params)
    _write_stabilization(ws, cim_data, params)
    _write_unit_mix(ws, cim_data, params)
    _write_other_income(ws, cim_data)
    _write_vacancy(ws, cim_data, params)
    _write_opex(ws, cim_data, financial_analysis, mgmt_fee_target_pct)
    _write_capex(ws, cim_data)
    _write_reversion(ws, cim_data, financial_analysis, costs, scenario_results)
    _write_waterfall(ws, waterfall, am_fee)
    _write_divergence_disclosures(ws)
    _write_summary_notes(ws_summary, cim_data)

    wb.save(out_path)
    wb.close()

    logger.info("  Template: %s", out_path)
    return out_path


# ── Template shape ───────────────────────────────────────────────────

class TemplateShapeError(RuntimeError):
    """The workbook is not the version this writer's cell map describes."""


def _assert_template_shape(ws):
    """Refuse to write into a workbook whose labels are not v1.3's.

    **This is the guard the v1.2→v1.3 upgrade proved necessary.** Every
    address in the old map still EXISTS in v1.3 — they simply mean
    different things, and several changed units. A misaimed write is
    therefore not a crash; it is a plausible workbook with expenses off
    by two orders of magnitude and every internal error-check still
    reading "OK". Nothing downstream can catch that, so it is caught
    here, before the first write.

    Cheap on purpose: five label reads, no version-string parsing. The
    "Version" sheet records a number a user can edit; the labels are
    what the formulas are actually built around.
    """
    wrong = {address: ws[address].value
             for address, expected in _SHAPE_MARKERS.items()
             if ws[address].value != expected}
    if wrong:
        raise TemplateShapeError(
            "Template does not match the v1.3 cell map — refusing to "
            f"write. Expected {_SHAPE_MARKERS}, found {wrong}. If the "
            "template was upgraded, output/template_writer.py must be "
            "recalibrated against it (see the module docstring); if it "
            "was downgraded to v1.2, restore v1.3."
        )


# ── Resolved-assumption helpers ──────────────────────────────────────

def _total_units(cim_data) -> int:
    """Unit count from the CIM's own mix — the denominator v1.3 needs.

    v1.3 states operating expenses and other income per UNIT, so a deal
    with no unit mix has no basis on which to write either block. The
    callers treat 0 as "leave the template's own defaults alone and say
    so", never as a divisor.
    """
    from_mix = sum((unit.count or _NO_UNITS)
                   for unit in (cim_data.unit_mix or []))
    # The mix is the better source — it is what the workbook's own C126
    # will total once the rows are written, so per-unit figures written
    # off it are consistent with the denominator the formulas use.
    # `total_units` is the fallback for a CIM that stated a count but no
    # breakdown; a workbook written from it carries expenses on the
    # right basis even though its rent rows stay blank.
    return from_mix or (cim_data.total_units or _NO_UNITS)


def _per_unit(total_dollars: float, units: int) -> float:
    """Annual dollars → dollars per unit per year, v1.3's expense basis.

    The one-line function that separates a correct workbook from one
    whose expenses are ~100x too large: v1.2 took these cells as $/SF.
    """
    return round(total_dollars / units, _DOLLARS_DP)


def _base_scenario_params(scenario_results: dict = None) -> dict:
    """The base case's resolved scenario parameters.

    Prefers the RUN's own params over re-reading config: `scenario_results`
    carries the values the projection actually used, and re-resolving
    would answer with whatever config says now. Falls back to config only
    when no scenario ran — the CLI's degraded path and tests.
    """
    base = (scenario_results or {}).get(ScenarioType.BASE) or {}
    return base.get("params") or cfg.SCENARIO_DEFAULTS[ScenarioType.BASE]


def _physical_occupancy(cim_data) -> float:
    """Physical occupancy, or the config assumption with a warning.

    The audit's complaint was not the 0.90 itself but that it was silent:
    a deal whose CIM never states occupancy was underwritten at 90% and
    nothing anywhere said so.
    """
    occ = cim_data.physical_occupancy
    # `is not None`, not truthiness: `physical_occupancy` defaults to None
    # when the parser finds nothing, so None is genuinely "missing" — but
    # a CIM that STATES 0% is a fully vacant building, and `or` would
    # overwrite that fact with 90% while logging that it was absent. A
    # warning that misreports its own trigger is worse than the silence
    # this helper exists to end.
    if occ is not None:
        return float(occ)
    assumed = cfg.XLSM_TEMPLATE_INPUTS["assumed_physical_occupancy"]
    logger.warning(
        "Physical occupancy missing from CIM data — XLSM written against "
        "the assumed %.0f%% in config.XLSM_TEMPLATE_INPUTS. Verify before "
        "relying on the workbook's vacancy or stabilization timing.",
        assumed * 100)
    return float(assumed)


def _next_month_start(today: datetime) -> datetime:
    """First of the month after `today` — the template's analysis begin
    date. Calendar arithmetic, not an assumption, which is why the month
    and day components live here and not in a write path."""
    if today.month == 12:
        return datetime(today.year + 1, 1, 1)
    return datetime(today.year, today.month + 1, 1)


def _unit_label(unit, index: int) -> str:
    """The unit type's own label, or a 1-based positional stand-in."""
    return unit.size_label or f"Type {index + 1}"


def _vacancy_from(occupancy: float) -> float:
    """Vacancy is the complement of occupancy — a definition, not an
    assumption, which is why the arithmetic lives here and not in a
    write path."""
    return max(0.0, 1.0 - float(occupancy))


def _is_stabilized(occupancy: float, params: dict) -> bool:
    """Stabilized against the BASE case's stabilized-occupancy target.

    `config.GATES["stabilized_occupancy"]` (0.85) is a different question
    — the ramp test for post-2020 vintages — and reconciling the two is
    item T's, per scoped-backlog rule 3.
    """
    return float(occupancy) >= float(params["stabilized_occ"])


# ── Property Description (rows 7-18) ─────────────────────────────────

def _write_property_description(ws, cim_data, hold_years: int):
    """Fill property description section.

    v1.3 moved the right-hand column: K8 is County (v1.2 repeated the
    name there), K9 is State, K10 is Zip. State is the one of the three
    the parser extracts, so it is now written rather than left blank.
    """
    ws["F8"] = cim_data.property_name or ""
    ws["K8"] = ""  # County — not extracted, user fills
    ws["F9"] = cim_data.address or ""
    ws["K9"] = cim_data.state or ""
    ws["F10"] = cim_data.city or ""
    ws["K10"] = ""  # Zip — CIMData carries no ZIP field, user fills

    if cim_data.acreage:
        ws["F11"] = cim_data.acreage

    # Buildings / stories — not reliably extracted, leave blank
    ws["K11"] = ""
    ws["K12"] = ""

    if cim_data.year_built:
        ws["F16"] = cim_data.year_built

    # Analysis begin date: first of next month
    ws["F17"] = _next_month_start(datetime.now())

    # Sale month — the hold period as the template expresses it.
    ws["D175"] = hold_years * MONTHS_PER_YEAR


# ── Investment Cash Flows (rows 20-47) ───────────────────────────────

def _write_investment_cf(ws, cim_data, costs: dict, capex: float = None):
    """Fill purchase price, acquisition closing costs and capex.

    Row 24 sits inside the template's own ACQUISITION COST block
    (K27 = SUM(K23:K26), which rolls into total project cost), so
    closing costs land where the template already totals them. Without
    this the .xlsm computed its purchase-side outlay as price + capex
    only, and its IRR disagreed with the memo and the .xlsx — which
    report a cost-inclusive basis — on every deal (review finding).
    v1.3 ships the row already labelled "Closing Costs" and the
    deferred-maintenance row already labelled, so the writer no longer
    supplies either label.

    No double count with the Architecture/Const.-Mgmt soft-cost rows:
    those are formulas off HARD costs (K34), a different base.
    """
    timing = cfg.XLSM_TEMPLATE_INPUTS
    if cim_data.asking_price:
        ws["K23"] = cim_data.asking_price
        acquisition_cost = (cim_data.asking_price
                            * costs["acquisition_closing_pct"])
        ws["K24"] = round(acquisition_cost, _DOLLARS_DP)
        ws["E24"] = _AT_CLOSING_START_MONTH   # same timing as the price row
        ws["F24"] = _AT_CLOSING_DURATION

    # `capex` arrives resolved when the caller knows the basis; a bare
    # cim_data read would write 0.02 into a dollar cell for a deal whose
    # CapEx was entered as a percentage of price.
    capex = ((cim_data.capex_estimate or _NO_CAPEX) if capex is None
             else capex)
    if capex > 0:
        ws["K25"] = capex
        ws["E25"] = timing["capex_start_month"]
        ws["F25"] = timing["capex_duration_months"]


# ── Financing (rows 64-73) ────────────────────────────────────────────

def _write_financing(ws, terms, sources_uses: dict = None):
    """Write the run's own loan into the template's debt block.

    **Why this is no longer all-equity.** The scoped backlog wrote its
    rule when leverage was opt-in per deal ("LTC stays 0, but the terms
    cells still carry the resolved values, so a user who flips LTC in
    Excel gets the terms the app would have used"). Item E3a reversed
    that on 2026-08-01: every deal is now sized at `config.DEBT_TERMS`
    and carries an LP net IRR, so the opted-out state that rule
    described no longer exists. Shipping an all-equity workbook beside a
    levered memo would leave exactly the contradiction this item exists
    to close, so H65 carries the run's actual leverage.

    `sources_uses["ltv"]` is loan / TOTAL USES — despite the key name,
    the denominator is the full cost stack, which is precisely what the
    template's H65 means (K65 = H65 * $K$56, and K56 is Total Uses). The
    workbook re-derives the dollar loan off its OWN K56, so the two agree
    on the loan only insofar as the two cost stacks agree; the K24
    closing-cost row above is what keeps them close.

    With no Sources & Uses (the CLI, or a deal that never sized) the
    block stays all-equity and the terms cells still carry the resolved
    loan, which is the backlog's original "flips LTC in Excel" case.

    v1.3 addresses: LTC H64→H65, junior H65→H66, loan terms row 73→74.
    """
    ltc = (sources_uses or {}).get("ltv")
    ws["H65"] = round(float(ltc), _RATE_DP) if ltc else _NO_DEBT
    ws["H66"] = _NO_JUNIOR_DEBT
    ws["F74"] = terms.term_years * MONTHS_PER_YEAR
    ws["G74"] = terms.io_months
    ws["H74"] = terms.amort_years * MONTHS_PER_YEAR
    ws["I74"] = terms.all_in_rate()


# ── Growth Rates (rows 100-106) ──────────────────────────────────────

def _write_growth_rates(ws, params: dict):
    """Write the base case's growth banding across years 1-6.

    Deliberate behavior change (scoped-backlog rule 4): this was a flat
    0%-then-3% ladder on all six rows regardless of scenario. Revenue
    rows now grow at the resolved revenue CAGR and expense rows at
    `exp_growth`, so the workbook and the projection answer the same
    question the same way.

    Year 1 stays flat — see `_GROWTH_BEGINS_YEAR`. The app expresses
    year 1 through `yr1_noi_bump`, which the template has no cell for.
    """
    rev_early = params["rev_cagr_yr1_3"]
    rev_late = params["rev_cagr_yr4_5"]
    exp_growth = params["exp_growth"]

    for row in _REVENUE_GROWTH_ROWS + _EXPENSE_GROWTH_ROWS:
        is_revenue = row in _REVENUE_GROWTH_ROWS
        for year in range(1, _GROWTH_YEARS + 1):
            if year < _GROWTH_BEGINS_YEAR:
                value = _NO_GROWTH
            elif not is_revenue:
                value = exp_growth
            elif year <= _REV_CAGR_BAND_END_YEAR:
                value = rev_early
            else:
                value = rev_late
            column = _GROWTH_FIRST_COL + year - 1
            ws.cell(row=row, column=column).value = value


# ── Stabilization ────────────────────────────────────────────────────

def _write_stabilization(ws, cim_data, params: dict):
    """Set stabilization timing."""
    occ = _physical_occupancy(cim_data)

    ws["K102"] = _STABILIZATION_BEGIN_MONTH
    if _is_stabilized(occ, params):
        # Already stabilized — complete in the month it begins.
        ws["K103"] = _STABILIZATION_BEGIN_MONTH
    else:
        ws["K103"] = cfg.VALUE_ADD_SCENARIOS[
            ScenarioType.BASE]["months_to_stabilize"]


# ── Unit Mix (rows 112-125) ──────────────────────────────────────────

def _write_unit_mix(ws, cim_data, params: dict):
    """Populate unit mix rows from CIMData.unit_mix."""
    units = cim_data.unit_mix or []

    # Clear all unit mix rows first (rows 112-125)
    for row in range(UNIT_MIX_START_ROW, UNIT_MIX_END_ROW + 1):
        for column, blank in _BLANK_UNIT_ROW.items():
            ws.cell(row=row, column=column).value = blank

    # Fill with actual data
    stabilized = _is_stabilized(_physical_occupancy(cim_data), params)

    for i, unit in enumerate(units):
        if i >= (UNIT_MIX_END_ROW - UNIT_MIX_START_ROW + 1):
            logger.warning("Unit mix has %d types but template has %d slots",
                           len(units), UNIT_MIX_END_ROW - UNIT_MIX_START_ROW + 1)
            break

        row = UNIT_MIX_START_ROW + i
        label = _unit_label(unit, i)
        rate = unit.rate or _BLANK_UNIT_ROW[_COL_IN_PLACE]

        ws.cell(row=row, column=_COL_UNIT_LABEL).value = label
        ws.cell(row=row, column=_COL_UNIT_COUNT).value = (
            unit.count or _BLANK_UNIT_ROW[_COL_UNIT_COUNT])
        ws.cell(row=row, column=_COL_UNIT_SF).value = (
            unit.sf or _BLANK_UNIT_ROW[_COL_UNIT_SF])
        ws.cell(row=row, column=_COL_UNIT_STABILIZED_PCT).value = (
            _UNIT_STABILIZED if stabilized else _UNIT_NOT_STABILIZED)
        ws.cell(row=row, column=_COL_IN_PLACE).value = rate
        # I column: stabilized rent — use in-place for now, user adjusts
        # Note: I{row} has formula =+G{row} in template for rows that had
        # data. For overwritten rows we set explicitly.
        ws.cell(row=row, column=_COL_STABILIZED).value = rate

        # Climate type
        if unit.climate_controlled:
            ws.cell(row=row, column=_COL_UNIT_CLIMATE).value = "Climate"
        else:
            ws.cell(row=row, column=_COL_UNIT_CLIMATE).value = "Non-Climate"


# ── Other Income (rows 132-136) ──────────────────────────────────────

def _write_other_income(ws, cim_data):
    """Populate other income lines — v1.3's PER-UNIT basis.

    v1.2 took an annual dollar total here. v1.3 does not:

        J132 = ((1-$E$126)*G132 + $E$126*I132) * $C$126 * (12/$E$128)

    so the cell holds dollars per unit per billing period, and the
    period length is the workbook's own E128. Writing our annual total
    into it would overstate other income by (units x 12/E128) — a
    four-figure error dressed as a plausible number, with no formula
    reporting anything wrong. So the total is divided back down, and
    E128 is READ from the workbook rather than assumed, which is what
    makes the resulting J132 equal the figure the memo prints.

    With no unit count there is no per-unit basis at all; the rows are
    zeroed (the CIM's other income is not silently dropped into a cell
    that would misstate it) and the warning says the block needs a
    hand.
    """
    other = cim_data.other_income or _NO_INCOME

    # Clear defaults — the template ships five populated example lines.
    for row in _OTHER_INCOME_ROWS:
        ws.cell(row=row, column=_COL_IN_PLACE).value = _NO_INCOME
        ws.cell(row=row, column=_COL_STABILIZED).value = _NO_INCOME

    if other <= _NO_INCOME:
        return

    units = _total_units(cim_data)
    if not units:
        logger.warning(
            "Other income of %s not written to the XLSM: v1.3 states it "
            "per unit and this CIM has no unit count. Enter it by hand "
            "on rows %s.", other, _OTHER_INCOME_ROWS)
        return

    period_months = ws[_OTHER_INCOME_PERIOD_CELL].value
    if not isinstance(period_months, (int, float)) or period_months <= _NO_INCOME:
        period_months = MONTHS_PER_YEAR
        logger.warning(
            "%s (the other-income annualization divisor) is not a number "
            "in this template — falling back to a 12-month period.",
            _OTHER_INCOME_PERIOD_CELL)

    periods_per_year = MONTHS_PER_YEAR / period_months
    per_unit = round(other / (units * periods_per_year), _DOLLARS_DP)
    ws.cell(row=_OTHER_INCOME_TARGET_ROW,
            column=_COL_IN_PLACE).value = per_unit
    ws.cell(row=_OTHER_INCOME_TARGET_ROW,
            column=_COL_STABILIZED).value = per_unit


# ── Vacancy (rows 139-140) ───────────────────────────────────────────

def _write_vacancy(ws, cim_data, params: dict):
    """Set vacancy and credit loss assumptions.

    Stabilized vacancy is the complement of the base case's
    `stabilized_occ`, not a standing 10%. At the 0.88 default that is
    0.12 — which is also the template's own shipped default, so the
    literal this replaced was the one value in the block that agreed with
    neither the workbook nor the model.
    """
    inputs = cfg.XLSM_TEMPLATE_INPUTS

    ws["G139"] = round(_vacancy_from(_physical_occupancy(cim_data)),
                       _PERCENT_DP)
    ws["I139"] = round(_vacancy_from(params["stabilized_occ"]), _PERCENT_DP)

    ws["G140"] = inputs["credit_loss_in_place"]
    ws["I140"] = inputs["credit_loss_stabilized"]


# ── Operating Expenses (rows 150-159, 164) ───────────────────────────

def _write_opex(ws, cim_data, financial_analysis: dict,
                mgmt_fee_target_pct=None):
    """
    Populate OpEx from CIM data and analyst adjustments.

    **v1.3's basis is $/UNIT/year, not $/SF/year.** Every row here
    computes `(...) * $C$126` — the unit count — where v1.2 multiplied
    by square footage. The conversion is `_per_unit`, and it is the
    reason this function no longer touches `cim_data.nrsf` at all.

    In-Place column (G): CIM actual $/unit/year
    Stabilized column (I): analyst-adjusted $/unit/year

    A CIM with no unit count cannot be expressed on this basis, so the
    block is left at the template's own defaults with a warning. That
    is deliberately NOT a zero-fill: zeroing would print a workbook
    claiming the property has no operating expenses, which is a worse
    lie than an untouched example column the reader can see is generic.
    """
    units = _total_units(cim_data)
    expense_analysis = financial_analysis.get("expense_analysis", {})
    expense_lines = expense_analysis.get("lines", [])

    if not units:
        logger.warning(
            "No unit count on this CIM — the XLSM's operating-expense "
            "block (v1.3 states it per unit) keeps the template's own "
            "defaults and must be filled by hand.")
        return

    # Build lookup: benchmark_key → expense line
    exp_lookup = {}
    for line in expense_lines:
        key = line.get("benchmark_key")
        if key:
            exp_lookup[key] = line

    for benchmark_key, row in OPEX_ROW_MAP.items():
        line = exp_lookup.get(benchmark_key, {})
        cim_value = line.get("cim_value")
        adjusted_value = line.get("adjusted_value")

        # In-place: CIM actual as $/unit/year
        if cim_value is not None:
            in_place = _per_unit(cim_value, units)
        else:
            in_place = _NO_EXPENSE

        # Stabilized: analyst-adjusted as $/unit/year
        if adjusted_value is not None:
            stabilized = _per_unit(adjusted_value, units)
        else:
            stabilized = in_place

        ws.cell(row=row, column=_COL_IN_PLACE).value = in_place
        ws.cell(row=row, column=_COL_STABILIZED).value = stabilized

    # Management fee — % of EGR (row 150). The CIM's own rate when it
    # states one; otherwise the resolved pro-forma target.
    #
    # This used to read the top of the benchmark band directly. Same
    # NUMBER as the target's default, but arriving by a second route, so
    # a per-deal `mgmt_fee_target_pct` moved the memo and the .xlsx while
    # this workbook — a real deliverable — kept writing 6%. Two output
    # files asserting different management fees on the same deal, with
    # nothing on screen saying so. It reads the ONE resolver now.
    #
    # `is not None` for the same reason as `_physical_occupancy`: a
    # self-managed property stating a 0% fee is data, not a blank.
    mgmt_pct = (cim_data.mgmt_fee_pct if cim_data.mgmt_fee_pct is not None
                else resolve_mgmt_fee_target(mgmt_fee_target_pct))
    ws["G150"] = mgmt_pct
    ws["I150"] = mgmt_pct

    # Bank/merchant fees (row 148). Config states this as a % of EGR;
    # v1.3's cell wants dollars per unit, so the rate is applied to the
    # run's own EGR and divided down. Same assumption, the template's
    # unit — which is the whole reason config keeps a percentage: a
    # per-unit dollar default would be a number nobody could check
    # against the memo.
    #
    # With no EGR (a CIM thin enough that `analyze_financials` could not
    # derive one) the row keeps the template's default rather than
    # taking a percentage as though it were dollars.
    egr = (financial_analysis.get("income_analysis") or {}).get("egr")
    if egr:
        inputs = cfg.XLSM_TEMPLATE_INPUTS
        ws["G148"] = _per_unit(inputs["bank_fee_pct_in_place"] * egr, units)
        ws["I148"] = _per_unit(inputs["bank_fee_pct_stabilized"] * egr, units)
    else:
        logger.warning(
            "No EGR on this run — the XLSM's bank/merchant fee row keeps "
            "the template's default (v1.3 states it per unit, and the "
            "config assumption is a percentage of EGR).")


# ── Capital Expenditures (row 157) ───────────────────────────────────

def _write_capex(ws, cim_data):
    """Set capital reserve assumption.

    In-place is zero because the CIM reports no reserve line; stabilized
    is the bottom of the benchmark band, which is the underwriting
    number.

    The band is stated in $/NRSF/year and v1.3's row is $/unit/year, so
    the benchmark passes through NRSF to dollars and back down to units.
    Without both figures the row keeps the template's default.
    """
    units = _total_units(cim_data)
    nrsf = cim_data.nrsf
    if not units or not nrsf:
        logger.warning(
            "Capital reserve not written to the XLSM: v1.3 states it per "
            "unit and this CIM lacks a unit count or NRSF.")
        return

    band_psf = cfg.EXPENSE_BENCHMARKS["cap_reserve"][_BAND_LOW]
    ws.cell(row=_CAP_RESERVE_ROW,
            column=_COL_IN_PLACE).value = _NO_IN_PLACE_CAPITAL_RESERVE
    ws.cell(row=_CAP_RESERVE_ROW,
            column=_COL_STABILIZED).value = _per_unit(band_psf * nrsf, units)


# ── Reversion / Sale Assumptions ─────────────────────────────────────

def _write_reversion(ws, cim_data, financial_analysis: dict, costs: dict,
                     scenario_results: dict = None):
    """Set cap rate and sale assumptions.

    K173 ("Market Cap Rate Today") and K174 ("Cap Rate at Sale") come from
    the run's resolved market anchor and base-case exit cap, so the .xlsm
    prints the same two rates as the memo and the .xlsx. They used to be
    the ENTRY cap and the workbook's own `= K173 + 0.5%`, which made this
    a second underwriting model that disagreed with the Python one on
    every deal.

    Overwriting K174 replaces a formula with a value, deliberately: the
    terminal cap is no longer "entry + 50 bp", so a cell that keeps
    tracking K173 by that rule would drift away from the published exit
    the moment anyone edited K173. Nothing depends on K174 REMAINING a
    formula — the interpolated cap at stabilization reads its VALUE, and
    still interpolates correctly between an anchor and a wider terminal
    cap, which is the drift model.

    **The thin-data path writes nothing rather than inventing an anchor.**
    It used to fall back to a 6.5% literal, and on a deal with no NOI and
    no price that 6.5% was the template deciding its own cap rate. Now
    K173 is written only from a resolved market anchor or a computed
    entry cap; with neither, both cells keep the workbook's own defaults
    — including K174's entry+50bp formula. That formula is the pre-#31
    contradiction, so it survives ONLY where there is no resolved exit
    cap for it to contradict, and a deal that reaches here produced no
    scenarios and no returns at all. The warning says so.
    """
    noi = financial_analysis.get("adjusted_ttm_noi", {}).get("analyst_adjusted_noi")
    price = cim_data.asking_price

    base = (scenario_results or {}).get(ScenarioType.BASE) or {}
    detail = base.get("exit_cap_detail") or {}
    market_cap = detail.get("market_cap")
    exit_cap = base.get("exit_cap")
    if market_cap is not None and exit_cap is not None:
        ws["K173"] = round(float(market_cap), _RATE_DP)
        ws["K174"] = round(float(exit_cap), _RATE_DP)
    elif noi and price and price > 0:
        # No scenario ran, so there is no resolved anchor. Write the entry
        # cap and leave K174's own formula alone rather than inventing a
        # terminal cap off a number that is not a market cap.
        ws["K173"] = round(noi / price, _PERCENT_DP)
    else:
        logger.warning(
            "No resolved market cap and no entry cap (NOI=%r, price=%r) — "
            "leaving the XLSM's own cap-rate defaults in place, including "
            "its entry+50bp terminal formula, which the model does not use.",
            noi, price)
    #
    # K175 ("Selling Costs") is the template's cost of sale — the cell
    # our disposition assumption belongs in. (The scoped backlog pointed
    # at the waterfall block's disposition FEE instead — F245 in v1.3 —
    # which is part of the promote structure and correctly left at 0
    # because we model no GP fees. Writing the broker cost there would
    # leave the template's own 5% still charging and double-count the
    # sale.)
    ws["K175"] = round(costs["disposition_cost_pct"], _PERCENT_DP)


# ── Distribution Waterfall (rows 242-252) ────────────────────────────

def _write_waterfall(ws, terms, am_fee_pct: float):
    """Write the fund's resolved partnership terms.

    Replaces a GP_EQUITY_SHARE / GP_AM_FEE_RATE / GP_PROMOTE_PCT
    environment-variable block that defaulted to 6% GP equity against a
    `GP_COINVEST_PCT` of 10%. The env vars are deleted, not re-defaulted
    (scoped-backlog rule 2): an assumption nobody can find in config is
    not an auditable assumption.

    **Tier mapping, and the way v1.3 changed it.** The template has four
    hurdle rows, which is not four hurdles — we collapse them to the
    single hurdle `model.waterfall` implements. In v1.2 that took ONE
    write: H259/H260/H261 each read `=+H258`. **v1.3 broke the chain.**
    H250 and H251 ship as their own 0.08 literals and only H252 reads
    `=H251`, so writing the pref into H249 alone would leave tiers 2-3
    on the template's shipped 8% and tier 4 following them — a
    four-hurdle structure on any deal whose pref is not 8%, quietly
    disagreeing with the memo. The rate is therefore written into every
    unchained hurdle row (`_HURDLE_ROWS`).

    Promote: v1.3 ships tier 4 at 30% against tiers 2-3 at 20%, so the
    single `promote_split` must be written to all three or the workbook
    runs a promote ladder we do not model. J250 =
    `I250+(1-I250)*$J$244` — promote plus the GP's pari-passu share of
    the residual — is the same construction as ours, promote computed on
    the LP-attributable residual only.

    What the mapping does NOT fix is that H248 is labelled "IRR Hurdle"
    while our pref is an accrual account. Same rate, different math; see
    the module docstring.
    """
    ws["H60"] = terms.gp_coinvest_pct

    ws["C244"] = cfg.GP_ENTITY_NAME
    ws["C245"] = cfg.LP_ENTITY_NAME

    # GP fees — we charge neither.
    ws["F244"] = _NO_GP_FEE        # Acquisition fee
    ws["F245"] = _NO_GP_FEE        # Disposition fee

    # Asset management fee
    ws["G244"] = _AM_FEE_BASIS_LABEL
    ws["G245"] = am_fee_pct
    ws["I245"] = _YES              # Fees accrue

    # Include GP fees in analysis
    ws["I242"] = _YES

    # Preferred return. In v1.2 this overwrote a formula that made the
    # pref depend on whether the deal was levered — not a term in the
    # LPA. v1.3 ships plain literals instead, and they are overwritten
    # for the same reason: the rate is the fund's, not the template's.
    for row in _HURDLE_ROWS:
        ws.cell(row=row, column=_COL_HURDLE).value = terms.pref_rate

    # Promote tiers 2-4, all sitting at the single hurdle above.
    for row in _PROMOTE_TIER_ROWS:
        ws.cell(row=row, column=_COL_PROMOTE).value = terms.promote_split


def _covered_by_merge(ws, cell) -> bool:
    """True if `cell` sits in ANY merged range — anchor included.

    Both halves break a write. A non-anchor member is a `MergedCell`
    whose `.value` is read-only, so assigning raises and the run dies
    before `wb.save`. The anchor IS writable, but writing there paints
    text across a band that belongs to some label, which is the silent
    corruption the emptiness check alone cannot see.
    """
    return any(rng.min_row <= cell.row <= rng.max_row
               and rng.min_col <= cell.column <= rng.max_col
               for rng in ws.merged_cells.ranges)


def _free_disclosure_cells(ws, count: int) -> list:
    """The first `count` cells safe to write a disclosure into.

    Safe means: not inside a merged range, and empty — so nothing the
    template already says is destroyed. Returns fewer than `count` when
    the sheet has no room, which the caller treats as "say nothing"
    rather than "write anyway".
    """
    free = []
    for offset in range(_DISCLOSURE_SEARCH_ROWS):
        cell = ws.cell(row=_DISCLOSURE_FIRST_ROW + offset,
                       column=_DISCLOSURE_COL)
        if _covered_by_merge(ws, cell) or cell.value not in (None, ""):
            continue
        free.append(cell)
        if len(free) == count:
            break
    return free


def _write_divergence_disclosures(ws):
    """Stamp the two structural divergences INTO the workbook.

    Until 2026-08-10 they were recorded only in this module's docstring —
    the one place the analyst reading the workbook will never look. The
    module docstring carries the full reasoning; these two lines carry
    the direction and the mechanism to the reader who has only the file.
    Text only — the AST gate is numeric and untouched by design.

    The target cells are CHECKED, not assumed (see `_DISCLOSURE_COL`):
    this writer runs against a proprietary workbook nobody who chose the
    addresses can open, and a deliverable must not depend on that guess
    being right. Worst case here is a missing note and a loud warning;
    the alternative was a corrupted workbook or no workbook at all.
    """
    disclosures = (_PREF_DISCLOSURE, _AM_FEE_DISCLOSURE)
    cells = _free_disclosure_cells(ws, len(disclosures))
    if len(cells) < len(disclosures):
        logger.warning(
            "No free rows for the XLSM divergence disclosures near row %d "
            "(column B) — the workbook is written WITHOUT them. The pref "
            "and AM-fee divergences still apply; see the module docstring "
            "in output/template_writer.py, and move "
            "_DISCLOSURE_FIRST_ROW to a blank area of the template.",
            _DISCLOSURE_FIRST_ROW)
        return
    for cell, text in zip(cells, disclosures):
        cell.value = text


# ── Summary Sheet Notes ──────────────────────────────────────────────

def _write_summary_notes(ws, cim_data):
    """Clear deal-specific strengths/weaknesses for user to fill."""
    for row in _SUMMARY_STRENGTH_ROWS:
        ws.cell(row=row, column=_SUMMARY_NOTE_COL).value = ""
    for row in _SUMMARY_WEAKNESS_ROWS:
        ws.cell(row=row, column=_SUMMARY_NOTE_COL).value = ""

    # Label headers remain
    ws["F5"] = "STRENGTHS"
    ws["F11"] = "WEAKNESSES"


# ── Helpers ──────────────────────────────────────────────────────────
